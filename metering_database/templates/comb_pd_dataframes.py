import openpyxl as xl
import datetime
import pymysql
import json
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.types import String, SmallInteger,VARCHAR
import time
start = time.time()

wb = xl.load_workbook("C:\\Users\\KIMERA\\Desktop\\Instru UETCL0102.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[0]]
for i in range(1, ws.max_column):
    cell_value = ws.cell(row=1,column=i).value
    #if cell_value1 != '':
    if not cell_value == None:
        #cell_value = cell_value1.replace('-','')
        manu_1 = 'Elster'
        if cell_value.find(manu_1) == 1:
            print(f'manufacturer is Elster')
            break

meter_no_x = ws["B2"].value
meter_no = meter_no_x.split()[0].replace('-', '')
print(meter_no)

conn = pymysql.connect(host='localhost', port=3306, user='root', password='',
                                       database='meteringdatabase')
engine = create_engine('mysql://root:@localhost/meteringdatabase')
c = conn.cursor()
list_1 =[]
list_2 = []
list_3 = []

for x in range(1, 20):
    for r in range(1, ws.max_column+1):
        cell = ws.cell(r, x)
        global df
        if cell.value == 'Date' and ws.cell(r, x+1).value == 'Start' and ws.cell(r, x+2).value == 'End': #check this loop for x and r (row and column)
            for m in range(r+1,ws.max_row):
                use_date = ws.cell(m, x).value
                start_time = ws.cell(m, x+1).value
                end_time = ws.cell(m, x+2).value
                list_1.append(use_date)
                list_2.append(end_time)
                list_3.append(end_time)
            #print(list_1)
            df1 = pd.DataFrame(list_1)
            df1.columns = ['date']
            #print(df1)
            df2 = pd.DataFrame(list_2)
            df2.columns = ['start_time']
            #print(df2)
            df3 = pd.DataFrame(list_3)
            df3.columns = ['end_time']
            #print(df3)
            df = pd.concat([df1,df2,df3], axis=1)

            #print(df)

        list_ia = []
        #row_ia = 0
        global dfnew
        global df_ia
        if cell.value == 'A: PhA: Av':
            for m in range(r, ws.max_row):
                I_A = ws.cell(m, x - 1).value
                list_ia.append(I_A)
            df_ia = pd.DataFrame(list_ia)
            df_ia.columns = ['I_A']
            dfnew = pd.concat([df,df_ia],axis=1)
            #print(dfnew)

            break

        list_ib = []
        global df_ib
        if cell.value == 'A: PhB: Av':
            for m in range(r, ws.max_row):
                I_B = ws.cell(m, x - 1).value
                list_ib.append(I_B)
            df_ib_x = pd.DataFrame(list_ib)
            df_ib_x.columns = ['I_B']
            df_ib = pd.concat([dfnew, df_ib_x], axis=1)
            #print(df_ib)
            break

        list_ic = []
        global df_ic
        if cell.value == 'A: PhC: Av':
            for m in range(r, ws.max_row):
                I_C = ws.cell(m, x - 1).value
                list_ic.append(I_C)
            df_ic_x = pd.DataFrame(list_ic)
            df_ic_x.columns = ['I_C']
            df_ic = pd.concat([df_ib,df_ic_x], axis=1)
            print(df_ic)



            df_ic.to_sql('uetcl0103', con=engine, if_exists='append', index=False, dtype={'date': VARCHAR(length=30),
            'start_time': String(length=10),'end_time': String(length=10),'I_A': String(length=10),'I_B': String(length=10),
                                                                                          'I_C': String(length=10)                                                                             })
            break

end = time.time()
print("Elapsed time is  {}".format(end-start))
