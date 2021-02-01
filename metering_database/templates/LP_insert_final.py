import openpyxl as xl
#import datetime
import pymysql
import json
import xlsxwriter
import pandas as pd
from sqlalchemy import create_engine
from sqlalchemy.types import String, SmallInteger
import time
start = time.time()

workbook = xlsxwriter.Workbook('Example1.xlsx')
worksheet = workbook.add_worksheet()

wb = xl.load_workbook("C:\\Users\\KIMERA\\Desktop\\Instru UETCL0102.xlsx")
sheets = wb.sheetnames
ws = wb[sheets[0]]
for i in range(1, ws.max_column):
    cell_value = ws.cell(row=1,column=i).value
    #if cell_value1 != '':
    if not cell_value == None:
        #cell_value = cell_value1.replace('-','')
        manu_1 = 'Elster'
        if cell_value.find(manu_1) == 1:
            print(f'manufacturer is Elster')
            break

meter_no_x = ws["B2"].value
meter_no = meter_no_x.split()[0].replace('-', '')
print(meter_no)
#engine = create_engine('mysql://root:@localhost/meteringdatabase')
conn = pymysql.connect(host='localhost', port=3306, user='root', password='',
                                       database='meteringdatabase')
c = conn.cursor()
for x in range(1, 20):
    for r in range(1, ws.max_column):
        cell = ws.cell(r, x)
        list_0 = []
        list_01 = []
        list_02 = []
        if cell.value == 'Date' and ws.cell(r, x+1).value == 'Start' and ws.cell(r, x+2).value == 'End': #check this loop for x and r (row and column)
            rowx = 0
            columnx = 0
            row1 = 0
            row2 = 0
            for m in range(r,ws.max_row):
                use_date = ws.cell(m, x).value
                start_time = ws.cell(m, x+1).value
                end_time = ws.cell(m, x+2).value
                list_0.append(use_date)
                list_01.append(start_time)
                list_02.append(end_time)
            #print(list_01)
            #print(list_0)
            for nn in list_0:
                worksheet.write(rowx, columnx, nn)
                rowx += 1
            for np in list_01:
                worksheet.write(row1, 1, np)
                row1 += 1
            for pp in list_02:
                worksheet.write(row2, 2, pp)
                row2 += 1

        list_1 = []
        row_ia = 0
        if cell.value == 'A: PhA: Av':
            for m in range(r-1, ws.max_row):
                I_A = ws.cell(m, x - 1).value
                list_1.append(I_A)
            for ia in list_1:
                worksheet.write(row_ia, 3, ia)
                row_ia += 1
            #print(list_1)
            break
        list_2 = []
        row_ib = 0
        if cell.value == 'A: PhB: Av':
            for m in range(r - 1, ws.max_row):
                I_B = ws.cell(m, x - 1).value
                list_2.append(I_B)
            for ib in list_2:
                worksheet.write(row_ib, 4, ib)
                row_ib += 1
            #print(list_2)
            break

        list_3 = []
        row_ic = 0
        if cell.value == 'A: PhC: Av':
            for m in range(r - 1, ws.max_row):
                I_C = ws.cell(m, x - 1).value
                list_3.append(I_C)
            for ic in list_3:
                worksheet.write(row_ic, 5, ic)
                row_ic += 1
            #print(list_3)
            break

        list_4 = []
        row_va = 0
        if cell.value == 'V: PhA: Av': #voltages
            for m in range(r-1, ws.max_row):
                V_A = ws.cell(m, x - 1).value
                list_4.append(V_A)
            for va in list_4:
                worksheet.write(row_va, 6, va)
                row_va += 1
                #print(V_A)
            break
        list_5 = []
        row_vb = 0
        if cell.value == 'V: PhB: Av':
            for m in range(r-1, ws.max_row):
                V_B = ws.cell(m, x - 1).value
                list_5.append(V_B)
            for vb in list_5:
                worksheet.write(row_vb, 7, vb)
                row_vb += 1
                    # print(V_A)
            break
                #print(V_B)
        list_6 = []
        row_vc = 0
        if cell.value == 'V: PhC: Av':
            for m in range(r - 1, ws.max_row):
                V_C = ws.cell(m, x - 1).value
                # print(V_C)
                list_6.append(V_C)
            for vc in list_6:
                worksheet.write(row_vc, 8, vc)
                row_vc += 1
            break

        list_7 = []
        row_pw = 0
        if cell.value == 'kW: Sys: Av':
            for m in range(r-1, ws.max_row):
                PW = ws.cell(m, x - 1).value
                #print(PW)
                list_7.append(PW)
            for rpw in list_7:
                worksheet.write(row_pw, 9, rpw)
                row_pw += 1
            break

        list_8 = []
        row_rp = 0
        if cell.value == 'kvar: Sys: Av':
            for m in range(r-1, ws.max_row):
                P_var = ws.cell(m, x - 1).value
                list_8.append(P_var)
            for rp in list_8:
                worksheet.write(row_rp, 10, rp)
                row_rp += 1
            break
workbook.close()
df = pd.read_excel("Example1.xlsx", header=None, index_col=None,skiprows=1)
#df.head()
engine = create_engine('mysql://root:@localhost/meteringdatabase')
#df.to_sql('people', con=engine)
#print('success')
#df.rename(columns={'0':'date','1':'start_time','2':'end_time','3':'I_A','4':'I_B','5':'I_C','6':'V_A','7':'V_B','8':'V_C','9':'PW',
                   #'10':'P_var'}, inplace = True)
df.columns = ['date','start_time','end_time','I_A',
                     'I_B','I_C','V_A','V_B','V_C','PW','P_var']


#print(df)
df.to_sql('uetcl0102', con=engine, if_exists='append', index=False, dtype={'date': String(length=10),
'start_time': String(length=10),'end_time': String(length=10),'I_A': String(length=10),'I_B': String(length=10),'I_C': String(length=10),
'V_A': String(length=10),'V_B': String(length=10),'V_C': String(length=10),'PW': SmallInteger, 'P_var': String(length=10)})

end = time.time()
print("Elapsed time is  {}".format(end-start))










