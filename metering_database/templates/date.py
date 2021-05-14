from openpyxl import Workbook
import csv
import openpyxl as py
from pygrok import Grok
import pymysql
import datetime
from openpyxl import Workbook
import openpyxl as py
from datetime import datetime
import pandas as pd
import calendar
from sqlalchemy import create_engine
from sqlalchemy.types import String, SmallInteger,VARCHAR,FLOAT,Float,Numeric

wb = Workbook()
ws = wb.active
with open("C:\\Users\\KIMERA\Desktop\\mutundwe\\Historical UETCL0102.CSV", 'r') as f:
    for row in csv.reader(f):
        ws.append(row)
wb.save('C:\\Users\\KIMERA\Desktop\\mutundwe\\name.xlsx')
workbook = py.load_workbook('C:\\Users\\KIMERA\Desktop\\mutundwe\\name.xlsx')
sheet = workbook['Sheet']
man_value = sheet['B1'].value
manu = man_value.split()[0]  # meter manufacturer
# print(manu)
input_string = str(sheet["B2"].value)
#print(input_string)
# global meter
meter_no = input_string.split()[0].replace('-', '')  # meter number
#print(meter)
read_time = f'{input_string.split()[1]} {input_string.split()[2]}' #reading time

#global meter_no
# sql_1 = "SELECT CTratio, VTratio FROM meter_details WHERE meter_no=?",((meter,))
#c.execute("SELECT CTratios, VTratio FROM standalone_meter_details WHERE meter_no = %s", meter)
#data = c.fetchone()

date_pattern = '%{MONTHNUM:month}/%{MONTHDAY:day}/%{YEAR:year}'
grok = Grok(date_pattern)
xl = (grok.match(input_string))
mm = xl['month']
dd = xl['day']
yr = xl['year']
x = datetime(int(yr), int(mm), int(dd))
# ll = (x.strftime("%d-%m-%Y"))
D_O_R = (x.strftime("%d-%b-%Y"))  # reading date
list1 = []
list2 = []
list7 = []
for m in range(1, sheet.max_row+1):
    for n in range(1, sheet.max_column+1):
        #cell1 = sheet.cell(m,n)
        cell = sheet.cell(m,n).value
        cell_value = str(cell)
        if (cell_value.find('Historical data set') != -1):
            list1.append(cell_value)
            list2.append(m)
print(len(list1))
print(list2)
list3 = []
list4 = []
list5 = []
list6 = []
list7 = []
list8 = []
list9 = []
list10 = []
list11 = []
list12 = []
list13 = []
list14 = []
list15 = []
list16 = []
list17 = []
list18 = []
list19 = []
list20 = []
list21 = []
list22 = []
list23 = []
list24 = []
list25 = []
list26 = []
list27 = []
list28 = []
list29 = []
list30 = []
list31 = []
list32 = []
list1a = []
list33 = []
list34 = []
for i in range(1,len(list2)+1):
    if i<=len(list2)-1:
        row_1 = list2[i-1]
        row_2 = list2[i]
        #print(row_2)
        for r in range(row_1,row_2+1):
            for x in range(1,sheet.max_column+1):
                cell = sheet.cell(r, x)
                #global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                if cell.value == 'Cumulative totals':
                    cell_imp_whr = sheet.cell(r + 3, x + 1)
                    cell_unit = sheet.cell(r + 3, x + 2)
                    cell_expwh_unit = sheet.cell(r + 4, x + 2)
                    if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                        imp_whr = cell_imp_whr.value.replace(',', '')
                        imp_whh = round(float(imp_whr) / 1000000, 3)
                        imp_wh = "{:,}".format(imp_whh)
                        list1a.append(imp_wh)

                    cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                    if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                        exp_whr = cell_exp_whr.value.replace(',', '')
                        exp_whh = round(float(exp_whr) / 1000000, 3)
                        exp_wh = "{:,}".format(exp_whh)
                        list4.append(exp_wh)

                    cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                    cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                    if cell_app_pow_unit.value == 'VAh':
                        app_powr = cell_app_pow.value.replace(',', '')
                        app_powh = round(float(app_powr) / 1000000, 3)
                        app_pow = "{:,}".format(app_powh)
                        list5.append(app_pow)
                        # app_pow = float(app_powr)

                    # getting total import and export MVarh
                    unit1_MVArh = sheet.cell(r + 5, x + 2)
                    unit3_MVArh = sheet.cell(r + 7, x + 2)
                    global total_imp_MVArh
                    if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                        imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                        imp_MVArh1 = float(imp_MVArh1_r)
                        imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                        imp_MVArh2 = float(imp_MVArh2_r)
                        total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                        total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                        list6.append(total_imp_MVArh)

                    unit2_MVArh = sheet.cell(r + 6, x + 2)
                    unit4_MVArh = sheet.cell(r + 8, x + 2)
                    if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                        exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                        exp_MVArh1 = float(exp_MVArh1_r)
                        exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                        exp_MVArh2 = float(exp_MVArh2_r)
                        total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                        total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                        list7.append(total_exp_MVArh)

                        # total MVArh
                        total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(total_imp_MVArh.replace(',', ''))
                        total_MVArh_xx = round(total_MVArh_x, 3)
                        total_MVArh = "{:,}".format(total_MVArh_xx)
                        list8.append(total_MVArh)

                #global rate1,rate2,rate3,rate4,rate5,rate6
                if cell.value == 'Register':  # getting the various rates
                    unit1 = sheet.cell(r + 1, x + 2)
                    unit2 = sheet.cell(r + 2, x + 2)
                    unit3 = sheet.cell(r + 3, x + 2)
                    if unit3.value == unit2.value == 'Import Wh':
                        rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                        rate1_rh = round(float(rate1_r) / 1000000, 3)
                        rate1 = "{:,}".format(rate1_rh)
                        list9.append(rate1)

                    if unit1.value == unit3.value == 'Import Wh':
                        rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                        rate2_rh = round(float(rate2_r) / 1000000, 3)
                        rate2 = "{:,}".format(rate2_rh)
                        list10.append(rate2)

                    if unit2.value == unit1.value == 'Import Wh':
                        rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                        rate3_rh = round(float(rate3_r) / 1000000, 3)
                        rate3 = "{:,}".format(rate3_rh)
                        list11.append(rate3)

                    unit4 = sheet.cell(r + 4, x + 2)
                    unit5 = sheet.cell(r + 5, x + 2)
                    unit6 = sheet.cell(r + 6, x + 2)
                    if unit5.value == unit6.value == 'Export Wh':
                        rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                        rate4_rh = round(float(rate4_r) / 1000000, 3)
                        rate4 = "{:,}".format(rate4_rh)
                        list12.append(rate4)

                    if unit4.value == unit6.value == 'Export Wh':
                        rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                        rate5_rh = round(float(rate5_r) / 1000000, 3)
                        rate5 = "{:,}".format(rate5_rh)
                        list13.append(rate5)
                    #global rate6
                    if unit4.value == unit5.value == 'Export Wh':

                        rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                        rate6_rh = round(float(rate6_r) / 1000000, 3)
                        rate6 = "{:,}".format(rate6_rh)
                        list14.append(rate6)
                #global reset_time, resets
                if cell.value == 'Billing event details':
                    req_cell = sheet.cell(r + 2, x)
                    req_cell1 = sheet.cell(r + 3, x)
                    if req_cell.value == 'Billing reset number:':
                        resets = sheet.cell(r + 2, x + 1).value
                        # no_of_resets = float(no_of_resets_r)
                        list23.append(resets)

                    if req_cell1.value == 'Time of billing reset:':
                        reset_time = sheet.cell(r + 3, x + 1).value
                        list24.append(reset_time)
                #global M_O_S,M_O_S_r
                if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                            x).value == 'Billing period end date':
                    M_O_S_r = sheet.cell(r, x + 1).value
                    M_O_S_x = M_O_S_r.split('/')
                    M_O_S_i = int(M_O_S_x[1])
                    M_O_S = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'
                    list16.append(M_O_S)
                #global max_dem_1,max_dem_2,max_dem_3
                if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                             x).value == 'Register':  # getting maximum demands

                    unit_1 = sheet.cell(r + 3, x + 2)
                    unit_2 = sheet.cell(r + 4, x + 2)
                    unit_3 = sheet.cell(r + 5, x + 2)
                    if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                        max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                        max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                        max_dem_1 = "{:,}".format(max_dem_1_rh)
                        list17.append(max_dem_1)

                        max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                        max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                        max_dem_2 = "{:,}".format(max_dem_2_rh)
                        list18.append(max_dem_2)

                        max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                        max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                        max_dem_3 = "{:,}".format(max_dem_3_rh)
                        list19.append(max_dem_3)
                #global max_dem_1t, max_dem_2t, max_dem_3t

                if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(r + 2,
                                                                                                               x + 3).value == 'Time and date':  # getting max-demands time
                    unit_1t = sheet.cell(r + 3, x + 3)
                    unit_2t = sheet.cell(r + 4, x + 3)
                    unit_3t = sheet.cell(r + 5, x + 3)

                    max_dem_1t = unit_1t.value
                    max_dem_2t = unit_2t.value
                    max_dem_3t = unit_3t.value
                    list20.append(max_dem_1t)
                    list21.append(max_dem_2t)
                    list22.append(max_dem_3t)

                    #constants
                    unix = datetime.now().replace(second=0,microsecond=0)
                    inserted_by = 'kimera'
                    active_user = 'kimera'
                    pow_down_count = 'No Data'
                    pow_down_dt = 'No Data'
                    prog_count = 'No Data'
                    prog_count_dt = 'No Data'
                    #list25 = ['No data']
                    list25.append(pow_down_count)
                    list26.append(pow_down_dt)
                    list27.append(prog_count)
                    list28.append(prog_count_dt)
                    list29.append(unix)
                    list30.append(inserted_by)
                    list31.append(meter_no)
                    list32.append(D_O_R)
                    list33.append(read_time)



        i += 1
    elif i == len(list2):
        row_1 = list2[i-1]
        for r in range(row_1,sheet.max_row):
            for x in range(1,sheet.max_column+1):
                cell = sheet.cell(r, x)
                #global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                if cell.value == 'Cumulative totals':
                    cell_imp_whr = sheet.cell(r + 3, x + 1)
                    cell_unit = sheet.cell(r + 3, x + 2)
                    cell_expwh_unit = sheet.cell(r + 4, x + 2)
                    if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                        imp_whr = cell_imp_whr.value.replace(',', '')
                        imp_whh = round(float(imp_whr) / 1000000, 3)
                        imp_wh = "{:,}".format(imp_whh)
                        list1a.append(imp_wh)
                    cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                    if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                        exp_whr = cell_exp_whr.value.replace(',', '')
                        exp_whh = round(float(exp_whr) / 1000000, 3)
                        exp_wh = "{:,}".format(exp_whh)
                        list4.append(exp_wh)

                    cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                    cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                    if cell_app_pow_unit.value == 'VAh':
                        app_powr = cell_app_pow.value.replace(',', '')
                        app_powh = round(float(app_powr) / 1000000, 3)
                        app_pow = "{:,}".format(app_powh)
                        list5.append(app_pow)
                        # app_pow = float(app_powr)

                    # getting total import and export MVarh
                    unit1_MVArh = sheet.cell(r + 5, x + 2)
                    unit3_MVArh = sheet.cell(r + 7, x + 2)
                    #global total_imp_MVArh
                    if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                        imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                        imp_MVArh1 = float(imp_MVArh1_r)
                        imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                        imp_MVArh2 = float(imp_MVArh2_r)
                        total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                        total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                        list6.append(total_imp_MVArh)

                    unit2_MVArh = sheet.cell(r + 6, x + 2)
                    unit4_MVArh = sheet.cell(r + 8, x + 2)
                    if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                        exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                        exp_MVArh1 = float(exp_MVArh1_r)
                        exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                        exp_MVArh2 = float(exp_MVArh2_r)
                        total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                        total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                        list7.append(total_exp_MVArh)

                        # total MVArh
                        total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(
                            total_imp_MVArh.replace(',', ''))
                        total_MVArh_xx = round(total_MVArh_x, 3)
                        total_MVArh = "{:,}".format(total_MVArh_xx)
                        list8.append(total_MVArh)

                    # global rate1,rate2,rate3,rate4,rate5,rate6
                if cell.value == 'Register':  # getting the various rates
                    unit1 = sheet.cell(r + 1, x + 2)
                    unit2 = sheet.cell(r + 2, x + 2)
                    unit3 = sheet.cell(r + 3, x + 2)
                    if unit3.value == unit2.value == 'Import Wh':
                        rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                        rate1_rh = round(float(rate1_r) / 1000000, 3)
                        rate1 = "{:,}".format(rate1_rh)
                        list9.append(rate1)

                    if unit1.value == unit3.value == 'Import Wh':
                        rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                        rate2_rh = round(float(rate2_r) / 1000000, 3)
                        rate2 = "{:,}".format(rate2_rh)
                        list10.append(rate2)

                    if unit2.value == unit1.value == 'Import Wh':
                        rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                        rate3_rh = round(float(rate3_r) / 1000000, 3)
                        rate3 = "{:,}".format(rate3_rh)
                        list11.append(rate3)

                    unit4 = sheet.cell(r + 4, x + 2)
                    unit5 = sheet.cell(r + 5, x + 2)
                    unit6 = sheet.cell(r + 6, x + 2)
                    if unit5.value == unit6.value == 'Export Wh':
                        rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                        rate4_rh = round(float(rate4_r) / 1000000, 3)
                        rate4 = "{:,}".format(rate4_rh)
                        list12.append(rate4)

                    if unit4.value == unit6.value == 'Export Wh':
                        rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                        rate5_rh = round(float(rate5_r) / 1000000, 3)
                        rate5 = "{:,}".format(rate5_rh)
                        list13.append(rate5)
                    # global rate6
                    if unit4.value == unit5.value == 'Export Wh':
                        rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                        rate6_rh = round(float(rate6_r) / 1000000, 3)
                        rate6 = "{:,}".format(rate6_rh)
                        list14.append(rate6)
                    # global reset_time, resets
                if cell.value == 'Billing event details':
                    req_cell = sheet.cell(r + 2, x)
                    req_cell1 = sheet.cell(r + 3, x)
                    if req_cell.value == 'Billing reset number:':
                        resets = sheet.cell(r + 2, x + 1).value
                        # no_of_resets = float(no_of_resets_r)
                        list23.append(resets)

                    if req_cell1.value == 'Time of billing reset:':
                        reset_time = sheet.cell(r + 3, x + 1).value
                        list24.append(reset_time)
                    # global M_O_S,M_O_S_r
                if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                            x).value == 'Billing period end date':
                    M_O_S_r = sheet.cell(r, x + 1).value
                    M_O_S_x = M_O_S_r.split('/')
                    M_O_S_i = int(M_O_S_x[1])
                    M_O_S = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'
                    list16.append(M_O_S)
                    # global max_dem_1,max_dem_2,max_dem_3
                if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                             x).value == 'Register':  # getting maximum demands

                    unit_1 = sheet.cell(r + 3, x + 2)
                    unit_2 = sheet.cell(r + 4, x + 2)
                    unit_3 = sheet.cell(r + 5, x + 2)
                    if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                        max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                        max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                        max_dem_1 = "{:,}".format(max_dem_1_rh)
                        list17.append(max_dem_1)

                        max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                        max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                        max_dem_2 = "{:,}".format(max_dem_2_rh)
                        list18.append(max_dem_2)

                        max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                        max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                        max_dem_3 = "{:,}".format(max_dem_3_rh)
                        list19.append(max_dem_3)
                    # global max_dem_1t, max_dem_2t, max_dem_3t

                if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(r + 2,
                                                                                                               x + 3).value == 'Time and date':  # getting max-demands time
                    unit_1t = sheet.cell(r + 3, x + 3)
                    unit_2t = sheet.cell(r + 4, x + 3)
                    unit_3t = sheet.cell(r + 5, x + 3)

                    max_dem_1t = unit_1t.value
                    max_dem_2t = unit_2t.value
                    max_dem_3t = unit_3t.value
                    list20.append(max_dem_1t)
                    list21.append(max_dem_2t)
                    list22.append(max_dem_3t)

                    # constants
                    unix = datetime.now().replace(second=0, microsecond=0)
                    inserted_by = 'kimera'
                    active_user = 'kimera'
                    pow_down_count = 'No Data'
                    pow_down_dt = 'No Data'
                    prog_count = 'No Data'
                    prog_count_dt = 'No Data'
                    #list25 = ['No data']
                    list25.append(pow_down_count)
                    list26.append(pow_down_dt)
                    list27.append(prog_count)
                    list28.append(prog_count_dt)
                    list29.append(unix)
                    list30.append(inserted_by)
                    list31.append(meter_no)
                    list32.append(D_O_R)
                    list33.append(read_time)




df1_a = pd.DataFrame(list29)
df1_a.columns = ['time_stamp']
df1_b = pd.DataFrame(list30)
df1_b.columns = ['inserted_by']
df1_c = pd.DataFrame(list30)
df1_c.columns = ['meter_read_by']
df1_e = pd.DataFrame(list33)
df1_e.columns = ['reading_datetime']
df1_f = pd.DataFrame(list31)
df1_f.columns = ['meter_no']
df1 = pd.DataFrame(list16)
df1.columns = ['Energy_For']
df2 = pd.DataFrame(list1a)
df2.columns = ['Cum_Import']
df3 = pd.DataFrame(list4)
df3.columns = ['Cum_Export']
df4 = pd.DataFrame(list5)
df4.columns = ['Apparent_Power']
df5 = pd.DataFrame(list9)
df5.columns = ['Rate_1']
df6 = pd.DataFrame(list10)
df6.columns = ['Rate_2']
df7 = pd.DataFrame(list11)
df7.columns = ['Rate_3']
df8 = pd.DataFrame(list12)
df8.columns = ['Rate_4']
df9 = pd.DataFrame(list13)
df9.columns = ['Rate_5']
df10 = pd.DataFrame(list14)
df10.columns = ['Rate_6']
df11 = pd.DataFrame(list17)
df11.columns = ['Max_Dem_1']
df12 = pd.DataFrame(list20)
df12.columns = ['Max_Dem1_time']
df13 = pd.DataFrame(list18)
df13.columns = ['Max_Dem_2']
df14 = pd.DataFrame(list21)
df14.columns = ['Max_Dem2_time']
df15 = pd.DataFrame(list19)
df15.columns = ['Max_Dem_3']
df16 = pd.DataFrame(list22)
df16.columns = ['Max_Dem3_time']
df17 = pd.DataFrame(list6)
df17.columns = ['Import_MVArh']
df18 = pd.DataFrame(list7)
df18.columns = ['Export_MVArh']
df19 = pd.DataFrame(list8)
df19.columns = ['Total_MVAh']
df20 = pd.DataFrame(list23)
df20.columns = ['No_of_Resets']
df21 = pd.DataFrame(list24)
df21.columns = ['Last_Reset']
df22 = pd.DataFrame(list25)
df22.columns = ['Power_Down_Count']
df23 = pd.DataFrame(list26)
df23.columns = ['Lst_pwr_dwn_date_and_time']
df24 = pd.DataFrame(list27)
df24.columns = ['prog_count']
df25 = pd.DataFrame(list28)
df25.columns = ['last_prog_date']


df = pd.concat([df1_a,df1_b,df1_c,df1_e,df1_f,df1,df2,df3,df4,df5,df6,df7,df8,df9,df10,df11,df12,df13,df14,df15,df16,
                df17,df18,df19,df20,df21,df22,df23,df24,df25], axis=1)
engine = create_engine('mysql://root:@localhost/meteringdatabase')
df.to_sql('data_sets', con=engine, if_exists='append', index=False,
dtype={'time_stamp': VARCHAR(length=25),'inserted_by': VARCHAR(length=20),'meter_read_by': VARCHAR(length=10),
       'reading_datetime': VARCHAR(length=25),'meter_no': VARCHAR(length=12), 'Energy_For': VARCHAR(length=15),'Cum_Import': VARCHAR(length=13),
        'Cum_Export':VARCHAR(length=13),'Apparent_Power':VARCHAR(length=13),'Rate_1':VARCHAR(length=13),'Rate_2':VARCHAR(length=13),'Rate_3':VARCHAR(length=13),
        'Rate_4':VARCHAR(length=13),'Rate_5':VARCHAR(length=13),'Rate_6':VARCHAR(length=13),'Max_Dem_1':VARCHAR(length=13),'Max_Dem1_time':VARCHAR(length=15),
        'Max_Dem_2':VARCHAR(length=13),'Max_Dem2_time':VARCHAR(length=15),'Max_Dem_3':VARCHAR(length=13),'Max_Dem3_time':VARCHAR(length=15),
        'Import_MVArh':VARCHAR(length=13),'Export_MVArh':VARCHAR(length=13),'Total_MVAh':VARCHAR(length=13),'No_of_Resets':VARCHAR(5),
        'Last_Reset':VARCHAR(length=25),'Power_Down_Count':VARCHAR(length=5),'Lst_pwr_dwn_date_and_time':VARCHAR(length=5),
        'prog_count':VARCHAR(5),'last_prog_date':VARCHAR(length=25)
                                        })
df.to_excel('C:\\Users\\KIMERA\Desktop\\mutundwe\\output.xlsx')
print(df)


