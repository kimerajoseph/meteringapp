from django.shortcuts import render
from django.http import HttpResponse, request, JsonResponse, Http404,HttpResponseRedirect
import datetime
from datetime import datetime,date
from datetime import timedelta
import json
import time
import pymysql
import pandas as pd
import openpyxl as py
from pygrok import Grok
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.models import AnonymousUser
import openpyxl as xl
#import datetime
import xlsxwriter
from sqlalchemy import create_engine
from sqlalchemy.types import String, SmallInteger,VARCHAR,INT,Float
import matplotlib.pyplot as plt
import mpld3
from django.core.files.storage import FileSystemStorage
import io
import calendar
from openpyxl import Workbook
import csv
import os
import math
import dash
import dash_core_components as dcc
import dash_html_components as html
import plotly.graph_objects as go
from dash.dependencies import Input,Output
from django_plotly_dash import DjangoDash
import plotly.express as px
import tabula
from tabula import read_pdf
import glob
import dateutil
from dateutil.relativedelta import relativedelta
import numpy as np
#from metering_database.forms import inputform

from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.decorators import login_required


#
# conn = pymysql.connect(host='localhost', port=3306,
#                        user='root', password='', database='meteringdatabase')

conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                     user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
c = conn.cursor()


#GETTING LIST OF ALL CEWE PROMETER LISTS
cewe_P100_list_m = []
cewe_P100_list_c = []
tables = ['standalone_meters', 'substation_meters', 'ipp_meters']
for table in tables:
    sql = f"SELECT meter_no from {table} WHERE meter_manufacturer = 'Cewe' AND meter_type = 'Prometer 100'"
    sql1 = f"SELECT meter_no_ch from {table} WHERE meter_manufacturer_ch = 'Cewe' AND meter_type_ch = 'Prometer 100'"
    c.execute(sql)
    data_all = c.fetchall()
    c.execute(sql1)
    data_all1 = c.fetchall()
    for data in data_all:
        data1 = data[0]
        cewe_P100_list_m.append(data1)

    for data1 in data_all1:
        data2 = data1[0]
        cewe_P100_list_c.append(data2)

#GETTING A LIST OF MAIN ENERGY METERS WHERE WORKS HAVE BEEN CARRIED OUT
c.execute("select Meter_No from metering_database_work")
datrex = c.fetchall()
meter_work_list = []
for dd in datrex:
    #global work_list
    meterr = dd[0]
    meter_work_list.append(meterr)
month_names_in_year = []
for i in range(0,13):
    monttt = calendar.month_name[i]
    month_names_in_year.append(monttt)

use_year_list = []
unix = datetime.now().year
for i in range(0,3):
    yearr = unix - i
    use_year_list.append(yearr)
#print(year_list)

#GETTING LIST OF SUBSTATION, IPP AND STAND ALONE METERS
global final_ipp_listall,final_sub_listall,final_stand_listall
c.execute("SELECT name_ipp FROM ipp_meters")
ipp_data = c.fetchall()
ipp_listall = []
for ipp in ipp_data:
    # print(ipp)
    ippx = ipp[0]
    ipp_listall.append(ippx)
# print(ipp_listall)
final_ipp_listall = list(dict.fromkeys(ipp_listall))

c.execute("SELECT substation FROM substation_meters")
sub_data = c.fetchall()
# print(sub_data)
sub_listall = []
for sub in sub_data:
    subx = sub[0]
    sub_listall.append(subx)
final_sub_listall = list(dict.fromkeys(sub_listall))
#print(final_sub_listall)

c.execute("SELECT district FROM standalone_meters")
stand_data = c.fetchall()
# print(stand_data)
stand_listall = []
for stand in stand_data:
    standx = stand[0]
    stand_listall.append(standx)
final_stand_listall = list(dict.fromkeys(stand_listall))
#print(final_stand_listall)


def index(request):
    return render(request, 'index.html')

def homepage(request):
    return render(request, 'index.html')

def welcome(request):
    if request.method == 'POST':
        global user,active_user, AnonymousUser
        user_act = request.POST['username']
        password_given = request.POST['password']
        user = authenticate(username = user_act, password = password_given)
        if user is not None and user.is_staff: #UETCL STAFF
            #global active_user,AnonymousUser
            active_user = user_act
            #AnonymousUser = user
            return render(request, 'welcome.html',{'user':user})

        elif user is not None and user.groups.filter(name='IPPs').exists(): #IPPs LANDING PAGE
            #global
            active_user = user_act
            AnonymousUser = user
            return render(request, 'ipp_landing_page.html', {'user': user})

        elif user is not None and user.groups.filter(name='umeme').exists(): #UMEME LANDING PAGE
            #global
            active_user = user_act
            AnonymousUser = user
            return render(request, 'umeme_page.html', {'user': user})

        else:
            error = 'Please check your username and password'
            return render(request, 'index.html', {'error':error})

#CHANGE PASSWORD
def change_pw(request):
    formx = PasswordChangeForm(user=user)
    return render(request, 'change_pw.html', {'user': user, 'formx':formx})

#@login_required
def comfirm_change_pw(request):
    if request.method == 'POST':
        form = PasswordChangeForm(data=request.POST, user=request.user)
        # print(user)
        # print(data)
        AnonymousUser = user
        #form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            print(user)
            update_session_auth_hash(request, user)  # Important!
            return render(request, 'index.html')
        else:
            return HttpResponse("PASSWORD NOT CHANGED")
            #return render(request, 'index.html')
            #messages.error(request, 'Please correct the error below.')
    # else:
    #     form = PasswordChangeForm(request.user)

    # return render(request, 'index.html')




def meterrecord(request):  #inserting meter records for standalone meter
    if request.method == 'POST':
        unix = datetime.now().replace (microsecond=0)
        time_added = unix
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        village = request.POST['village_st']
        subcounty = request.POST['sub_county_st']
        district = request.POST['district_st']
        region = request.POST['region_st']
        x_cordinates = request.POST['x_cordinates_st']
        y_cordinates = request.POST['y_cordinates_st']
        feeder_name = request.POST['feeder_name_st']
        voltage = request.POST['voltage_st']
        distributor_1 = request.POST['distributor_1']
        distributor_2 = request.POST['distributor_2']
        meter_owner = request.POST['meter_owner_st']

        manu_st = request.POST['manu_meter_st'] #second take
        unit_type = request.POST['meter_type_st']
        D_O_M = request.POST['DOM_st']
        MU_Serial_No = request.POST['mu_meter_no_st']
        no_of_elements = request.POST['elements_st']
        wired_as = request.POST['wired_as_st']
        comm_date = request.POST['comm_date_st']
        DLP = request.POST['DLP_st']
        metering_cores = request.POST['meter_cores_st']
        CTratios = request.POST['core_ratios_st']
        core_used = request.POST['core_used_st']
        core_used_accuracy = request.POST['accuracy_st']
        avail_spares = request.POST['avail_spares_st']
        spares_class = request.POST['spares_class_st']
        VTratio = request.POST['vt_ratio_st']
        VT_accuracy = request.POST['vt_accuracy_st']
        #meter_no = request.POST['meter_no_st']
        #global meter_no
        #meter_no = request.POST['meter_no']
        global active_user
        added_by = active_user
        #global meter

        meter_manuf = request.POST['energy_meter_manu'] #main energy meter details
        meter_type = request.POST['energy_meter_type']
        meter_no = request.POST['energy_meter_no']
        Y_O_M = request.POST['meter_YOM']
        meter_accuracy = request.POST['e_meter_accuracy']
        no_of_meter_elements = request.POST['e_meter_elements']
        meter_wired_as = request.POST['e_meter_wired']
        meter_install_date = request.POST['e_instal_date']
        meter_decom_date = request.POST['e_decom_date']
        access = request.POST['access']
        IP_Address = request.POST['ip_address']
        avail_interfaces = request.POST['interfaces']
        comm_protocol = request.POST['comm_protocol']
        comm_protocol_used = request.POST['comm_protocol_used']

        meter_manuf_ch = request.POST['energy_meter_manu_ch']  # main energy meter details
        meter_type_ch = request.POST['energy_meter_type_ch']
        meter_no_ch = request.POST['energy_meter_no_ch']
        Y_O_M_ch = request.POST['meter_YOM_ch']
        meter_accuracy_ch = request.POST['e_meter_accuracy_ch']
        no_of_meter_elements_ch = request.POST['e_meter_elements_ch']
        meter_wired_as_ch = request.POST['e_meter_wired_ch']
        meter_install_date_ch = request.POST['e_instal_date_ch']
        meter_decom_date_ch = request.POST['e_decom_date_ch']
        access_ch = request.POST['access_ch']
        IP_Address_ch = request.POST['ip_address_ch']
        avail_interfaces_ch = request.POST['interfaces_ch']
        comm_protocol_ch = request.POST['comm_protocol_ch']
        comm_protocol_used_ch = request.POST['comm_protocol_used_ch']

        #writting to excel sheet
        workbook = xlsxwriter.Workbook(f'Stand_Alone_Meters/{meter_no}_DETAILS.xlsx') #write to excel
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True, 'font_size':16,'align':'center'})
        bold1 = workbook.add_format({'bold': True, 'font_size': 13})
        worksheet.merge_range('A1:D1', 'Merged Range', bold)

        worksheet.write('A1', f'{meter_no}_DETAILS', bold)
        worksheet.write('A3', 'Time Added',bold1)
        worksheet.write('B3', str(unix))
        worksheet.write('C3', 'Added By', bold1)
        worksheet.write('D3', active_user)
        worksheet.write('A4', 'Village',bold1)
        worksheet.write('B4', village)
        worksheet.write('C4', 'Sub-County', bold1)
        worksheet.write('D4', subcounty)
        worksheet.write('A5', 'District', bold1)
        worksheet.write('B5', district)
        worksheet.write('C5', 'Region', bold1)
        worksheet.write('D5', region)
        worksheet.write('A6', 'X Cordinates', bold1)
        worksheet.write('B6', x_cordinates)
        worksheet.write('C6', 'Y Cordinates', bold1)
        worksheet.write('D6', y_cordinates)
        worksheet.write('A7', 'Feeder Name', bold1)
        worksheet.write('B7', feeder_name)
        worksheet.write('C7', 'Voltage', bold1)
        worksheet.write('D7', voltage)
        worksheet.write('A8', 'Distributor 1', bold1)
        worksheet.write('B8', distributor_1)
        worksheet.write('C8', 'Distributor 2', bold1)
        worksheet.write('D8', distributor_2)
        worksheet.write('A9', 'Meter Owner', bold1)
        worksheet.write('B9', meter_owner)

        worksheet.merge_range('A11:D11', 'Merged Range', bold)
        worksheet.write('A11', "METERING UNIT DETAILS",bold)
        worksheet.write('A12', 'Manufacturer', bold1)
        worksheet.write('B12', manu_st)
        worksheet.write('C12', 'Unit Type', bold1)
        worksheet.write('D12', unit_type)
        worksheet.write('A13', 'Date of Manufacture', bold1)
        worksheet.write('B13', D_O_M)
        worksheet.write('C13', 'Serial No', bold1)
        worksheet.write('D13', MU_Serial_No)
        worksheet.write('A14', 'No of Elements', bold1)
        worksheet.write('B14', no_of_elements)
        worksheet.write('C14', 'Wired As', bold1)
        worksheet.write('D14', wired_as)
        worksheet.write('A15', 'Commissioning Date', bold1)
        worksheet.write('B15', comm_date)
        worksheet.write('C15', 'DLP', bold1)
        worksheet.write('D15', DLP)
        worksheet.write('A16', 'No of Metering cores', bold1)
        worksheet.write('B16', metering_cores)
        worksheet.write('C16', 'CT Ratios', bold1)
        worksheet.write('D16', CTratios)
        worksheet.write('A17', 'Core Used', bold1)
        worksheet.write('B17', core_used)
        worksheet.write('C17', 'Accuracy of Used Core', bold1)
        worksheet.write('D17', core_used_accuracy)
        worksheet.write('A18', 'No. of Avail. Spares', bold1)
        worksheet.write('B18', avail_spares)
        worksheet.write('C18', 'Spares Class', bold1)
        worksheet.write('D18', spares_class)
        worksheet.write('A20', 'VT Ration', bold1)
        worksheet.write('B20', VTratio)
        worksheet.write('C20', 'VT Accuracy', bold1)
        worksheet.write('D20', VT_accuracy)

        worksheet.merge_range('A21:D21', 'Merged Range', bold)
        worksheet.write('A21', "MAIN METER DETAILS", bold)
        worksheet.write('A22', 'Meter Manufacturer', bold1)
        worksheet.write('B22', meter_manuf)
        worksheet.write('C22', 'Meter Type', bold1)
        worksheet.write('D22', meter_type)
        worksheet.write('A23', 'Meter No', bold1)
        worksheet.write('B23', meter_no)
        worksheet.write('C23', 'Meter Y.O.M', bold1)
        worksheet.write('D23', Y_O_M)
        worksheet.write('A24', 'Meter Accuracy', bold1)
        worksheet.write('B24', meter_accuracy)
        worksheet.write('C24', 'No. of Elements', bold1)
        worksheet.write('D24', no_of_meter_elements)
        worksheet.write('A25', 'Wired AS', bold1)
        worksheet.write('B25', meter_wired_as)
        worksheet.write('C25', 'Installation Date', bold1)
        worksheet.write('D25', meter_install_date)
        worksheet.write('A26', 'Decommisioning Date', bold1)
        worksheet.write('B26', meter_decom_date)
        worksheet.write('C26', 'Access', bold1)
        worksheet.write('D26', access)
        worksheet.write('A27', 'IP Address', bold1)
        worksheet.write('B27', IP_Address)
        worksheet.write('C27', 'Avail. Interfaces', bold1)
        worksheet.write('D27', avail_interfaces)
        worksheet.write('A28', 'Comm Protocol', bold1)
        worksheet.write('B28', comm_protocol)
        worksheet.write('C28', 'Comm Protocol Used', bold1)
        worksheet.write('D28', comm_protocol_used)

        worksheet.merge_range('A30:D30', 'Merged Range', bold)
        worksheet.write('A30', "CHECK METER DETAILS", bold)
        worksheet.write('A31', 'Meter Manufacturer', bold1)
        worksheet.write('B31', meter_manuf_ch)
        worksheet.write('C31', 'Meter Type', bold1)
        worksheet.write('D31', meter_type_ch)
        worksheet.write('A32', 'Meter No', bold1)
        worksheet.write('B32', meter_no_ch)
        worksheet.write('C32', 'Meter Y.O.M', bold1)
        worksheet.write('D32', Y_O_M_ch)
        worksheet.write('A33', 'Meter Accuracy', bold1)
        worksheet.write('B33', meter_accuracy_ch)
        worksheet.write('C33', 'No. of Elements', bold1)
        worksheet.write('D33', no_of_meter_elements_ch)
        worksheet.write('A34', 'Wired AS', bold1)
        worksheet.write('B34', meter_wired_as_ch)
        worksheet.write('C34', 'Installation Date', bold1)
        worksheet.write('D34', meter_install_date_ch)
        worksheet.write('A35', 'Decommisioning Date', bold1)
        worksheet.write('B35', meter_decom_date_ch)
        worksheet.write('C35', 'Access', bold1)
        worksheet.write('D35', access_ch)
        worksheet.write('A36', 'IP Address', bold1)
        worksheet.write('B36', IP_Address_ch)
        worksheet.write('C36', 'Avail. Interfaces', bold1)
        worksheet.write('D36', avail_interfaces_ch)
        worksheet.write('A367', 'Comm Protocol', bold1)
        worksheet.write('B367', comm_protocol_ch)
        worksheet.write('C367', 'Comm Protocol Used', bold1)
        worksheet.write('D367', comm_protocol_used_ch)


        workbook.close()


        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                              user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        sql_1 = """INSERT INTO standalone_meters (time_added,added_by,village,subcounty,district,region,x_cordinates,y_cordinates,feeder_name,
         voltage,distributor_1,distributor_2,meter_owner,manufacturer,unit_type,D_O_M,MU_Serial_No,no_of_elements,wired_as,comm_date,DLP,metering_cores,CTratios,
          core_used,core_used_accuracy,avail_spares,spares_class,VTratio,VT_accuracy,meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy,no_of_meter_elements,
          meter_wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocol,comm_protocol_used,meter_manufacturer_ch,meter_type_ch,meter_no_ch,
          meter_YOM_ch,meter_accuracy_ch,no_of_meter_elements_ch,meter_wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocol_ch,
          comm_protocol_used_ch) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
        %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        values = (unix,added_by,village,subcounty,district,region,x_cordinates,y_cordinates,feeder_name,voltage,distributor_1,distributor_2,meter_owner,manu_st,
                  unit_type,D_O_M,MU_Serial_No,no_of_elements,wired_as,comm_date,DLP,metering_cores,CTratios,core_used,core_used_accuracy,avail_spares,spares_class,VTratio,VT_accuracy,
                  meter_manuf,meter_type,meter_no,Y_O_M,meter_accuracy,no_of_meter_elements,meter_wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,
                  comm_protocol,comm_protocol_used,meter_manuf_ch,meter_type_ch,meter_no_ch,Y_O_M_ch,meter_accuracy_ch,no_of_meter_elements_ch,meter_wired_as_ch,
                  meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocol_ch,comm_protocol_used_ch)
        c.execute(sql_1, values)
        c.close()
        conn.commit()
        conn.close()
        return render(request, 'form_success.html')

def sub_meterrecord(request):  #inserting substation meter records
    if request.method == 'POST':

        if 'config_file' in request.FILES: #GETTING CONFIGURATION FILE
            pp = f'Configuration Files Main'
            # pp = f'files_for_{active_user}'
            storage = str(pp)
            fs = FileSystemStorage()
            fs1 = FileSystemStorage(location=storage)
            if not os.path.exists(f"{storage}"):
                os.makedirs(f"{storage}")

            config_file = request.FILES['config_file']
            config_filename = fs.save(config_file.name, config_file)
            configuration_file = fs.url(config_filename)
            filex = fs1.save(config_file.name, config_file)
            print(config_file)
            print(configuration_file)

        else:
            configuration_file = ''

        if 'config_file_ch' in request.FILES: #GETTING CHECK METER CONFIGURATION FILE
            pp = f'Configuration Files Check'
            # pp = f'files_for_{active_user}'
            storage = str(pp)
            fs = FileSystemStorage()
            fs1 = FileSystemStorage(location=storage)
            if not os.path.exists(f"{storage}"):
                    os.makedirs(f"{storage}")
            config_file_ch = request.FILES['config_file_ch']
            config_filename_ch = fs.save(config_file_ch.name, config_file_ch)
            configuration_file_ch = fs.url(config_filename_ch)
            filex_ch = fs1.save(config_file_ch.name, config_file_ch)

        else:
            configuration_file_ch = ''
        unix = datetime.now().replace (microsecond=0)
        time_added = unix

        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        substation = request.POST['sub']
        voltage = request.POST['voltage']
        district = request.POST['sub_dist']
        region = request.POST['sub_region']
        x_cordinates = request.POST['sub_cord_x']
        y_cordinates = request.POST['sub_cord_y']

        component = request.POST['comp']        #ADDITIONS TO FORM
        capacity = request.POST['capacity']
        network_type = request.POST['net_type']


        component_name = request.POST['sub_feeder'] #second take
        component_voltage = request.POST['sub_feeder_voltage']
        meter_owner = request.POST['sub_meter_owner']
        distributor = request.POST['distributor']
        Comments = request.POST['subs_contractor']
        ct_manufacturer = request.POST['sub_ct_manu'] #Third take
        ct_type = request.POST['sub_ct_type']
        no_ct_cores = request.POST['sub_ct_cores']
        ct_ratios = request.POST['sub_ct_core_ratios']
        core_used = request.POST['sub_core_used']
        accuracy_class = request.POST['sub_accuracy_class']
        avail_spares = request.POST['sub_avail_spares']
        spares_class = request.POST['sub_spares_class']
        VT_manufacturer = request.POST['sub_vt_manu']
        VT_type = request.POST['sub_VT_type']
        VT_ratio = request.POST['sub_vt_ratio']
        VT_accuracy = request.POST['sub_vt_accuracy']
        meter_manufacturer = request.POST['sub_meter_manu'] #fourth take
        meter_type = request.POST['sub_meter_type']
        meter_no = request.POST['sub_meter_no']
        meter_YOM = request.POST['sub_YOM']
        meter_accuracy_class = request.POST['sub_meter_acc_class']
        no_of_elements = request.POST['sub_meter_elements']
        wired_as = request.POST['sub_meter_wire']
        meter_install_date = request.POST['sub_meter_install_date']
        meter_decom_date = request.POST['sub_meter_decom_date']
        access = request.POST['sub_meter_access']
        IP_Address = request.POST['sub_ip_address']
        avail_interfaces = request.POST['sub_meter_interfaces']
        comm_protocols = request.POST['sub_meter_comm_protocols']
        protocol_used = request.POST['sub_meter_protocol_used']


        meter_manufacturer_ch = request.POST['sub_meter_manu_ch']  # fifth take
        meter_type_ch = request.POST['sub_meter_type_ch']
        meter_no_ch = request.POST['sub_meter_no_ch']
        meter_YOM_ch = request.POST['sub_YOM_ch']
        meter_accuracy_class_ch = request.POST['sub_meter_acc_class_ch']
        no_of_elements_ch = request.POST['sub_meter_elements_ch']
        wired_as_ch = request.POST['sub_meter_wire_ch']
        meter_install_date_ch = request.POST['sub_meter_install_date_ch']
        meter_decom_date_ch = request.POST['sub_meter_decom_date_ch']
        access_ch = request.POST['sub_meter_access_ch']
        IP_Address_ch = request.POST['sub_ip_address_ch']
        avail_interfaces_ch = request.POST['sub_meter_interfaces_ch']
        comm_protocols_ch = request.POST['sub_meter_comm_protocols_ch']
        protocol_used_ch = request.POST['sub_meter_protocol_used_ch']

        global active_user
        added_by = active_user

        # writting to excel sheet
        workbook = xlsxwriter.Workbook(f'Substation_Meters/{meter_no}_{component_name}_METERING_POINT_DETAILS.xlsx')  # write to excel
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center'})
        bold2 = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        bold1 = workbook.add_format({'bold': True, 'font_size': 13})
        worksheet.merge_range('A1:D1', 'Merged Range', bold)
        worksheet.merge_range('A2:D2', 'Merged Range', bold2)

        worksheet.write('A1', f'{component_name}_{meter_no}_DETAILS', bold)
        worksheet.write('A2', 'LOCATION DETAILS', bold2)
        worksheet.write('A3', 'Time Added', bold1)
        worksheet.write('B3', str(unix))
        worksheet.write('C3', 'Added By', bold1)
        worksheet.write('D3', active_user)
        worksheet.write('A4', 'Substation', bold1)
        worksheet.write('B4', substation)
        worksheet.write('C4', 'Voltage', bold1)
        worksheet.write('D4', voltage)
        worksheet.write('A5', 'District', bold1)
        worksheet.write('B5', district)
        worksheet.write('C5', 'Region', bold1)
        worksheet.write('D5', region)
        worksheet.write('A6', 'X Cordinates', bold1)
        worksheet.write('B6', x_cordinates)
        worksheet.write('C6', 'Y Cordinates', bold1)
        worksheet.write('D6', y_cordinates)

        worksheet.merge_range('A8:D8', 'Merged Range', bold2)
        worksheet.write('A8', 'COMPONENT DETAILS', bold2)
        # component = request.POST['comp']  # ADDITIONS TO FORM
        # capacity = request.POST['capacity']
        # network_type = request.POST['network_type']
        worksheet.write('A9', 'Component', bold1)
        worksheet.write('B9', component)
        worksheet.write('C9', 'Capacity', bold1)
        worksheet.write('D9', capacity)

        worksheet.write('A10', 'Network Type', bold1)
        worksheet.write('B10', network_type)
        worksheet.write('C10', 'Component Name', bold1)
        worksheet.write('D10', component_name)

        # worksheet.write('A9', 'Component Name', bold1)
        # worksheet.write('B9', component_name)
        worksheet.write('A11', 'Component Voltage', bold1)
        worksheet.write('B11', component_voltage)
        worksheet.write('C11', 'Meeter Owner', bold1)
        worksheet.write('D11', meter_owner)
        worksheet.write('A12', 'Distributor', bold1)
        worksheet.write('B12', distributor)
        worksheet.write('C12', 'Comments', bold1)
        worksheet.write('D12', Comments)

        worksheet.merge_range('A13:D13', 'Merged Range', bold2)
        worksheet.write('A13', 'CT & VT DETAILS', bold2)
        worksheet.write('A14', 'CT Manufacturer', bold1)
        worksheet.write('B14', ct_manufacturer)
        worksheet.write('C14', 'CT Type', bold1)
        worksheet.write('D14', ct_type)
        worksheet.write('A15', 'No of Metering CT cores', bold1)
        worksheet.write('B15', no_ct_cores)
        worksheet.write('C15', 'CT Ratios', bold1)
        worksheet.write('D15', ct_ratios)
        worksheet.write('A16', 'Core Used', bold1)
        worksheet.write('B16', core_used)
        worksheet.write('C16', 'Accuracy Class', bold1)
        worksheet.write('D16', accuracy_class)
        worksheet.write('A17', 'Available Spares', bold1)
        worksheet.write('B17', avail_spares)
        worksheet.write('C17', 'Spares Class', bold1)
        worksheet.write('D17', spares_class)
        worksheet.write('A18', 'VT Manufacturer', bold1)
        worksheet.write('B18', VT_manufacturer)
        worksheet.write('C18', 'VT Type', bold1)
        worksheet.write('D18', VT_type)
        worksheet.write('A19', 'VT Ratio', bold1)
        worksheet.write('B19', VT_ratio)
        worksheet.write('C19', 'VT Accuracy', bold1)
        worksheet.write('D19', VT_accuracy)

        worksheet.merge_range('A21:D21', 'MAIN METER DETAILS', bold)
        #worksheet.write('A21', "MAIN METER DETAILS", bold)
        worksheet.write('A22', 'Meter Manufacturer', bold1)
        worksheet.write('B22', meter_manufacturer)
        worksheet.write('C22', 'Meter Type', bold1)
        worksheet.write('D22', meter_type)
        worksheet.write('A23', 'Meter No', bold1)
        worksheet.write('B23', meter_no)
        worksheet.write('C23', 'Meter Y.O.M', bold1)
        worksheet.write('D23', meter_YOM)
        worksheet.write('A24', 'Meter Accuracy', bold1)
        worksheet.write('B24', meter_accuracy_class)
        worksheet.write('C24', 'No. of Elements', bold1)
        worksheet.write('D24', no_of_elements)
        worksheet.write('A25', 'Wired AS', bold1)
        worksheet.write('B25', wired_as)
        worksheet.write('C25', 'Installation Date', bold1)
        worksheet.write('D25', meter_install_date)
        worksheet.write('A26', 'Decommisioning Date', bold1)
        worksheet.write('B26', meter_decom_date)
        worksheet.write('C26', 'Access', bold1)
        worksheet.write('D26', access)
        worksheet.write('A27', 'IP Address', bold1)
        worksheet.write('B27', IP_Address)
        worksheet.write('C27', 'Avail. Interfaces', bold1)
        worksheet.write('D27', avail_interfaces)
        worksheet.write('A28', 'Comm Protocol', bold1)
        worksheet.write('B28', comm_protocols)
        worksheet.write('C28', 'Comm Protocol Used', bold1)
        worksheet.write('D28', protocol_used)

        worksheet.merge_range('A30:D30', 'Merged Range', bold)
        worksheet.write('A30', "CHECK METER DETAILS", bold)
        worksheet.write('A31', 'Meter Manufacturer', bold1)
        worksheet.write('B31', meter_manufacturer_ch)
        worksheet.write('C31', 'Meter Type', bold1)
        worksheet.write('D31', meter_type_ch)
        worksheet.write('A32', 'Meter No', bold1)
        worksheet.write('B32', meter_no_ch)
        worksheet.write('C32', 'Meter Y.O.M', bold1)
        worksheet.write('D32', meter_YOM_ch)
        worksheet.write('A33', 'Meter Accuracy', bold1)
        worksheet.write('B33', meter_accuracy_class_ch)
        worksheet.write('C33', 'No. of Elements', bold1)
        worksheet.write('D33', no_of_elements_ch)
        worksheet.write('A34', 'Wired AS', bold1)
        worksheet.write('B34', wired_as_ch)
        worksheet.write('C34', 'Installation Date', bold1)
        worksheet.write('D34', meter_install_date_ch)
        worksheet.write('A35', 'Decommisioning Date', bold1)
        worksheet.write('B35', meter_decom_date_ch)
        worksheet.write('C35', 'Access', bold1)
        worksheet.write('D35', access_ch)
        worksheet.write('A36', 'IP Address', bold1)
        worksheet.write('B36', IP_Address_ch)
        worksheet.write('C36', 'Avail. Interfaces', bold1)
        worksheet.write('D36', avail_interfaces_ch)
        worksheet.write('A37', 'Comm Protocol', bold1)
        worksheet.write('B37', comm_protocols_ch)
        worksheet.write('C37', 'Comm Protocol Used', bold1)
        worksheet.write('D37', protocol_used_ch)

        workbook.close()

        # conn = pymysql.connect(host='localhost', port=3306,
        #                user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        sql_2 = """INSERT INTO substation_meters (time_added,added_by,substation,voltage, district,region,x_cordinates,y_cordinates,component,capacity,network_type,component_name,component_voltage,meter_owner,
                    distributor,Comments,ct_manufacturer,ct_type,no_ct_cores,ct_ratios,core_used,accuracy_class,avail_spares,spares_class,VT_manufacturer,VT_type,VT_ratio,VT_accuracy,
                    meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy_class,no_of_elements,wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocols,protocol_used,
                    meter_manufacturer_ch,meter_type_ch,meter_no_ch,meter_YOM_ch,meter_accuracy_class_ch,no_of_elements_ch,wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,
                    comm_protocols_ch,protocol_used_ch,config_file,config_file_ch) 
                    VALUES (%s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                     %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,%s)"""
        values = (unix,added_by,substation,voltage,district,region,x_cordinates,y_cordinates,component,capacity,network_type,component_name,component_voltage,meter_owner,distributor,Comments,
                  ct_manufacturer,ct_type,no_ct_cores,ct_ratios,core_used,accuracy_class,avail_spares,spares_class,VT_manufacturer,VT_type,VT_ratio,VT_accuracy,
                  meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy_class,no_of_elements,wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocols,protocol_used,
                  meter_manufacturer_ch,meter_type_ch,meter_no_ch,meter_YOM_ch,meter_accuracy_class_ch,no_of_elements_ch,wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,
                  comm_protocols_ch,protocol_used_ch,configuration_file,configuration_file_ch)

        c.execute(sql_2, values)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')

def new_submission(request):
    #user = user_act
    #user = active_user
    return render(request, 'welcome.html',{'user':user})

def success(request):
    return render(request, 'uetcl_meter.html')
    #if request.method == 'POST':
        #user = request.POST['username']
        #password_given = request.POST['password']
        #unix = int(time.time())
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
now = datetime.now()
#assigning the excel sheet a dynamic name
xlm = now.strftime('%B')
#import csv
def fillform(request):
    if request.method == 'POST':
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()

        global sales_month,imp_wh, exp_wh, app_pow, rate1, rate2, rate3, rate4, rate5, rate6, max_dem_1, max_dem_1t, max_dem_2
        global max_dem_2t, max_dem_3, max_dem_3t, imp_mvarh, exp_mvarh, total_mvarh, resets, reset_time, pow_down_count
        global pow_down_dt, prog_count, prog_count_dt
        global total_exp_MVArh, total_imp_MVArh,sales,total_exp_MVArh, imp_wh

        listz = []
        csv_list = []
        global active_user, our_month,unix,our_year,our_file_name,df
        unix = datetime.now().__format__("%Y-%m-%d %H:%M")
        attache = str(unix)
        pp = f'files_for_{active_user}'
        storage = str(pp)
        for f in request.FILES.getlist('filename_m'):
            fs = FileSystemStorage(location=storage)
            file_m = fs.save(f.name,f)
            listz.append(f.name)
        #print(listz)
        list_len = len(listz)
        #print(list_len)
        excel_list = []
        read_months = []
        global read_datetime, meter_no
        for element in listz:
            #print(f'THIS IS THE ELEMENT : {element}')
            my_cols = ['A','B','C','D']
            df = pd.read_csv(f'{storage}/{element}',names = my_cols)
            data = df.iat[1, 1]
            # stream = data.replace('-', '')
            meter_nox = data.split()[0]
            meter_no = meter_nox.replace('-', '')
            #csv_date = data.split()[1]
            #actual_date = csv_date.split('/')
            #i = int(actual_date[1])
            row = df.index[df['A'] == 'Billing period start date'].tolist()
            save_values = df.loc[row[0], 'B'].split('/')
            our_month = calendar.month_name[int(save_values[1])]
            year_month = f'{our_month}_{save_values[2]}'
            our_year = int(save_values[2])
            #our_month = calendar.month_name[i]
            our_file_name = f'{meter_no}_{our_month}_{our_year}'
            if not os.path.exists(f'{storage}/Elster/historical/{our_month}_{our_year}_files'):
                os.makedirs(f'{storage}/Elster/historical/{our_month}_{our_year}_files')
            df.to_excel(f'{storage}/Elster/historical/{our_month}_{our_year}_files/{our_file_name}.xlsx',index=None,header=False,sheet_name='sheet1')
            excel_list.append(f'{our_file_name}.xlsx')
            read_months.append(year_month)
        root = os.getcwd()
        dirlist = [item for item in os.listdir(root) if os.path.isdir(os.path.join(root, item))]
        #read_monthlist = ['January_2020_files','November_2020_files','December_2020_files']
        for excel in excel_list:
            for dir in dirlist:
                y = excel.split('_')
                excel_name_x = f'{y[1]}_{y[2]}'
                excel_name = excel_name_x.split('.')[0]
                if dir.find(excel_name) != -1:
                    #df1 = pd.read_excel(f'{our_month}_files/{excel}')
                    #print(df1)
                    read_months.append(our_month)
                    workbook = py.load_workbook(f"{storage}/Elster/historical/{excel_name}_files/{excel}")
                    sheet = workbook['sheet1']
                    man_value = sheet['B1'].value
                    manu = man_value.split()[0]  # meter manufacturer
                    # print(manu)
                    input_string = str(sheet["B2"].value)
                    meter = input_string.split()[0].replace('-', '')  # meter number
                    # print(meter)
                    read_datetime = f'{input_string.split()[1]} {input_string.split()[2]}'
                    date_pattern = '%{MONTHNUM:month}/%{MONTHDAY:day}/%{YEAR:year}'
                    grok = Grok(date_pattern)
                    xl = (grok.match(input_string))
                    mm = xl['month']
                    dd = xl['day']
                    yr = xl['year']
                    x = datetime(int(yr), int(mm), int(dd))
                    # ll = (x.strftime("%d-%m-%Y"))
                    D_O_R = (x.strftime("%d-%b-%Y"))  # reading date
                    list3, list4, list5, list6, list7, list8, list9, list10, list11, list12 = [], [], [], [], [], [], [], [], [], []
                    list13, list14, list15, list16, list17, list18, list20, list20, list21, list22 = [], [], [], [], [], [], [], [], [], []
                    list23, list24, list25, list26, list27, list28, list29, list30, list31, list32 = [], [], [], [], [], [], [], [], [], []
                    list1a, list33, list34 = [], [], []

                    for r in range(1, sheet.max_row):
                        for x in range(1, sheet.max_column):
                            cell = sheet.cell(r, x)
                            #print(sheet.max_row)
                            if cell.value == 'Cumulative totals':
                                cell_imp_whr = sheet.cell(r + 3, x + 1)
                                cell_unit = sheet.cell(r + 3, x + 2)
                                cell_expwh_unit = sheet.cell(r + 4, x + 2)
                                #imp_wh = ''
                                if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                                    imp_whr = cell_imp_whr.value.replace(',', '')
                                    imp_whh = round(float(imp_whr) / 1000000, 3)
                                    imp_wh = "{:,}".format(imp_whh)
                                    # print(float(imp_wh))
                                #exp_wh = ''
                                cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                                if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                                    exp_whr = cell_exp_whr.value.replace(',', '')
                                    exp_whh = round(float(exp_whr) / 1000000, 3)
                                    exp_wh = "{:,}".format(exp_whh)

                                #app_pow = ''
                                cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                                cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                                if cell_app_pow_unit.value == 'VAh':
                                    app_powr = cell_app_pow.value.replace(',', '')
                                    app_powh = round(float(app_powr) / 1000000, 3)
                                    app_pow = "{:,}".format(app_powh)
                                    # app_pow = float(app_powr)

                                # getting total import and export MVarh
                                #total_imp_MVArh = ''

                                unit1_MVArh = sheet.cell(r + 5, x + 2)
                                unit3_MVArh = sheet.cell(r + 7, x + 2)
                                if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                                    imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                    imp_MVArh1 = float(imp_MVArh1_r)
                                    imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                                    imp_MVArh2 = float(imp_MVArh2_r)
                                    total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                                    total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                                #total_exp_MVArh = ''
                                #total_MVArh = ''
                                unit2_MVArh = sheet.cell(r + 6, x + 2)
                                unit4_MVArh = sheet.cell(r + 8, x + 2)
                                if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting import MVArh. We add Q3 & Q4
                                    exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                    exp_MVArh1 = float(exp_MVArh1_r)
                                    exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                                    exp_MVArh2 = float(exp_MVArh2_r)
                                    total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                                    total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                                    print(total_exp_MVArh)

                                #total MVArh

                                    total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(total_imp_MVArh.replace(',', ''))
                                    total_MVArh_xx = round(total_MVArh_x, 3)
                                    total_mvarh = "{:,}".format(total_MVArh_xx)

                            #rate1,rate2,rate1,rate3,rate4,rate6 = '','','','','',''
                            if cell.value == 'Register' and sheet.cell(r-2,x).value == 'Rates' and sheet.cell(r+1,x+2).value == 'Import Wh':  # getting the various rates
                                unit1 = sheet.cell(r + 1, x + 2)
                                unit2 = sheet.cell(r + 2, x + 2)
                                unit3 = sheet.cell(r + 3, x + 2)
                                if unit3.value == unit2.value == 'Import Wh':
                                    rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                                    rate1_rh = round(float(rate1_r) / 1000000, 3)
                                    rate1 = "{:,}".format(rate1_rh)

                                if unit1.value == unit3.value == 'Import Wh':
                                    rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                                    rate2_rh = round(float(rate2_r) / 1000000, 3)
                                    rate2 = "{:,}".format(rate2_rh)

                                if unit2.value == unit1.value == 'Import Wh':
                                    rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                                    rate3_rh = round(float(rate3_r) / 1000000, 3)
                                    rate3 = "{:,}".format(rate3_rh)

                                unit4 = sheet.cell(r + 4, x + 2)
                                unit5 = sheet.cell(r + 5, x + 2)
                                unit6 = sheet.cell(r + 6, x + 2)
                                if unit5.value == unit6.value == 'Export Wh':
                                    rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                                    rate4_rh = round(float(rate4_r) / 1000000, 3)
                                    rate4 = "{:,}".format(rate4_rh)

                                if unit4.value == unit6.value == 'Export Wh':
                                    rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                    rate5_rh = round(float(rate5_r) / 1000000, 3)
                                    rate5 = "{:,}".format(rate5_rh)

                                if unit4.value == unit5.value == 'Export Wh':
                                    rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                    rate6_rh = round(float(rate6_r) / 1000000, 3)
                                    rate6 = "{:,}".format(rate6_rh)

                            if cell.value == 'Billing event details':
                                req_cell = sheet.cell(r + 2, x)
                                req_cell1 = sheet.cell(r + 3, x)
                                if req_cell.value == 'Billing reset number:':
                                    resets = sheet.cell(r + 2, x + 1).value
                                    # no_of_resets = float(no_of_resets_r)

                                if req_cell1.value == 'Time of billing reset:':
                                    reset_time = sheet.cell(r + 3, x + 1).value
                            #global sales_month

                            #sales_month = 0
                            if cell.value == 'Time of billing reset:' and sheet.cell(r + 1,x).value == 'Billing period end date':
                                #M_O_S_r = sheet.cell(r + 1, x + 1).value
                                M_O_S_r = sheet.cell(r, x + 1).value
                                #print(M_O_S_r)
                                M_O_S_x = M_O_S_r.split('/')
                                M_O_S_i = int(M_O_S_x[1])
                                sales_month = rate5
                                #sales_month = f'{calendar.month_name[M_O_S_i]}-{int(M_O_S_x[2])}'


                            #max_dem_1,max_dem_2,max_dem_3 = '','',''
                            pow_down_count, pow_down_dt, prog_count, prog_count_dt = '', '', '', ''
                            if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,x).value == 'Register':  # getting maximum demand
                                unit_1 = sheet.cell(r + 3, x + 2)
                                unit_2 = sheet.cell(r + 4, x + 2)
                                unit_3 = sheet.cell(r + 5, x + 2)
                                if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                                    max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                                    max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 4)
                                    max_dem_1 = "{:,}".format(max_dem_1_rh)

                                    max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                                    max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 4)
                                    max_dem_2 = "{:,}".format(max_dem_2_rh)

                                    max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                                    max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 4)
                                    max_dem_3 = "{:,}".format(max_dem_3_rh)

                            #max_dem_1t, max_dem_2t, max_dem_3t, = '','',''
                            if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                                    r + 2,
                                    x + 3).value == 'Time and date':  # getting max-demands time
                                unit_1t = sheet.cell(r + 3, x + 3)
                                unit_2t = sheet.cell(r + 4, x + 3)
                                unit_3t = sheet.cell(r + 5, x + 3)

                                max_dem_1t = unit_1t.value
                                max_dem_2t = unit_2t.value
                                max_dem_3t = unit_3t.value

                            if cell.value == 'Maximum demands' and sheet.cell(r + 2,
                                                                              x).value == 'Register' and sheet.cell(
                                    r + 2,
                                    x + 3).value != 'Time and date':
                                max_dem_1t = 'N/A'
                                max_dem_2t = 'N/A'
                                max_dem_3t = 'N/A'

                            if cell.value == 'Billing period start date' and sheet.cell(r+1,x).value == 'Billing period end date':
                                sales_value = sheet.cell(r,x+1).value
                                sales_x = sales_value.split('/')
                                month_int = int(sales_x[1])
                                this_month = calendar.month_name[month_int]
                                sales = f'{this_month} {sales_x[2]}'

                    c.execute(f"""CREATE TABLE IF NOT EXISTS {meter} (id int(5) NOT NULL AUTO_INCREMENT PRIMARY KEY,time_inserted VARCHAR(20),
                    inserted_by varchar(10),meter_read_by varchar(10),reading_datetime varchar(20),meter_no varchar(10),Energy_For varchar(20),
                    Cum_Import VARCHAR(13),Cum_Export VARCHAR(13),Apparent_Power VARCHAR(13),Rate_1 VARCHAR(13),Rate_2 VARCHAR(13),
                    Rate_3 VARCHAR(13),Rate_4 VARCHAR(13),Rate_5 VARCHAR(13),Rate_6 VARCHAR(13),Max_Dem_1 VARCHAR(13),Max_Dem1_time VARCHAR(13),
                    Max_Dem_2 VARCHAR(13),Max_Dem2_time VARCHAR(13),Max_Dem_3 VARCHAR(13),Max_Dem3_time VARCHAR(13),Import_MVArh VARCHAR(13),
                    Export_MVArh VARCHAR(13),Total_MVAh VARCHAR(13),No_of_Resets int(5),Last_Reset varchar(15),Power_Down_Count varchar(5),
                    Lst_pwr_dwn_date_and_time varchar(20),prog_count int(5),last_prog_date varchar(20))""")

                    sql = f"""INSERT INTO {meter} (time_inserted,inserted_by,meter_read_by,reading_datetime,meter_no, Energy_For, Cum_Import, Cum_Export, Apparent_Power,Rate_1,
                    Rate_2, Rate_3, Rate_4, Rate_5, Rate_6, Max_Dem_1, Max_Dem1_time, Max_Dem_2,Max_Dem2_time, Max_Dem_3, Max_Dem3_time, Import_MVArh,
                    Export_MVArh,Total_MVAh,No_of_Resets,Last_Reset,Power_Down_Count,Lst_pwr_dwn_date_and_time,prog_count,last_prog_date)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
                    val = (unix, active_user, active_user, read_datetime, meter,sales, imp_wh, exp_wh, app_pow, rate1, rate2,
                    rate3, rate4, rate5,rate6, max_dem_1, max_dem_1t, max_dem_2, max_dem_2t, max_dem_3, max_dem_3t, total_imp_MVArh, total_exp_MVArh,
                    total_mvarh, resets, reset_time, pow_down_count, pow_down_dt, prog_count, prog_count_dt)

                    c.execute(sql,val)
                    conn.commit()

    return render(request, 'form_success.html')
    #c.close()
    conn.close()

def submission(request): #iNSERTING NONTHLY RECORD
    if request.method == 'POST':
        global active_user
        unix = datetime.now().replace (microsecond=0)
        #date = str(datetime.datetime.fromtimestamp(unix).strftime('%Y-%m-%d %H:%M:%S'))
        #substation = request.POST['subst']
        M_O_S = request.POST['month_of_sale']
        R_D = request.POST['reading_date']
        R_T = request.POST['reading_time']
        meter_no = request.POST['main_meter_no']
        cum_imp = request.POST['cum_imp_m'].replace(',','')
        cum_exp = request.POST['cum_exp_m'].replace(',','')
        app_pow = request.POST['app_pow_m'].replace(',','')
        rate1 = request.POST['rate1_m'].replace(',','')
        rate2 = request.POST['rate2_m'].replace(',','')
        rate3 = request.POST['rate3_m'].replace(',','')
        rate4 = request.POST['rate4_m'].replace(',','')
        rate5 = request.POST['rate5_m'].replace(',','')
        rate6 = request.POST['rate6_m'].replace(',','')
        max_dem1 = request.POST['max_dem_m']
        max_dem1_time = request.POST['max_dem_t_m']
        max_dem2 = request.POST['max_dem2_m']
        max_dem2_time = request.POST['max_dem2_t_m']
        max_dem3 = request.POST['max_dem3_m']
        max_dem3_time = request.POST['max_dem3_t_m']
        imp_mvarh = request.POST['imp_mvarh_m'].replace(',','')
        exp_mvarh = request.POST['exp_mvarh_m'].replace(',','')
        total_mvarh = request.POST['total_mvarh_m'].replace(',','')
        resets = request.POST['resets_m']
        last_reset = request.POST['last_reset_m']
        pow_down_count = request.POST['pow_down_count_m']
        pow_down_dt = request.POST['pow_down_dt_m']
        prog_count = request.POST['prog_count_m']
        prog_count_dt = request.POST['prog_count_dt_m']
        vt_ratio = request.POST['vt_ratio_m']
        ct_ratio = request.POST['ct_ratio_m']

        #date for check meter
        meter_no_ch = request.POST['check_meter_no']
        cum_imp_ch = request.POST['cum_imp_c'].replace(',', '')
        cum_exp_ch = request.POST['cum_exp_c'].replace(',', '')
        app_pow_ch = request.POST['app_pow_c'].replace(',', '')
        rate1_ch = request.POST['rate1_c'].replace(',', '')
        rate2_ch = request.POST['rate2_c'].replace(',', '')
        rate3_ch = request.POST['rate3_c'].replace(',', '')
        rate4_ch = request.POST['rate4_c'].replace(',', '')
        rate5_ch = request.POST['rate5_c'].replace(',', '')
        rate6_ch = request.POST['rate6_c'].replace(',', '')
        max_dem1_ch = request.POST['max_dem_c']
        max_dem1_time_ch = request.POST['max_dem_t_c']
        max_dem2_ch = request.POST['max_dem2_c']
        max_dem2_time_ch = request.POST['max_dem2_t_c']
        max_dem3_ch = request.POST['max_dem3_c']
        max_dem3_time_ch = request.POST['max_dem3_t_c']
        imp_mvarh_ch = request.POST['imp_mvarh_c'].replace(',', '')
        exp_mvarh_ch = request.POST['exp_mvarh_c'].replace(',', '')
        total_mvarh_ch = request.POST['total_mvarh_c'].replace(',', '')
        resets_ch = request.POST['resets_c']
        last_reset_ch = request.POST['last_reset_c']
        pow_down_count_ch = request.POST['pow_down_count_c']
        pow_down_dt_ch = request.POST['pow_down_dt_c']
        prog_count_ch = request.POST['prog_count_c']
        prog_count_dt_ch = request.POST['prog_count_dt_c']
        vt_ratio_ch = request.POST['vt_ratio_c']
        ct_ratio_ch = request.POST['ct_ratio_c']
        #current_user = request.user

        #active_user = user_act
        #active_user = read_by
        #inserted_by = active_user
        read_time= request.POST['reading_time']
        read_date = request.POST['reading_date']
        #id = 0

        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                              user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        sql = """INSERT INTO lgg_main(time_stamp,inserted_by,meter_read_by,reading_date,reading_time,meter_no, Energy_For, Cum_Import, Cum_Export, Apparent_Power,Rate_1,
        Rate_2, Rate_3, Rate_4, Rate_5, Rate_6, Max_Dem_1, Max_Dem1_time, Max_Dem_2,Max_Dem2_time, Max_Dem3, Max_Dem3_time, Import_MVArh,
        Export_MVArh,Total_MVAh,No_of_Resets,Last_Reset,Power_Down_Count,Lst_pwr_dwn_date_and_time,prog_count,last_prog_date,VT_Ratio,CT_Ratio,
        meter_no_ch,Cum_Import_ch, Cum_Export_ch, Apparent_Power_ch,Rate_1_ch,
        Rate_2_ch, Rate_3_ch, Rate_4_ch, Rate_5_ch, Rate_6_ch, Max_Dem_1_ch, Max_Dem1_time_ch, Max_Dem_2_ch,Max_Dem2_time_ch, Max_Dem3_ch,
         Max_Dem3_time_ch, Import_MVArh_ch,Export_MVArh_ch,Total_MVAh_ch,No_of_Resets_ch,Last_Reset_ch,
        Power_Down_Count_ch,Lst_pwr_dwn_date_and_time_ch,prog_count_ch,last_prog_date_ch,VT_Ratio_ch,CT_Ratio_ch)
         VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s,
         %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s)"""
        #sql2 = 'read username FROM auth_user'

        val = (unix, active_user, active_user,read_date,read_time, meter_no, M_O_S, cum_imp, cum_exp, app_pow, rate1, rate2, rate3, rate4, rate5,
               rate6, max_dem1, max_dem1_time, max_dem2, max_dem2_time, max_dem3, max_dem3_time, imp_mvarh, exp_mvarh,
               total_mvarh, resets, last_reset, pow_down_count, pow_down_dt, prog_count, prog_count_dt, vt_ratio, ct_ratio,
               meter_no_ch,cum_imp_ch,cum_exp_ch,app_pow_ch,rate1_ch,rate2_ch,rate3_ch,rate4_ch,rate5_ch,rate6_ch,max_dem1_ch,max_dem1_time_ch,
               max_dem2_ch,max_dem2_time_ch,max_dem3_ch,max_dem3_time_ch,imp_mvarh_ch,exp_mvarh_ch,total_mvarh_ch,resets_ch,last_reset_ch,
               pow_down_count_ch,pow_down_dt_ch,prog_count_ch,prog_count_dt_ch,vt_ratio_ch,ct_ratio_ch)

        c.execute(sql,val)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')
#global df1,df2,df3,df_ia,df_ib,df_ic,df_va,df_vb,df_vc,df_pw,df_vars

def ipp(request):
    return render(request,'ipp.html')

def ipp_insert(request):
    if request.method == 'POST':
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        unix = datetime.now().replace (microsecond=0)
        time_added = unix
        name_ipp = request.POST['ipp_name']
        capacity = request.POST['install_cap']
        gen_type = request.POST['gen_type']
        ipp_district = request.POST['ipp_dist']
        ipp_region = request.POST['ipp_dist']
        x_cordinates = request.POST['ipp_cord_x']
        y_cordinates = request.POST['ipp_cord_y']

        feeder_name = request.POST['ipp_feeder'] #second take
        feeder_voltage = request.POST['ipp_feeder_voltage']
        conn_substation = request.POST['ipp_conn']
        distributor = request.POST['distributor']
        comments = request.POST['comments']


        ct_manufacturer = request.POST['ipp_ct_manu']  # Third take
        ct_type = request.POST['ipp_ct_type']
        no_ct_cores = request.POST['ipp_ct_cores']
        ct_ratios = request.POST['ipp_ct_core_ratios']
        core_used = request.POST['ipp_core_used']
        accuracy_class = request.POST['ipp_accuracy_class']
        avail_spares = request.POST['ipp_avail_spares']
        spares_class = request.POST['ipp_spares_class']
        VT_manufacturer = request.POST['ipp_vt_manu']
        VT_type = request.POST['ipp_VT_type']
        VT_ratio = request.POST['ipp_vt_ratio']
        VT_accuracy = request.POST['ipp_vt_accuracy']
        meter_manufacturer = request.POST['ipp_meter_manu']  # fourth take
        meter_type = request.POST['ipp_meter_type']
        meter_no = request.POST['ipp_meter_no']
        meter_YOM = request.POST['ipp_YOM']
        meter_accuracy_class = request.POST['ipp_meter_acc_class']
        no_of_elements = request.POST['ipp_meter_elements']
        wired_as = request.POST['ipp_meter_wire']
        meter_install_date = request.POST['ipp_meter_install_date']
        meter_decom_date = request.POST['ipp_meter_decom_date']
        access = request.POST['ipp_meter_access']
        IP_Address = request.POST['ipp_ip_address']
        avail_interfaces = request.POST['ipp_meter_interfaces']
        comm_protocols = request.POST['ipp_meter_comm_protocols']
        protocol_used = request.POST['ipp_meter_protocol_used']

        meter_manufacturer_ch = request.POST['ipp_meter_manu_ch']  # fifth take
        meter_type_ch = request.POST['ipp_meter_type_ch']
        meter_no_ch = request.POST['ipp_meter_no_ch']
        meter_YOM_ch = request.POST['ipp_YOM_ch']
        meter_accuracy_class_ch = request.POST['ipp_meter_acc_class_ch']
        no_of_elements_ch = request.POST['ipp_meter_elements_ch']
        wired_as_ch = request.POST['ipp_meter_wire_ch']
        meter_install_date_ch = request.POST['ipp_meter_install_date_ch']
        meter_decom_date_ch = request.POST['ipp_meter_decom_date_ch']
        access_ch = request.POST['ipp_meter_access_ch']
        IP_Address_ch = request.POST['ipp_ip_address_ch']
        avail_interfaces_ch = request.POST['ipp_meter_interfaces_ch']
        comm_protocols_ch = request.POST['ipp_meter_comm_protocols_ch']
        protocol_used_ch = request.POST['ipp_meter_protocol_used_ch']

        global active_user
        added_by = active_user

        workbook = xlsxwriter.Workbook(
            f'ipp_meters/{meter_no}_{name_ipp}_METERING_POINT_DETAILS.xlsx')  # write to excel
        worksheet = workbook.add_worksheet()

        bold = workbook.add_format({'bold': True, 'font_size': 16, 'align': 'center'})
        bold2 = workbook.add_format({'bold': True, 'font_size': 14, 'align': 'center'})
        bold1 = workbook.add_format({'bold': True, 'font_size': 13})
        worksheet.merge_range('A1:D1', 'Merged Range', bold)
        worksheet.merge_range('A2:D2', 'Merged Range', bold2)

        worksheet.write('A1', f'{name_ipp}_{meter_no}_DETAILS', bold)
        worksheet.write('A2', 'LOCATION DETAILS', bold2)
        worksheet.write('A3', 'Time Added', bold1)
        worksheet.write('B3', str(unix))
        worksheet.write('C3', 'Added By', bold1)
        worksheet.write('D3', active_user)
        worksheet.write('A4', 'Name of IPP', bold1)
        worksheet.write('B4', name_ipp)
        worksheet.write('C4', 'Installed Capacity', bold1)
        worksheet.write('D4', capacity)
        worksheet.write('A5', 'Generation Type', bold1)
        worksheet.write('B5', gen_type)
        worksheet.write('C5', 'District', bold1)
        worksheet.write('D5', ipp_district)
        worksheet.write('A6', 'Region', bold1)
        worksheet.write('B6', ipp_region)
        worksheet.write('C6', 'X Cordinates', bold1)
        worksheet.write('D6', x_cordinates)
        worksheet.write('A7', 'Y Cordinates', bold1)
        worksheet.write('B7', y_cordinates)

        worksheet.merge_range('A9:D9', 'FEEDER DETAILS AND DISTRIBUTOR', bold2)
        #worksheet.write('A8', 'FEEDER DETAILS AND DISTRIBUTOR', bold2)
        worksheet.write('A10', 'Feeder Name', bold1)
        worksheet.write('B10', feeder_name)
        worksheet.write('C10', 'Feeder Voltage', bold1)
        worksheet.write('D10', feeder_voltage)
        worksheet.write('A11', 'Connecting Substation', bold1)
        worksheet.write('B11', conn_substation)
        worksheet.write('C11', 'Distributor', bold1)
        worksheet.write('D11', distributor)
        worksheet.write('A12', 'Comments', bold1)
        worksheet.write('B12', comments)

        worksheet.merge_range('A14:D14', 'CT & VT DETAILS', bold2)
        #worksheet.write('A13', 'CT & VT DETAILS', bold2)
        worksheet.write('A15', 'CT Manufacturer', bold1)
        worksheet.write('B15', ct_manufacturer)
        worksheet.write('C15', 'CT Type', bold1)
        worksheet.write('D15', ct_type)
        worksheet.write('A16', 'No of Metering CT cores', bold1)
        worksheet.write('B16', no_ct_cores)
        worksheet.write('C16', 'CT Ratios', bold1)
        worksheet.write('D16', ct_ratios)
        worksheet.write('A17', 'Core Used', bold1)
        worksheet.write('B17', core_used)
        worksheet.write('C17', 'Accuracy Class', bold1)
        worksheet.write('D17', accuracy_class)
        worksheet.write('A18', 'Available Spares', bold1)
        worksheet.write('B18', avail_spares)
        worksheet.write('C18', 'Spares Class', bold1)
        worksheet.write('D18', spares_class)
        worksheet.write('A19', 'VT Manufacturer', bold1)
        worksheet.write('B19', VT_manufacturer)
        worksheet.write('C19', 'VT Type', bold1)
        worksheet.write('D19', VT_type)
        worksheet.write('A20', 'VT Ratio', bold1)
        worksheet.write('B20', VT_ratio)
        worksheet.write('C20', 'VT Accuracy', bold1)
        worksheet.write('D20', VT_accuracy)

        worksheet.merge_range('A22:D22', 'MAIN METER DETAILS', bold)
        # worksheet.write('A21', "MAIN METER DETAILS", bold)
        worksheet.write('A23', 'Meter Manufacturer', bold1)
        worksheet.write('B23', meter_manufacturer)
        worksheet.write('C23', 'Meter Type', bold1)
        worksheet.write('D23', meter_type)
        worksheet.write('A24', 'Meter No', bold1)
        worksheet.write('B24', meter_no)
        worksheet.write('C24', 'Meter Y.O.M', bold1)
        worksheet.write('D24', meter_YOM)
        worksheet.write('A25', 'Meter Accuracy', bold1)
        worksheet.write('B25', meter_accuracy_class)
        worksheet.write('C25', 'No. of Elements', bold1)
        worksheet.write('D25', no_of_elements)
        worksheet.write('A26', 'Wired AS', bold1)
        worksheet.write('B26', wired_as)
        worksheet.write('C26', 'Installation Date', bold1)
        worksheet.write('D26', meter_install_date)
        worksheet.write('A27', 'Decommisioning Date', bold1)
        worksheet.write('B27', meter_decom_date)
        worksheet.write('C27', 'Access', bold1)
        worksheet.write('D27', access)
        worksheet.write('A28', 'IP Address', bold1)
        worksheet.write('B28', IP_Address)
        worksheet.write('C28', 'Avail. Interfaces', bold1)
        worksheet.write('D28', avail_interfaces)
        worksheet.write('A29', 'Comm Protocol', bold1)
        worksheet.write('B29', comm_protocols)
        worksheet.write('C29', 'Comm Protocol Used', bold1)
        worksheet.write('D29', protocol_used)

        worksheet.merge_range('A31:D31', 'CHECK METER DETAILS', bold)
        #worksheet.write('A30', "CHECK METER DETAILS", bold)
        worksheet.write('A32', 'Meter Manufacturer', bold1)
        worksheet.write('B32', meter_manufacturer_ch)
        worksheet.write('C32', 'Meter Type', bold1)
        worksheet.write('D32', meter_type_ch)
        worksheet.write('A33', 'Meter No', bold1)
        worksheet.write('B33', meter_no_ch)
        worksheet.write('C33', 'Meter Y.O.M', bold1)
        worksheet.write('D33', meter_YOM_ch)
        worksheet.write('A34', 'Meter Accuracy', bold1)
        worksheet.write('B34', meter_accuracy_class_ch)
        worksheet.write('C34', 'No. of Elements', bold1)
        worksheet.write('D34', no_of_elements_ch)
        worksheet.write('A35', 'Wired AS', bold1)
        worksheet.write('B35', wired_as_ch)
        worksheet.write('C35', 'Installation Date', bold1)
        worksheet.write('D35', meter_install_date_ch)
        worksheet.write('A36', 'Decommisioning Date', bold1)
        worksheet.write('B36', meter_decom_date_ch)
        worksheet.write('C36', 'Access', bold1)
        worksheet.write('D36', access_ch)
        worksheet.write('A37', 'IP Address', bold1)
        worksheet.write('B37', IP_Address_ch)
        worksheet.write('C37', 'Avail. Interfaces', bold1)
        worksheet.write('D37', avail_interfaces_ch)
        worksheet.write('A38', 'Comm Protocol', bold1)
        worksheet.write('B38', comm_protocols_ch)
        worksheet.write('C38', 'Comm Protocol Used', bold1)
        worksheet.write('D38', protocol_used_ch)

        workbook.close()

        c = conn.cursor()
        sql_2 = """INSERT INTO ipp_meters (time_added,added_by,name_ipp,capacity,gen_type, ipp_district,ipp_region,x_cordinates,y_cordinates,feeder_name,feeder_voltage,conn_substation,
                           distributor,comments,ct_manufacturer,ct_type,no_ct_cores,ct_ratios,core_used,accuracy_class,avail_spares,spares_class,VT_manufacturer,VT_type,VT_ratio,VT_accuracy,
                           meter_manufacturer,meter_type,meter_no,meter_YOM,meter_accuracy_class,no_of_elements,wired_as,meter_install_date,meter_decom_date,access,IP_Address,avail_interfaces,comm_protocols,protocol_used,
                           meter_manufacturer_ch,meter_type_ch,meter_no_ch,meter_YOM_ch,meter_accuracy_class_ch,no_of_elements_ch,wired_as_ch,meter_install_date_ch,meter_decom_date_ch,access_ch,IP_Address_ch,avail_interfaces_ch,comm_protocols_ch,protocol_used_ch) 
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                            %s,%s, %s, %s, %s, %s,%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        values = (
        unix, added_by, name_ipp, capacity, gen_type, ipp_district,ipp_region, x_cordinates, y_cordinates, feeder_name, feeder_voltage,
        conn_substation, distributor, comments,
        ct_manufacturer, ct_type, no_ct_cores, ct_ratios, core_used, accuracy_class, avail_spares, spares_class,
        VT_manufacturer, VT_type, VT_ratio, VT_accuracy,
        meter_manufacturer, meter_type, meter_no, meter_YOM, meter_accuracy_class, no_of_elements, wired_as,
        meter_install_date, meter_decom_date, access, IP_Address, avail_interfaces, comm_protocols, protocol_used,
        meter_manufacturer_ch, meter_type_ch, meter_no_ch, meter_YOM_ch, meter_accuracy_class_ch, no_of_elements_ch,
        wired_as_ch, meter_install_date_ch, meter_decom_date_ch, access_ch, IP_Address_ch, avail_interfaces_ch,
        comm_protocols_ch, protocol_used_ch)

        c.execute(sql_2, values)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')




def LP_plot(request): #Plotting Load profiles
    #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
    global all_main_meters
    all_main_meters = []
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                          user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    c = conn.cursor()
    c.execute("SELECT name_ipp FROM ipp_meters")
    ipp_data = c.fetchall()
    ipp_listall = []
    for ipp in ipp_data:
        #print(ipp)
        ippx = ipp[0]
        ipp_listall.append(ippx)
    #print(ipp_listall)
    final_ipp_listall = list(dict.fromkeys(ipp_listall))

    c.execute("SELECT substation FROM substation_meters")
    sub_data = c.fetchall()
    #print(sub_data)
    sub_listall = []
    for sub in sub_data:
        subx = sub[0]
        sub_listall.append(subx)
    final_sub_listall = list(dict.fromkeys(sub_listall))
    print(final_sub_listall)

    c.execute("SELECT district FROM standalone_meters")
    stand_data = c.fetchall()
    # print(stand_data)
    stand_listall = []
    for stand in stand_data:
        standx = stand[0]
        stand_listall.append(standx)
    final_stand_listall = list(dict.fromkeys(stand_listall))
    print(final_stand_listall)

    sql = "SELECT start_time, PW FROM uetcl0103 WHERE month = 'September' AND date = '22/09/2020'"
    tables = ['standalone_meters', 'substation_meters', 'ipp_meters']
    for table in tables:
        sql = f"select meter_no from {table}"
        c.execute(sql)
        data_all = c.fetchall()
        for data in data_all:
            data1 = data[0]
            all_main_meters.append(data1)

    sub_meters = []
    sql1 = "select meter_no from substation_meters"
    c.execute(sql1)
    data_all = c.fetchall()
    for data in data_all:
        data1 = data[0]
        sub_meters.append(data1)

    standalone = []
    sql2 = "select meter_no from standalone_meters"
    c.execute(sql2)
    data_all = c.fetchall()
    for data in data_all:
        data1 = data[0]
        standalone.append(data1)

    ipp_meter = []
    sql3 = "select meter_no from standalone_meters"
    c.execute(sql3)
    data_all = c.fetchall()
    for data in data_all:
        data1 = data[0]
        ipp_meter.append(data1)

    conn.commit()
    c.close()
    conn.close()
    month_list = []
    for month in calendar.month_name:
        month_list.append(month)
    month_list.pop(0)

    year_list = []
    unix = datetime.now().year
    for i in range(0, 3):
        yearr = unix - i
        year_list.append(yearr)

    return render(request, 'load_profiles.html',{'year_list':year_list,'month_list':month_list,'ipp_meter':ipp_meter,
    'standalone':standalone,'all_main_meters':all_main_meters,'sub_meters':sub_meters,'ipp_listall':final_ipp_listall,
    'final_sub_listall':final_sub_listall,'final_stand_listall':final_stand_listall}) #end of function



def umeme(request):
    #global meters_umeme
    meters_umeme = []
    #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    c = conn.cursor()
    sql = "SELECT meter_no FROM substation_meters WHERE distributor = 'umeme'"
    c.execute(sql)
    data = c.fetchall()
    for row in data:
        meters_umeme.append(row[0])
    #print(meters)
    conn.commit()
    c.close()
    conn.close()
    month_list, year_list = [], []
    for i in range(1, 13):
        print(calendar.month_name[i])
        month = calendar.month_name[i]
        month_list.append(month)
    #print(month_list)

    year = datetime.now().year
    for x in range(0, 3):
        new_year = year - x
        year_list.append(new_year)
    #print(year_list)
    # year = date.today().year
    # xxl = date.today().month - 1
    # # print(now)
    # month = calendar.month_name[xxl]
    # bill_date = f'{month} {year}'

    return render(request, 'gen_bill.html',{'meters_umeme':meters_umeme,'month_list':month_list,'year_list':year_list})

def bill_gen(request):
    if request.method == 'POST':

        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                             user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        meter_no = request.POST['meter_no']
        month = request.POST['month']
        year = request.POST['year']
        month_year = f"{month} {year}"
        date1 = datetime.strptime(month_year, '%B %Y')
        a_month = dateutil.relativedelta.relativedelta(months=1)
        b_month = dateutil.relativedelta.relativedelta(months=2)
        a_year = dateutil.relativedelta.relativedelta(years=1)

        bill_month_f = date1 - a_month
        prev_month = date1 - a_month
        prev_year = date1 - a_year   #SAME MONTH IN THE PREVIOUS YEAR
        bill_month = date1.strftime('%B %Y')
        prev_b_month = prev_month.strftime('%B %Y')
        prev_b_year = prev_year.strftime('%B %Y')
        # print(date1.strftime('%B %Y'))
        # print(prev_month.strftime('%B %Y'))
        # print(prev_year.strftime('%B %Y'))
        c = conn.cursor()
        data_line = []
        energy_list = [bill_month,prev_b_month,prev_b_year]
        for item in energy_list:
            c.execute(f"""SELECT reading_datetime,meter_no,Energy_For,Cum_Import,Cum_Export,Apparent_Power,Rate_1,Rate_2,Rate_3,Rate_4,Rate_5,Rate_6,No_of_Resets 
            FROM {meter_no}_monthly WHERE Energy_For = %s""",(item))
            data = c.fetchone()
            data_line.append(data)

        #data = list(datax)
        #print(data_line)
        bill_list = data_line[0]
        diff = data_line[0][3]-data_line[1][3]
        change_1 = (diff/data_line[1][3])*100
        change_1_f = round(change_1, 2)
        diff2 = data_line[0][3]- data_line[2][3]
        change_2 = (diff2/ data_line[2][3])*100
        change_2_f = round(change_2,2)
        conn.commit()
        conn.close()

        return render(request,'bill.html',{'data_line':data_line,'bill_list':bill_list,'change_1_f':change_1_f,
                                           'change_2_f':change_2_f})


def newmeter(request):
    return render(request,'new_meter_home.html')

def standalone(request):
    return render(request,'standalone_meter.html')

def substation(request):
    return render(request,'substation_meter.html')



def querries(request):
    return render(request, 'querries.html')

def querriesback(request):
    return render(request, 'welcome.html',{'user':user})



def bill_sub(request):
    return render(request, 'bill_success.html')


def hist_records(request): #inserting historical data
    if request.method == 'POST':
        global active_user
        uploaded_file = request.FILES['hist']
        fs = FileSystemStorage()
        file_m = fs.save(uploaded_file.name, uploaded_file)
        wb = Workbook()
        ws = wb.active
        with open(f'media/{file_m}', 'r') as f:
            for row in csv.reader(f):
                ws.append(row)
        wb.save(f'historical_records/{file_m}.xlsx')
        workbook = py.load_workbook(f'historical_records/{file_m}.xlsx')
        sheet = workbook['Sheet']
        man_value = sheet['B1'].value
        manu = man_value.split()[0]  # meter manufacturer
        # print(manu)
        input_string = str(sheet["B2"].value)
        # print(input_string)
        # global meter
        meter_no = input_string.split()[0].replace('-', '')  # meter number
        # print(meter)
        read_time = f'{input_string.split()[1]} {input_string.split()[2]}'  # reading time

        # global meter_no
        # sql_1 = "SELECT CTratio, VTratio FROM meter_details WHERE meter_no=?",((meter,))
        # c.execute("SELECT CTratios, VTratio FROM standalone_meter_details WHERE meter_no = %s", meter)
        # data = c.fetchone()

        date_pattern = '%{MONTHNUM:month}/%{MONTHDAY:day}/%{YEAR:year}'
        grok = Grok(date_pattern)
        xl = (grok.match(input_string))
        mm = xl['month']
        dd = xl['day']
        yr = xl['year']
        x = datetime(int(yr), int(mm), int(dd))
        # ll = (x.strftime("%d-%m-%Y"))
        D_O_R = (x.strftime("%d-%b-%Y"))  # reading date
        list1 = []
        list2 = []
        list7 = []
        for m in range(1, sheet.max_row + 1):
            for n in range(1, sheet.max_column + 1):
                # cell1 = sheet.cell(m,n)
                cell = sheet.cell(m, n).value
                cell_value = str(cell)
                if (cell_value.find('Historical data set') != -1):
                    list1.append(cell_value)
                    list2.append(m)
        #print(len(list1))
        #print(list2)
        list3 = []
        list4 = []
        list5 = []
        list6 = []
        list7 = []
        list8 = []
        list9 = []
        list10 = []
        list11 = []
        list12 = []
        list13 = []
        list14 = []
        list15 = []
        list16 = []
        list17 = []
        list18 = []
        list20 = []
        list20 = []
        list21 = []
        list22 = []
        list23 = []
        list24 = []
        list25 = []
        list26 = []
        list27 = []
        list28 = []
        list29 = []
        list30 = []
        list31 = []
        list32 = []
        list1a = []
        list33 = []
        list34 = []

        for i in range(1, len(list2) + 1):
            if i <= len(list2) - 1:
                row_1 = list2[i - 1]
                row_2 = list2[i]
                # print(row_2)
                for r in range(row_1, row_2 + 1):
                    for x in range(1, sheet.max_column + 1):
                        cell = sheet.cell(r, x)
                        # global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                        if cell.value == 'Cumulative totals':
                            cell_imp_whr = sheet.cell(r + 3, x + 1)
                            cell_unit = sheet.cell(r + 3, x + 2)
                            cell_expwh_unit = sheet.cell(r + 4, x + 2)
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                                imp_whr = cell_imp_whr.value.replace(',', '')
                                imp_whh = round(float(imp_whr) / 1000000, 3)
                                imp_wh = "{:,}".format(imp_whh)
                                list1a.append(imp_wh)

                            cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                                exp_whr = cell_exp_whr.value.replace(',', '')
                                exp_whh = round(float(exp_whr) / 1000000, 3)
                                exp_wh = "{:,}".format(exp_whh)
                                list4.append(exp_wh)

                            cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                            cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                            if cell_app_pow_unit.value == 'VAh':
                                app_powr = cell_app_pow.value.replace(',', '')
                                app_powh = round(float(app_powr) / 1000000, 3)
                                app_pow = "{:,}".format(app_powh)
                                list5.append(app_pow)
                                # app_pow = float(app_powr)

                            # getting total import and export MVarh
                            unit1_MVArh = sheet.cell(r + 5, x + 2)
                            unit3_MVArh = sheet.cell(r + 7, x + 2)
                            global total_imp_MVArh
                            if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                                imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                imp_MVArh1 = float(imp_MVArh1_r)
                                imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                                imp_MVArh2 = float(imp_MVArh2_r)
                                total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                                total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                                list6.append(total_imp_MVArh)

                            unit2_MVArh = sheet.cell(r + 6, x + 2)
                            unit4_MVArh = sheet.cell(r + 8, x + 2)
                            if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                                exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                exp_MVArh1 = float(exp_MVArh1_r)
                                exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                                exp_MVArh2 = float(exp_MVArh2_r)
                                total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                                total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                                list7.append(total_exp_MVArh)

                                # total MVArh
                                total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(
                                    total_imp_MVArh.replace(',', ''))
                                total_MVArh_xx = round(total_MVArh_x, 3)
                                total_MVArh = "{:,}".format(total_MVArh_xx)
                                list8.append(total_MVArh)

                        # global rate1,rate2,rate3,rate4,rate5,rate6
                        if cell.value == 'Register':  # getting the various rates
                            unit1 = sheet.cell(r + 1, x + 2)
                            unit2 = sheet.cell(r + 2, x + 2)
                            unit3 = sheet.cell(r + 3, x + 2)
                            if unit3.value == unit2.value == 'Import Wh':
                                rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                                rate1_rh = round(float(rate1_r) / 1000000, 3)
                                rate1 = "{:,}".format(rate1_rh)
                                list9.append(rate1)

                            if unit1.value == unit3.value == 'Import Wh':
                                rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                                rate2_rh = round(float(rate2_r) / 1000000, 3)
                                rate2 = "{:,}".format(rate2_rh)
                                list10.append(rate2)

                            if unit2.value == unit1.value == 'Import Wh':
                                rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                                rate3_rh = round(float(rate3_r) / 1000000, 3)
                                rate3 = "{:,}".format(rate3_rh)
                                list11.append(rate3)

                            unit4 = sheet.cell(r + 4, x + 2)
                            unit5 = sheet.cell(r + 5, x + 2)
                            unit6 = sheet.cell(r + 6, x + 2)
                            if unit5.value == unit6.value == 'Export Wh':
                                rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                                rate4_rh = round(float(rate4_r) / 1000000, 3)
                                rate4 = "{:,}".format(rate4_rh)
                                list12.append(rate4)

                            if unit4.value == unit6.value == 'Export Wh':
                                rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                rate5_rh = round(float(rate5_r) / 1000000, 3)
                                rate5 = "{:,}".format(rate5_rh)
                                list13.append(rate5)
                            # global rate6
                            if unit4.value == unit5.value == 'Export Wh':
                                rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                rate6_rh = round(float(rate6_r) / 1000000, 3)
                                rate6 = "{:,}".format(rate6_rh)
                                list14.append(rate6)
                        # global reset_time, resets
                        if cell.value == 'Billing event details':
                            req_cell = sheet.cell(r + 2, x)
                            req_cell1 = sheet.cell(r + 3, x)
                            if req_cell.value == 'Billing reset number:':
                                resets = sheet.cell(r + 2, x + 1).value
                                # no_of_resets = float(no_of_resets_r)
                                list23.append(resets)

                            if req_cell1.value == 'Time of billing reset:':
                                reset_time = sheet.cell(r + 3, x + 1).value
                                list24.append(reset_time)
                        # global M_O_S,M_O_S_r
                        if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                                    x).value == 'Billing period end date':
                            M_O_S_r = sheet.cell(r, x + 1).value
                            #print(M_O_S_r)
                            M_O_S_x = M_O_S_r.split('/')
                            M_O_S_i = int(M_O_S_x[1])
                            M_O_S = f'{calendar.month_name[M_O_S_i]} {int(M_O_S_x[2])}'
                            #print(f"THIS IS THE ENERGY_FOR: {M_O_S}")
                            list16.append(M_O_S)
                        # global max_dem_1,max_dem_2,max_dem_3
                        if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                                     x).value == 'Register':  # getting maximum demands

                            unit_1 = sheet.cell(r + 3, x + 2)
                            unit_2 = sheet.cell(r + 4, x + 2)
                            unit_3 = sheet.cell(r + 5, x + 2)
                            if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                                max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                                max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                                max_dem_1 = "{:,}".format(max_dem_1_rh)
                                list17.append(max_dem_1)

                                max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                                max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                                max_dem_2 = "{:,}".format(max_dem_2_rh)
                                list18.append(max_dem_2)

                                max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                                max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                                max_dem_3 = "{:,}".format(max_dem_3_rh)
                                list20.append(max_dem_3)
                        # global max_dem_1t, max_dem_2t, max_dem_3t

                        if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                                r + 2,
                                x + 3).value == 'Time and date':  # getting max-demands time
                            unit_1t = sheet.cell(r + 3, x + 3)
                            unit_2t = sheet.cell(r + 4, x + 3)
                            unit_3t = sheet.cell(r + 5, x + 3)

                            max_dem_1t = unit_1t.value
                            max_dem_2t = unit_2t.value
                            max_dem_3t = unit_3t.value
                            list20.append(max_dem_1t)
                            list21.append(max_dem_2t)
                            list22.append(max_dem_3t)

                            # constants
                            unix = datetime.now().replace(second=0, microsecond=0)
                            inserted_by = active_user
                            #active_user = 'kimera'
                            pow_down_count = 'No Data'
                            pow_down_dt = 'No Data'
                            prog_count = 'No Data'
                            prog_count_dt = 'No Data'
                            # list25 = ['No data']
                            list25.append(pow_down_count)
                            list26.append(pow_down_dt)
                            list27.append(prog_count)
                            list28.append(prog_count_dt)
                            list29.append(unix)
                            list30.append(inserted_by)
                            list31.append(meter_no)
                            list32.append(D_O_R)
                            list33.append(read_time)

                i += 1
            elif i == len(list2):
                row_1 = list2[i - 1]
                for r in range(row_1, sheet.max_row):
                    for x in range(1, sheet.max_column + 1):
                        cell = sheet.cell(r, x)
                        # global total_MVArh, imp_wh,exp_wh,app_pow,total_MVArh
                        if cell.value == 'Cumulative totals':
                            cell_imp_whr = sheet.cell(r + 3, x + 1)
                            cell_unit = sheet.cell(r + 3, x + 2)
                            cell_expwh_unit = sheet.cell(r + 4, x + 2)
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':  # cum import
                                imp_whr = cell_imp_whr.value.replace(',', '')
                                imp_whh = round(float(imp_whr) / 1000000, 3)
                                imp_wh = "{:,}".format(imp_whh)
                                list1a.append(imp_wh)
                            cell_exp_whr = sheet.cell(r + 4, x + 1)  # cum export
                            if cell_unit.value == 'Import Wh' and cell_expwh_unit.value == 'Export Wh':
                                exp_whr = cell_exp_whr.value.replace(',', '')
                                exp_whh = round(float(exp_whr) / 1000000, 3)
                                exp_wh = "{:,}".format(exp_whh)
                                list4.append(exp_wh)

                            cell_app_pow = sheet.cell(r + 9, x + 1)  # getting apparent power
                            cell_app_pow_unit = sheet.cell(r + 9, x + 2)
                            if cell_app_pow_unit.value == 'VAh':
                                app_powr = cell_app_pow.value.replace(',', '')
                                app_powh = round(float(app_powr) / 1000000, 3)
                                app_pow = "{:,}".format(app_powh)
                                list5.append(app_pow)
                                # app_pow = float(app_powr)

                            # getting total import and export MVarh
                            unit1_MVArh = sheet.cell(r + 5, x + 2)
                            unit3_MVArh = sheet.cell(r + 7, x + 2)
                            # global total_imp_MVArh
                            if unit1_MVArh.value == 'Q1 varh' and unit3_MVArh.value == 'Q3 varh':  # getting import MVArh. We add Q1 & Q2
                                imp_MVArh1_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                imp_MVArh1 = float(imp_MVArh1_r)
                                imp_MVArh2_r = sheet.cell(r + 7, x + 1).value.replace(',', '')
                                imp_MVArh2 = float(imp_MVArh2_r)
                                total_imp_MVArh_r = round((imp_MVArh1 + imp_MVArh2) / 1000000, 3)

                                total_imp_MVArh = "{:,}".format(total_imp_MVArh_r)
                                list6.append(total_imp_MVArh)

                            unit2_MVArh = sheet.cell(r + 6, x + 2)
                            unit4_MVArh = sheet.cell(r + 8, x + 2)
                            if unit2_MVArh.value == 'Q2 varh' and unit4_MVArh.value == 'Q4 varh':  # getting export MVArh. We add Q3 & Q4
                                exp_MVArh1_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                exp_MVArh1 = float(exp_MVArh1_r)
                                exp_MVArh2_r = sheet.cell(r + 8, x + 1).value.replace(',', '')
                                exp_MVArh2 = float(exp_MVArh2_r)
                                total_exp_MVArh_r = round((exp_MVArh1 + exp_MVArh2) / 1000000, 3)
                                total_exp_MVArh = "{:,}".format(total_exp_MVArh_r)
                                list7.append(total_exp_MVArh)

                                # total MVArh
                                total_MVArh_x = float(total_exp_MVArh.replace(',', '')) - float(
                                    total_imp_MVArh.replace(',', ''))
                                total_MVArh_xx = round(total_MVArh_x, 3)
                                total_MVArh = "{:,}".format(total_MVArh_xx)
                                list8.append(total_MVArh)

                            # global rate1,rate2,rate3,rate4,rate5,rate6
                        if cell.value == 'Register':  # getting the various rates
                            unit1 = sheet.cell(r + 1, x + 2)
                            unit2 = sheet.cell(r + 2, x + 2)
                            unit3 = sheet.cell(r + 3, x + 2)
                            if unit3.value == unit2.value == 'Import Wh':
                                rate1_r = sheet.cell(r + 1, x + 1).value.replace(',', '')
                                rate1_rh = round(float(rate1_r) / 1000000, 3)
                                rate1 = "{:,}".format(rate1_rh)
                                list9.append(rate1)

                            if unit1.value == unit3.value == 'Import Wh':
                                rate2_r = sheet.cell(r + 2, x + 1).value.replace(',', '')
                                rate2_rh = round(float(rate2_r) / 1000000, 3)
                                rate2 = "{:,}".format(rate2_rh)
                                list10.append(rate2)

                            if unit2.value == unit1.value == 'Import Wh':
                                rate3_r = sheet.cell(r + 3, x + 1).value.replace(',', '')
                                rate3_rh = round(float(rate3_r) / 1000000, 3)
                                rate3 = "{:,}".format(rate3_rh)
                                list11.append(rate3)

                            unit4 = sheet.cell(r + 4, x + 2)
                            unit5 = sheet.cell(r + 5, x + 2)
                            unit6 = sheet.cell(r + 6, x + 2)
                            if unit5.value == unit6.value == 'Export Wh':
                                rate4_r = sheet.cell(r + 4, x + 1).value.replace(',', '')
                                rate4_rh = round(float(rate4_r) / 1000000, 3)
                                rate4 = "{:,}".format(rate4_rh)
                                list12.append(rate4)

                            if unit4.value == unit6.value == 'Export Wh':
                                rate5_r = sheet.cell(r + 5, x + 1).value.replace(',', '')
                                rate5_rh = round(float(rate5_r) / 1000000, 3)
                                rate5 = "{:,}".format(rate5_rh)
                                list13.append(rate5)
                            # global rate6
                            if unit4.value == unit5.value == 'Export Wh':
                                rate6_r = sheet.cell(r + 6, x + 1).value.replace(',', '')
                                rate6_rh = round(float(rate6_r) / 1000000, 3)
                                rate6 = "{:,}".format(rate6_rh)
                                list14.append(rate6)
                            # global reset_time, resets
                        if cell.value == 'Billing event details':
                            req_cell = sheet.cell(r + 2, x)
                            req_cell1 = sheet.cell(r + 3, x)
                            if req_cell.value == 'Billing reset number:':
                                resets = sheet.cell(r + 2, x + 1).value
                                # no_of_resets = float(no_of_resets_r)
                                list23.append(resets)

                            if req_cell1.value == 'Time of billing reset:':
                                reset_time = sheet.cell(r + 3, x + 1).value
                                list24.append(reset_time)
                            # global M_O_S,M_O_S_r
                        if cell.value == 'Billing period start date' and sheet.cell(r + 1,
                                                                                    x).value == 'Billing period end date':
                            M_O_S_r = sheet.cell(r, x + 1).value
                            #print(f"THIS IS T: {M_O_S_r}")
                            M_O_S_x = M_O_S_r.split('/')
                            M_O_S_i = int(M_O_S_x[1])
                            M_O_S = f'{calendar.month_name[M_O_S_i]} {int(M_O_S_x[2])}'
                            #print(M_O_S)
                            list16.append(M_O_S)
                            # global max_dem_1,max_dem_2,max_dem_3
                        if cell.value == 'Cumulative maximum demands' and sheet.cell(r + 2,
                                                                                     x).value == 'Register':  # getting maximum demands

                            unit_1 = sheet.cell(r + 3, x + 2)
                            unit_2 = sheet.cell(r + 4, x + 2)
                            unit_3 = sheet.cell(r + 5, x + 2)
                            if unit_1.value == 'VA' and unit_2.value == 'VA' and unit_3.value == 'VA':
                                max_dem_1_r = sheet.cell(r + 3, x + 1).value.replace(',', '')  # max demand 1
                                max_dem_1_rh = round(float(max_dem_1_r) / 1000000, 3)
                                max_dem_1 = "{:,}".format(max_dem_1_rh)
                                list17.append(max_dem_1)

                                max_dem_2_r = sheet.cell(r + 4, x + 1).value.replace(',', '')  # max demand 2
                                max_dem_2_rh = round(float(max_dem_2_r) / 1000000, 3)
                                max_dem_2 = "{:,}".format(max_dem_2_rh)
                                list18.append(max_dem_2)

                                max_dem_3_r = sheet.cell(r + 5, x + 1).value.replace(',', '')  # max demand 3
                                max_dem_3_rh = round(float(max_dem_3_r) / 1000000, 3)
                                max_dem_3 = "{:,}".format(max_dem_3_rh)
                                list20.append(max_dem_3)
                            # global max_dem_1t, max_dem_2t, max_dem_3t

                        if cell.value == 'Maximum demands' and sheet.cell(r + 2, x).value == 'Register' and sheet.cell(
                                r + 2,
                                x + 3).value == 'Time and date':  # getting max-demands time
                            unit_1t = sheet.cell(r + 3, x + 3)
                            unit_2t = sheet.cell(r + 4, x + 3)
                            unit_3t = sheet.cell(r + 5, x + 3)

                            max_dem_1t = unit_1t.value
                            max_dem_2t = unit_2t.value
                            max_dem_3t = unit_3t.value
                            list20.append(max_dem_1t)
                            list21.append(max_dem_2t)
                            list22.append(max_dem_3t)

                            # constants
                            unix = datetime.now().replace(second=0, microsecond=0)
                            inserted_by = active_user
                            #active_user = 'kimera'
                            pow_down_count = 'No Data'
                            pow_down_dt = 'No Data'
                            prog_count = 'No Data'
                            prog_count_dt = 'No Data'
                            # list25 = ['No data']
                            list25.append(pow_down_count)
                            list26.append(pow_down_dt)
                            list27.append(prog_count)
                            list28.append(prog_count_dt)
                            list29.append(unix)
                            print(list29)
                            list30.append(inserted_by)
                            list31.append(meter_no)
                            list32.append(D_O_R)
                            list33.append(read_time)

        df1_a = pd.DataFrame(list29)
        df1_a.columns = ['time_inserted']
        df1_b = pd.DataFrame(list30)
        df1_b.columns = ['inserted_by']
        df1_c = pd.DataFrame(list30)
        df1_c.columns = ['meter_read_by']
        df1_e = pd.DataFrame(list33)
        df1_e.columns = ['reading_datetime']
        df1_f = pd.DataFrame(list31)
        df1_f.columns = ['meter_no']
        df1 = pd.DataFrame(list16)
        df1.columns = ['Energy_For']
        df2 = pd.DataFrame(list1a)
        df2.columns = ['Cum_Import']
        df3 = pd.DataFrame(list4)
        df3.columns = ['Cum_Export']
        df4 = pd.DataFrame(list5)
        df4.columns = ['Apparent_Power']
        df5 = pd.DataFrame(list9)
        df5.columns = ['Rate_1']
        df6 = pd.DataFrame(list10)
        df6.columns = ['Rate_2']
        df7 = pd.DataFrame(list11)
        df7.columns = ['Rate_3']
        df8 = pd.DataFrame(list12)
        df8.columns = ['Rate_4']
        df9 = pd.DataFrame(list13)
        df9.columns = ['Rate_5']
        df10 = pd.DataFrame(list14)
        df10.columns = ['Rate_6']
        df11 = pd.DataFrame(list17)
        df11.columns = ['Max_Dem_1']
        df12 = pd.DataFrame(list20)
        df12.columns = ['Max_Dem1_time']
        df13 = pd.DataFrame(list18)
        df13.columns = ['Max_Dem_2']
        df14 = pd.DataFrame(list21)
        df14.columns = ['Max_Dem2_time']
        df15 = pd.DataFrame(list20)
        df15.columns = ['Max_Dem_3']
        df16 = pd.DataFrame(list22)
        df16.columns = ['Max_Dem3_time']
        df17 = pd.DataFrame(list6)
        df17.columns = ['Import_MVArh']
        df18 = pd.DataFrame(list7)
        df18.columns = ['Export_MVArh']
        df20 = pd.DataFrame(list8)
        df20.columns = ['Total_MVAh']
        df20 = pd.DataFrame(list23)
        df20.columns = ['No_of_Resets']
        df21 = pd.DataFrame(list24)
        df21.columns = ['Last_Reset']
        df22 = pd.DataFrame(list25)
        df22.columns = ['Power_Down_Count']
        df23 = pd.DataFrame(list26)
        df23.columns = ['Lst_pwr_dwn_date_and_time']
        df24 = pd.DataFrame(list27)
        df24.columns = ['prog_count']
        df25 = pd.DataFrame(list28)
        df25.columns = ['last_prog_date']

        df = pd.concat(
            [df1_a, df1_b, df1_c, df1_e, df1_f, df1, df2, df3, df4, df5, df6, df7, df8, df9, df10, df11, df12, df13,
             df14, df15, df16,
             df17, df18, df20, df20, df21, df22, df23, df24, df25], axis=1)
        df.dropna(subset=['Apparent_Power'], inplace=True)
        #print(df)

        #CREATE ENGINE
        #engine = create_engine('mysql://root:@localhost/meteringdatabase')
        engine = create_engine(
            'mysql://root:uxiGJV2uwUWAS533CRDd@database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com:3306/meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                             user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        c.execute("SHOW TABLES ")
        table_list = []
        data = c.fetchall()
        for x in data:
            table_name = x[0]
            table_list.append(table_name)
        current_table = f'{meter_no}'

        #CHECK IF TABLE EXISTS THEN INSERT
        if current_table in table_list:
            df.to_sql(f'{meter_no}', con=engine, if_exists='append', index=False,
                      dtype={'time_inserted': VARCHAR(length=25), 'inserted_by': VARCHAR(length=20),
                             'meter_read_by': VARCHAR(length=10),
                             'reading_datetime': VARCHAR(length=25), 'meter_no': VARCHAR(length=12),
                             'Energy_For': VARCHAR(length=15), 'Cum_Import': VARCHAR(length=13),
                             'Cum_Export': VARCHAR(length=13), 'Apparent_Power': VARCHAR(length=13),
                             'Rate_1': VARCHAR(length=13), 'Rate_2': VARCHAR(length=13), 'Rate_3': VARCHAR(length=13),
                             'Rate_4': VARCHAR(length=13), 'Rate_5': VARCHAR(length=13), 'Rate_6': VARCHAR(length=13),
                             'Max_Dem_1': VARCHAR(length=13), 'Max_Dem1_time': VARCHAR(length=15),
                             'Max_Dem_2': VARCHAR(length=13), 'Max_Dem2_time': VARCHAR(length=15),
                             'Max_Dem_3': VARCHAR(length=13), 'Max_Dem3_time': VARCHAR(length=15),
                             'Import_MVArh': VARCHAR(length=13), 'Export_MVArh': VARCHAR(length=13),
                             'Total_MVAh': VARCHAR(length=13), 'No_of_Resets': VARCHAR(5),
                             'Last_Reset': VARCHAR(length=25), 'Power_Down_Count': VARCHAR(length=5),
                             'Lst_pwr_dwn_date_and_time': VARCHAR(length=5),
                             'prog_count': VARCHAR(5), 'last_prog_date': VARCHAR(length=25)
                             })

        else:
            df.to_sql(f'{meter_no}', con=engine, if_exists='append', index=False,
                      dtype={'time_inserted': VARCHAR(length=25), 'inserted_by': VARCHAR(length=20),
                             'meter_read_by': VARCHAR(length=10),
                             'reading_datetime': VARCHAR(length=25), 'meter_no': VARCHAR(length=12),
                             'Energy_For': VARCHAR(length=15), 'Cum_Import': VARCHAR(length=13),
                             'Cum_Export': VARCHAR(length=13), 'Apparent_Power': VARCHAR(length=13),
                             'Rate_1': VARCHAR(length=13), 'Rate_2': VARCHAR(length=13), 'Rate_3': VARCHAR(length=13),
                             'Rate_4': VARCHAR(length=13), 'Rate_5': VARCHAR(length=13), 'Rate_6': VARCHAR(length=13),
                             'Max_Dem_1': VARCHAR(length=13), 'Max_Dem1_time': VARCHAR(length=15),
                             'Max_Dem_2': VARCHAR(length=13), 'Max_Dem2_time': VARCHAR(length=15),
                             'Max_Dem_3': VARCHAR(length=13), 'Max_Dem3_time': VARCHAR(length=15),
                             'Import_MVArh': VARCHAR(length=13), 'Export_MVArh': VARCHAR(length=13),
                             'Total_MVAh': VARCHAR(length=13), 'No_of_Resets': VARCHAR(5),
                             'Last_Reset': VARCHAR(length=25), 'Power_Down_Count': VARCHAR(length=5),
                             'Lst_pwr_dwn_date_and_time': VARCHAR(length=5),
                             'prog_count': VARCHAR(5), 'last_prog_date': VARCHAR(length=25)
                             })

            c.execute(f"ALTER TABLE {current_table} ADD id int(5) NOT NULL PRIMARY KEY AUTO_INCREMENT FIRST")

        conn.commit()
        c.close()
        conn.close()


    return render(request, 'form_success.html')

def meter_summary(request):

    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
    c = conn.cursor()
    tables = ['standalone_meters','substation_meters','ipp_meters']
    manufacturers = ['elster','cewe','schneider','landis']
    no_of_manu = len(manufacturers)
    secox = 2 * no_of_manu

    elster,cewe,landis,scheneider = [],[],[],[]
    manu_list_mm = []
    manu_list_ch = []
    output_list = []
    this_year = date.today().year
    m_year_list = []
    for i in range(0,5):
        yx = this_year + i
        m_year_list.append(yx)

    replace_year = this_year - 10
    print(replace_year)
    replace_list_mm,replace_list_ch = [],[]
    for i in tables:
        sql = f"SELECT COUNT(*) FROM {i}" #COUNTITNG TOTAL METERS
        c.execute(sql)
        data = c.fetchone()
        data1 = list(data)
        final_data = data1[0]
        output_list.append(final_data)
        for m in manufacturers:  #COMPILING LIST PER MANUFACTURER
            c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_manufacturer = %s",(m))
            manu_f = c.fetchall()
            manu1 = list(manu_f)
            manu_count = list(manu1[0])
            manu_list_mm.append(manu_count)
            c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_manufacturer_ch = %s", (m))
            manu_f_ch = c.fetchall()
            manu1_ch = list(manu_f_ch)
            manu_count_ch = list(manu1_ch[0])
            manu_list_ch.append(manu_count_ch)
            c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_YOM <= %s and meter_manufacturer = %s", (replace_year,m))
            meters_fr = c.fetchall()
            meters_fr_1 = list(meters_fr[0])
            replace_list_mm.append(meters_fr_1)
            c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_YOM_ch <= %s and meter_manufacturer_ch = %s", (replace_year, m))
            meters_fr_ch = c.fetchall()
            meters_fr_1_ch = list(meters_fr_ch[0])
            replace_list_ch.append(meters_fr_1_ch)


    result_1 = sum(manu_list_mm, []) #main meter list
    result_2 = sum(manu_list_ch, [])    #check meter list

    standalone_main_meters = result_1[:no_of_manu]      #stand alone main meters per manufacturer
    standalone_check_meters = result_2[:no_of_manu]
    stand_main = sum(standalone_main_meters)
    stand_check = sum(standalone_check_meters)

    substation_main_meters = result_1[no_of_manu:secox]       #substation main meters per manufacturer
    subs_sum_main = sum(substation_main_meters)
    substation_check_meters = result_2[no_of_manu:secox]
    sub_sum_check = sum(substation_check_meters)
    ipp_main_meters = result_1[secox:]            #ipp main meters per manufacturer
    ipp_check_meters = result_2[secox:]
    ipp_sum_main = sum(ipp_main_meters)
    ipp_sum_check = sum(ipp_check_meters)

    total_meters = [x + y for x, y in zip(result_1, result_2)]
    total_standalone_meters = [x + y for x, y in zip(standalone_main_meters, standalone_check_meters)]
    total_substation_meters = [x + y for x, y in zip(substation_main_meters, substation_check_meters)]
    total_ipp_meters = [x + y for x, y in zip(ipp_main_meters, ipp_check_meters)]
    total_each_manu = [x + y + z for x, y,z in zip(total_standalone_meters, total_substation_meters,total_ipp_meters)]
    total_stand = sum(total_standalone_meters)
    total_substation = sum(total_substation_meters)
    total_ipp = sum(total_ipp_meters)
    overall = sum(total_each_manu)
    meter_sum = sum(total_standalone_meters) + sum(total_substation_meters) + sum(total_ipp_meters)
    total_each_manu_main = [x + y + z for x, y, z in zip(standalone_main_meters, substation_main_meters, ipp_main_meters)]
    total_each_manu_check = [x + y + z for x, y, z in
                            zip(standalone_check_meters, substation_check_meters, ipp_check_meters)]
    main_sum_each = sum(total_each_manu_main)
    check_sum_each = sum(total_each_manu_check)

    #Meter replacement table
    print(replace_list_mm)
    print(replace_list_ch)
    main_rep_list = sum(replace_list_mm, [])
    check_rep_list = sum(replace_list_ch, [])
    stand_main_rep = main_rep_list[:no_of_manu]
    stand_check_rep = check_rep_list[:no_of_manu]
    sub_main_rep = main_rep_list[no_of_manu:secox]
    sub_check_rep = check_rep_list[no_of_manu:secox]
    ipp_main_rep = main_rep_list[secox:]
    ipp_check_rep = check_rep_list[secox:]

    #Summing Up lists
    sum_stand_mrp = sum(stand_main_rep)
    sum_stand_crp = sum(stand_check_rep)
    sum_sub_mrp = sum(sub_main_rep)
    sum_sub_crp = sum(sub_check_rep)
    sum_ipp_mrp = sum(ipp_main_rep)
    sum_ipp_crp = sum(ipp_check_rep)

    total_rp_main = [x + y + z for x, y, z in zip(stand_main_rep, sub_main_rep, ipp_main_rep)]
    total_rp_check = [x + y + z for x, y, z in zip(stand_check_rep, sub_check_rep, ipp_check_rep)]
    sum_tt_rp_main = sum(total_rp_main)
    sum_tt_rp_check = sum(total_rp_check)


    conn.commit()
    c.close()
    conn.close()

    return render(request,'meter_summary.html',{'output_list':output_list,'manu_list_mm':manu_list_mm,'meter_sum':meter_sum,
     'manu_list_ch':manu_list_ch,'total_meters':total_meters,'total_standalone_meters':total_standalone_meters,'total_ipp':total_ipp,
    'total_substation_meters':total_substation_meters,'total_ipp_meters':total_ipp_meters,'total_stand':total_stand,
    'total_substation':total_substation,'total_each_manu':total_each_manu,'overall':overall,'standalone_main_meters':standalone_main_meters,
    'standalone_check_meters':standalone_check_meters,'substation_main_meters':substation_main_meters,'ipp_main_meters':ipp_main_meters,
    'substation_check_meters':substation_check_meters,'ipp_check_meters':ipp_check_meters,'stand_main':stand_main,'stand_check':stand_check,
    'subs_sum_main':subs_sum_main,'sub_sum_check':sub_sum_check,'ipp_sum_main':ipp_sum_main,'ipp_sum_check':ipp_sum_check,
    'total_each_manu_main':total_each_manu_main,'total_each_manu_check':total_each_manu_check,'main_sum_each':main_sum_each,
    'check_sum_each':check_sum_each,'stand_main_rep':stand_main_rep,'stand_check_rep':stand_check_rep,'sub_main_rep':sub_main_rep,
     'sub_check_rep':sub_check_rep,'ipp_main_rep':ipp_main_rep,'ipp_check_rep':ipp_check_rep,'sum_stand_mrp':sum_stand_mrp,'sum_stand_crp':sum_stand_crp,
      'sum_sub_mrp':sum_sub_mrp,'sum_sub_crp':sum_sub_crp,'sum_ipp_mrp':sum_ipp_mrp,'sum_ipp_crp':sum_ipp_crp,'total_rp_main':total_rp_main,
       'total_rp_check':total_rp_check,'sum_tt_rp_main':sum_tt_rp_main,'sum_tt_rp_check':sum_tt_rp_check,'this_year':this_year,
                                                'm_year_list':m_year_list})

#UPDATING METERS DUE FOR REPLACEMENT IN VARIOUS YEARS
def meter_update(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        # conn = pymysql.connect(host='localhost', port=3306,
        #                user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        year = request.POST.get("selected_year")
        replace_year = int(year) - 10
        replace_list_mm,replace_list_ch = [],[]
        tables = ['standalone_meters', 'substation_meters', 'ipp_meters']
        manufacturers = ['elster', 'cewe', 'schneider', 'landis']
        for i in tables:
            for m in manufacturers:
                c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_YOM = %s and meter_manufacturer = %s",(replace_year, m))
                meters_fr = c.fetchone()
                for x in meters_fr:
                    replace_list_mm.append(x)

                c.execute(f"SELECT COUNT(*) FROM {i} WHERE meter_YOM_ch = %s and meter_manufacturer_ch = %s",(replace_year, m))
                meters_fr_ch = c.fetchone()
                for l in meters_fr_ch:
                    replace_list_ch.append(l)

        finals = replace_list_mm + replace_list_ch
        c.close()
        conn.close()
        print(finals)
        sum_main_stand_f = finals[:4]           #GETTING VARIOUS SUMS FOR MAIN METERS
        sum_main_stand = sum(sum_main_stand_f)
        sum_main_sub_f = finals[4:8]
        sum_main_sub = sum(sum_main_sub_f)
        sum_main_ipp_f = finals[8:12]
        sum_main_ipp = sum(sum_main_ipp_f)
        sum_ch_stand_f = finals[12:16]          #GETTING VARIOUS SUMS FOR CHECK METERS
        sum_ch_stand = sum(sum_ch_stand_f)
        sum_ch_sub_f = finals[16:20]
        sum_ch_sub = sum(sum_ch_sub_f)
        sum_ch_ipp_f = finals[20:24]
        sum_ch_ipp = sum(sum_ch_ipp_f)

        #APPEND VARIOUS SUMS TO THE LIST
        finals.extend([sum_main_stand,sum_main_sub,sum_main_ipp,sum_ch_stand,sum_ch_sub,sum_ch_ipp])
        final_data1 = json.dumps({'finals': finals})
        print(final_data1)
        return HttpResponse(final_data1, content_type="application/json")


    return render(request,'new_job.html')

#INSERTING DETAILS OF A NEW JOB
def new_job(request):
    return render(request,'new_job.html')

def store_newjob(request):
    if request.method == 'POST':
        global active_user
        Meter_No = request.POST['job_meter']
        pp = f'Jobs/{Meter_No}'
        #pp = f'files_for_{active_user}'
        storage = str(pp)
        if not os.path.exists(f"{storage}"):
            os.makedirs(f"{storage}")
        #uploaded_file = request.FILES['hist']
        feeder_name = request.POST['feeder']
        job_type = request.POST['job_desc']
        job_description = request.POST['job_details']
        fs = FileSystemStorage()
        fs1 = FileSystemStorage(location=storage)
        if 'file_1' in request.FILES:
            ctfile = request.FILES['file_1']
            ctfilename = fs.save(ctfile.name, ctfile)
            CT_analysis = fs.url(ctfilename)
            filex = fs1.save(ctfile.name, ctfile)
            #CT_analysis = f'{storage}/{ctfilename}'

        else:
            CT_analysis = ''

        if 'file_2' in request.FILES:
            vtfile = request.FILES['file_2']
            #fs = FileSystemStorage(location=storage)
            vtfilename = fs.save(vtfile.name, vtfile)
            VT_analysis = fs.url(vtfilename)
            #VT_analysis = f'{storage}/{vtfilename}'
            filey = fs1.save(vtfile.name, vtfile)
        else:
            VT_analysis = ''

        if 'file_3' in request.FILES:
            EMA = request.FILES['file_3']
            #fs = FileSystemStorage(location=storage)
            EMAfilename = fs.save(EMA.name, EMA)
            Energy_Meter_Analysis = fs.url(EMAfilename)
            #Energy_Meter_Analysis = f'{storage}/{EMAfilename}'
            filez= fs1.save(EMA.name, EMA)

        else:
            Energy_Meter_Analysis = ''

        if 'file_4' in request.FILES:
            field_report_f = request.FILES['file_4']
            # fs = FileSystemStorage(location=storage)
            FR_filename = fs.save(field_report_f.name, field_report_f)
            field_report = fs.url(FR_filename)
            #field_report = f'{storage}/{FR_filename}'
            filezz = fs1.save(field_report_f.name, field_report_f)
        else:
            field_report = ''

        # conn = pymysql.connect(host='localhost', port=3306,
        #                        user='root', password='', database='meteringdatabase')

        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                             user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()

        sql = """INSERT INTO metering_database_work (Meter_No,feeder_name,job_type,job_description,CT_analysis,VT_analysis,
        Energy_Meter_Analysis,field_report) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
        """
        val = (Meter_No,feeder_name,job_type,job_description,CT_analysis,VT_analysis,Energy_Meter_Analysis,field_report)
        c.execute(sql,val)

        conn.commit()
        c.close()
        conn.close()
        return render(request, 'form_success.html')


def config_file_page(request):

    return render(request,'add_config_file.html',{'final_stand_listall':final_stand_listall,'final_sub_listall':final_sub_listall,
    'final_ipp_listall':final_ipp_listall})

def energy_loss(request):
    uetcl_substations = ['Lugogo','Kampala_North','Queensway','Kawanda','Namungona','Mutundwe']
    uetcl_ipp_list = ['Bujagali','Owen_Falls','Kiira_Extension','Mpanga']
    month_list = []
    for month in calendar.month_name:
        month_list.append(month)
    month_list.pop(0)

    year_list = []
    unix = datetime.now().year
    for i in range(0, 4):
        yearr = unix - i
        year_list.append(yearr)
    #print(year_list)
    return render(request,'energy_loss.html',{'uetcl_substations':uetcl_substations,'uetcl_ipp_list':uetcl_ipp_list,'month_list':month_list,
    'year_list':year_list})

def ipp_analysis_month(request):
    return render(request,'ipp_report.html')

def feeder_list(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        #sub = json.loads(request.body)
        sub = request.POST.get("selected_sub")
        print(sub)
        #sql = "select feeder_name from substation_meters where substation = %s",(sub)
        c.execute("select component_name,component_voltage from substation_meters where substation = %s",(sub))
        #c.execute(sql)
        data = c.fetchall()
        final_data = list(data)
        finals = []
        voltages = []
        for dt in final_data:
            dt1 = dt[0]
            finals.append(dt1)
            dt2 = dt[1]
            voltages.append(dt2)
        final_voltages = list(dict.fromkeys(voltages))
        print(final_voltages)
        print(finals)
        conn.commit()
        c.close()
        conn.close()
        final_data1 = json.dumps({'finals':finals})
        context = {"final_data1json":final_data1,'finals':finals}
        print(f'This is the final data: {finals}')
        print(f"finals {final_data1}")
        return HttpResponse(final_data1,content_type ="application/json" )

    else:
        kamp = ['kintu','kamara']
        conte = json.dumps(kamp)
        return JsonResponse(conte,content_type ="application/json",safe=False)

def ipp_feeder_list(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
        #               user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        #sub = json.loads(request.body)
        ipp = request.POST.get("selected_ipp")
        print(ipp)
        #sql = "select feeder_name from substation_meters where substation = %s",(sub)
        c.execute("select feeder_name,feeder_voltage from ipp_meters where name_ipp = %s",(ipp))
        #c.execute(sql)
        data = c.fetchall()
        final_data = list(data)
        finals_ipp = []
        voltages = []
        for dt in final_data:
            dt1 = dt[0]
            finals_ipp.append(dt1)
            dt2 = dt[1]
            voltages.append(dt2)
        final_voltages = list(dict.fromkeys(voltages))
        print(final_voltages)
        print(finals_ipp)
        conn.commit()
        c.close()
        conn.close()
        final_data1 = json.dumps({'finals_ipp':finals_ipp})
        context = {"final_data1json":final_data1,'finals_ipp':finals_ipp}
        print(f'This is the final data: {finals_ipp}')
        print(f"finals {final_data1}")

        return HttpResponse(final_data1,content_type ="application/json" )

    else:
        kamp = ['kintu','kamara']
        conte = json.dumps(kamp)
        return JsonResponse(conte,content_type ="application/json",safe=False)

def stand_feeder_list(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
        #               user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        #sub = json.loads(request.body)
        dist = request.POST.get("selected_dist")
        print(dist)
        #sql = "select feeder_name from substation_meters where substation = %s",(sub)
        c.execute("select feeder_name,voltage from standalone_meters where district = %s",(dist))
        #c.execute(sql)
        data = c.fetchall()
        final_data = list(data)
        finals_stand = []
        voltages = []
        for dt in final_data:
            dt1 = dt[0]
            finals_stand.append(dt1)
            dt2 = dt[1]
            voltages.append(dt2)
        final_voltages = list(dict.fromkeys(voltages))
        print(final_voltages)
        print(finals_stand)
        conn.commit()
        c.close()
        conn.close()
        final_data1 = json.dumps({'finals_stand':finals_stand})
        context = {"final_data1json":final_data1,'finals_stand':finals_stand}
        print(f'This is the final data: {finals_stand}')
        print(f"finals {final_data1}")

        return HttpResponse(final_data1,content_type ="application/json" )

    else:
        kamp = ['kintu','kamara']
        conte = json.dumps(kamp)
        return JsonResponse(conte,content_type ="application/json",safe=False)

#PLOTTING LOAD PROFILES
def plotly_lp(request):
    #conn = pymysql.connect(host='localhost', port=3306, user='root', password='', database='meteringdatabase')
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    c = conn.cursor
    df_2 = pd.read_sql(sql="select date,end_time, PW, P_var FROM UETCL0102_load_profile", con=conn)
    df_2.reset_index(inplace=False)
    df_2['datetime'] = df_2['date'] + df_2['end_time']

    data = []
    date_list = df_2['date'].tolist() #converting date column to list
    datee = []
    for it in date_list:
        pp = (str(it)).strip()
        new_date = datetime.strptime(pp, "%Y-%m-%d").date()
        datee.append(new_date)

    list1 = df_2['end_time'].tolist()  #converting end_time column to list
    time_list = []
    for item in list1:
        tt = (str(item)).strip()
        tt_f = tt.split(':')[0]
        time_list.append(tt_f)

    df_2 = df_2.drop(["date"], axis=1) #DROPPING DATE COLUMN FROM DF

    df_2.insert(2, 'f_endtime', time_list, True)
    df_2.insert(1, 'date', datee, True)

    df_length = len(df_2.index)
    df_2['PW'] = df_2['PW'].replace(',', '').astype(float)
    df_2['P_var'] = df_2['P_var'].replace(',', '').astype(float)
    # df_2['date'] = pd.to_datetime(df_2['P_var'],format="%Y/%M/%d")
    # Getting Apparent power
    for i in range(0, df_length):
        val1 = pow(df_2["PW"].loc[i], 2)
        val2 = pow(df_2["P_var"].loc[i], 2)
        val_t = val1 + val2
        valuesx = math.sqrt(val_t)
        values = round(valuesx, 3)
        data.append(values)

    df_2.insert(5,'App_pow',data,True)
    print(f"THIS IS DF2 SCHEME {df_2.info()}")
    new_df = df_2.copy()
    #print(new_df)
    new_df_startdate = new_df['date'].iloc[0]
    new_df_enddate = new_df['date'].iloc[-1]
    #print(new_df_enddate)
    #new_df.to_excel("C:\\Users\\KIMERA\\Desktop\\1.xlsx")
    maxx = new_df.loc[new_df['PW'].idxmax()].tolist()

    # dash app
    #app = dash.Dash(__name__)
    app = DjangoDash('SimpleExample')
    #app.css.serve_locally = True
    #app.scripts.config.serve_locally = True

    app.layout = html.Div(children=[html.Div(className='kimera',   #OVERALL DIV
     style={'width':'90%',},

     children=[
         html.Div(className='header_div',                   #FIRST DIV HAS TWO CHILDREN #'border':'1px solid black'
                  style={'width':'95%','background-color':'#cccccc','display': 'flex',
    'justify-content': 'center','padding':'2%', 'margin-left':'5%'},
          children=[
              html.Div(className='set_par',              #FIRST CHILD
               style={'float':'left','width':'30%','background-color':'#F8F9F9','border-radius':'10px',
                      'margin':'auto','padding':'20px'},
                       children=[
                           html.Label(children="Select Range:  ", style={'font-size': 30, 'color': 'blue', }),
                            html.Br(),
                           dcc.DatePickerRange(
                               id='date_range',
                               min_date_allowed=new_df_startdate,
                               max_date_allowed=new_df_enddate,
                               initial_visible_month=date(2020, 10, 2),
                               start_date=date(2020, 10, 1),
                               end_date=date(2020, 10, 3),
                               style={'border-radius': '10px', 'border': '3px solid #09ADE4'}),
                       ]
                  ),
              html.Br(),
              html.Br(),
              html.Div(className='result_par' ,              #SECOND CHILD
               style={'float':'left', 'width': '60%','background-color':'#F8F9F9','border-radius':'10px',
                      'margin-left':'2%','padding':'1%'},

                       children=[html.H2(children="Results", ),
                    html.Label(children="Max Demand (kW): ",style={'font-size':20,}),
              dcc.Input(id='max_pw', type="text", value=1,
                        style={'border':'none','color':'blue','font-size':20,'width':'130px','text-align':'center'}),

              # html.Label(children="Max Demand Date: ",style={'font-size':20,} ),
              # dcc.Input(id='date_max', type="text", value=1,
              #           style={'border':'none','color':'blue','font-size':20,'width':'130px','text-align':'center'}),
              # # html.Br(),
              # # html.Br(),
              html.Label(children="Max Demand Date Time : ",style={'fontSize': 20,'margin-left':'20px'} ),
              dcc.Input(id='time_max', type="text", value=1,
                style={'border':'none','color':'blue','font-size':20,'width':'220px','text-align':'center'}),
              # html.Label(children="Hours", style={'fontSize':20,'marginLeft':'10px'}),
              html.Br(),
              html.Br(),
              html.Label(children="Min Demand (kW): ", style={'fontSize': 20}),
              dcc.Input(id='Min_pw', type="text", value=1,
                        style={'border':'none','color':'blue','font-size':20,'width':'130px','text-align':'center'}),

              # html.Label(children="Min Demand Date: ", style={'fontSize': 20}),
              # dcc.Input(id='date_Min', type="text", value=1,
              #           style={'border':'none','color':'blue','font-size':20,'width':'130px','text-align':'center'}),

              html.Label(children="Min Demand Date Time: ",style={'fontSize': 20,'margin-left':'20px'} ),
              dcc.Input(id='time_Min', type="text", value=1,
                style={'border':'none','color':'blue','font-size':20,'width':'220px','text-align':'center',})
                                 ])
          ]) ,  #END OF FIRST MAIN CHILD

         ] ),
        # html.Div(className='graph_div',  # GRAPH DIV WITH TWO CHILDREN ALSO
        #          children=[
        #              html.Div(className='graph_header',  # FIRST CHILD
        #                       children=[html.Label(children='This is the Plot')],
        #                       ),
                     html.Div(className='graph',  # SECOND CHILD
                              children=[dcc.Graph(
                                  id="graph",
                                  figure={})],
                              ),
    ])

       # ])
    @app.callback(

        [Output(component_id="graph", component_property="figure"),
         Output(component_id="max_pw", component_property="value"),
         Output(component_id="time_max", component_property="value"),
         #Output(component_id="date_max", component_property="value"),
         Output(component_id="Min_pw", component_property="value"),
         Output(component_id="time_Min", component_property="value"),
         #Output(component_id="date_Min", component_property="value")
         ],
        [Input(component_id="date_range", component_property="start_date"),
         Input(component_id="date_range", component_property="end_date")]
    )

    def update_graph(start_date1, end_date1):
        new_dff = new_df.copy()
        new_dff['date'] = pd.to_datetime(new_dff['date'])

        datex1 = datetime.strptime(start_date1, "%Y-%m-%d")
        datex2 = datetime.strptime(end_date1,"%Y-%m-%d")

        #GETTING A DATAFRAME BETWEEN TWO DATES
        final_df = new_dff[new_dff['date'].between(datex1, datex2)]

        final_df['date'] = final_df['date'].astype(str)
        final_df['datetime'] = final_df[['date', 'end_time']].agg('.'.join, axis=1) #COMBINE TWO COLUMNS
        #print(f"THIS IS THE FINAL DF \n {final_df}")
        maxx1 = final_df.loc[final_df['PW'].idxmax()].tolist()
        max_val1 = maxx1[2]
        max_val = "{:,}".format(max_val1)
        date_maxx = maxx1[1]
        time_maxx = maxx1[0]
        date_time = maxx1[6]

        min1 = final_df.loc[final_df['PW'].idxmin()].tolist()
        print(f"THIS IS MIN: {min1}")
        min_val1 = min1[2]
        min_val = "{:,}".format(min_val1)
        date_minx = min1[1]
        time_minx = min1[0]
        date_time_min = min1[6]
        fig = go.Figure(
            data=[go.Line(y=final_df['PW'], x=final_df['datetime'],name='Real Power'),
                  go.Line(y=final_df['App_pow'], x=final_df['datetime'], name = 'Apparent Power'),
                  go.Line(y=final_df['P_var'], x=final_df['datetime'], name = 'Reactive Power'),]
            # data=[],
            #data=[go.Line(y=final_df['App_pow'], x=final_df['datetime'])],
        )
        # FIGURE SIZE
        fig.update_layout(
            autosize=False,
            #width=1800,
            height=600,
        ),
        fig.update_xaxes(title_text='DateTime',showgrid=False,title_font = {"size": 25,'color':'blue'},title_standoff = 25 )
        fig.update_yaxes(title_text='Power',showgrid=False,title_font = {"size": 25,'color':'blue'},title_standoff = 35)
        #return fig, max_val, time_maxx, date_maxx,min_val,time_minx,date_minx
        fig.update_layout(
            title={
                'text': "This is the Load Profile",
                'y': 0.9,
                'x': 0.5,
                'xanchor': 'center',
                'yanchor': 'top'},
        title_font=dict(
        size=35,
        color="RebeccaPurple"))

        return fig, max_val, date_time, min_val, date_time_min,
    return render(request, 'plotly_app.html')


#ONE DAY LOAD PROFILE
def plotly_lp_one(request):
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    #conn = pymysql.connect(host='localhost', port=3306, user='root', password='', database='meteringdatabase')
    c = conn.cursor
    df_2 = pd.read_sql(sql="select date,end_time, PW, P_var FROM UETCL0102_load_profile", con=conn)
    df_2.reset_index(inplace=False)
    # print(df_2)
    data = []
    date_list = df_2["end_time"].tolist()
    date_list1 = []

    # changing date format
    date_listx = df_2['date'].tolist()
    #print(date_listx)
    dattt = []
    for item in date_listx:
        new_l = datetime.strptime(item, "%Y-%m-%d")
        #new_l = datetime.strptime(l, "%Y/%m/%d").date()
        dattt.append(new_l)

    #print(dattt)

    df_date = pd.DataFrame(dattt)
    df_date.columns = ['date']

    df_2 = df_2.drop(["end_time"], axis=1)  # dropping end_time from dataframe
    df_2 = df_2.drop(["date"], axis=1)  # dropping date from dataframe

    # cleaning the end_time
    for item in date_list:
        itemx = item.split(':')
        new_date = f"{item.split(':')[0]}:{item.split(':')[1]}"
        new_date1 = new_date.strip()
        date_list1.append(new_date1)
    # print(date_list1)
    df_end_time = pd.DataFrame(date_list1)
    df_end_time.columns = ['end_time']

    df_length = len(df_2.index)
    df_2['PW'] = df_2['PW'].replace(',', '').astype(float)
    df_2['P_var'] = df_2['P_var'].replace(',', '').astype(float)
    # df_2['date'] = pd.to_datetime(df_2['P_var'],format="%Y/%M/%d")
    #df_2_enddate = df_2['date'].iloc[-1]

    # Getting Apparent power
    for i in range(0, df_length):
        val1 = pow(df_2["PW"].loc[i], 2)
        val2 = pow(df_2["P_var"].loc[i], 2)
        val_t = val1 + val2
        valuesx = math.sqrt(val_t)
        values = round(valuesx, 3)
        data.append(values)

    df_ia = pd.DataFrame(data)
    df_ia.columns = ['App_pow']

    new_df = pd.concat([df_end_time, df_date, df_2, df_ia], axis=1)
    #print(new_df)
    #new_df.to_excel("C:\\Users\\KIMERA\\Desktop\\1.xlsx")
    maxx = new_df.loc[new_df['PW'].idxmax()].tolist()
    new_df_enddate = new_df['date'].iloc[-1]
    new_df_startdate = new_df['date'].iloc[0]


    # app = dash.Dash("kimera")
    #app = dash.Dash(__name__)
    app = DjangoDash('SimpleExample')
    app.layout = html.Div(children=[html.Div(className='kimera',  # OVERALL DIV
                                             style={'width': '90%', },

     children=[ html.Div(className='header_div',            # FIRST DIV HAS TWO CHILDREN
      style={'width': '95%', 'background-color': '#cccccc','display': 'flex',
             'justify-content': 'center', 'padding': '2%','margin-left': '5%'},
      children=[html.Div(className='set_par',  # FIRST CHILD
       style={'float': 'left', 'width': '30%','background-color': '#F8F9F9','border-radius': '10px',
              'margin': 'auto', 'padding': '20px'},
       children=[
           html.Label(children="Select Day:  ",
                      style={'font-size': 30,'color': 'blue', }),
                       html.Br(),
                       dcc.DatePickerSingle(id="datex", min_date_allowed=new_df_startdate,
                                            initial_visible_month=new_df_startdate,
                                            max_date_allowed=new_df_enddate,
                                            date=new_df_startdate, display_format="DD/MM/YYYY"),
                       # date=datetime.today().date()
                   ]
                   ),
          html.Br(),
          html.Br(),
          html.Div(className='result_par',  # SECOND CHILD
           style={'float': 'left', 'width': '60%', 'background-color': '#F8F9F9','border-radius': '10px',
                  'margin-left': '2%', 'padding': '1%'},
           children=[html.H2(children="Results", ),
                     html.Label(children="Max Demand (kW): ",style={'font-size': 20, }),
                     dcc.Input(id='max_pw', type="text", value=1,
           style={'border': 'none', 'color': 'blue', 'font-size': 20,  'width': '130px', 'text-align': 'center'}),
             html.Label(children="Max Demand Time : ",
                         style={'fontSize': 20,  'margin-left': '20px'}),

         dcc.Input(id='time_max', type="text", value=1,
                   style={'border': 'none', 'color': 'blue',  'font-size': 20,   'width': '150px','text-align': 'center'}),
         html.Br(),
         html.Br(),
         html.Label(children="Min Demand (kW): ", style={'fontSize': 20}),
         dcc.Input(id='min_pw', type="text", value=1,
                   style={'border': 'none', 'color': 'blue', 'font-size': 20, 'width': '130px', 'text-align': 'center'}),

         html.Label(children="Min Demand Time: ",
             style={'fontSize': 20,'margin-left': '20px'}),
         dcc.Input(id='time_min', type="text",value=1,
                   style={'border': 'none','color': 'blue','font-size': 20,'width': '150px','text-align': 'center', })
         ])
      ]),  # END OF FIRST MAIN CHILD

             ]),

                html.Div(className='graph',  # SECOND CHILD
                         children=[dcc.Graph(
                             id="graph",
                             figure={})],
                         ),
                ])

    # #call back application
    @app.callback(

        [Output(component_id="graph", component_property="figure"),
         Output(component_id="max_pw", component_property="value"),
         Output(component_id="time_max", component_property="value"),
         Output(component_id="min_pw", component_property="value"),
         Output(component_id="time_min", component_property="value")],
        [Input(component_id="datex", component_property="date"),
         Input(component_id="datex", component_property="min_date_allowed")]
    )
    def update_graph(interest_day,min_date):
        # interest_day = interest_day.__format__("%d/%m/%Y")
        new_dff = new_df.copy()
        #print(new_dff.info())
        # print(new_dff)
        new_dff = new_dff[new_dff['date'] == interest_day]
        #print(interest_day)
        # plot_df = new_df.loc[new_df['date'] == interest_day]
        #print(f"graph df is \n{new_dff}")
        maxx1 = new_dff.loc[new_dff['PW'].idxmax()].tolist()
        max_val1 = maxx1[2]
        max_val = "{:,}".format(max_val1)
        time_maxx = maxx1[0]

        min1 = new_dff.loc[new_dff['PW'].idxmin()].tolist()
        print(f"THIS IS MIN: {min1}")
        min_val1 = min1[2]
        min_val = "{:,}".format(min_val1)
        time_min = min1[0]
        fig = go.Figure(
            data=[go.Line(y=new_dff['PW'], x=new_dff['end_time'],name='Real Power'),
                  go.Line(y=new_dff['App_pow'], x=new_dff['end_time'],name='Apparent Power'),
                  go.Line(y=new_dff['P_var'], x=new_dff['end_time'],name='Reactive Power'),],

        )
        # FIGURE SIZE
        fig.update_layout(
            autosize=False,
            # width=1800,
            height=500,
        ),
        fig.update_xaxes(title_text='Time', title_font={"size": 25, 'color': 'blue'}, title_standoff=25)
        fig.update_yaxes(title_text='Power', title_font={"size": 25, 'color': 'blue'}, title_standoff=35)
        fig.update_layout( title={
                'text': "This is the Load Profile", 'y': 0.9, 'x': 0.5, 'xanchor': 'center', 'yanchor': 'top'},
            title_font=dict(size=35,color="RebeccaPurple"))

        return fig, max_val, time_maxx,min_val,time_min

    return render(request, 'plotly_app.html')
    #
    # if __name__ == "__plotly_lp__":
    # app.run_server(debug=True, use_reloader=True)

def prometer(request):
    #global meter_no_hist, meter_no_tou, meter_no_maxd
    global bill_date_hist, bill_date_tou, bill_date_maxd,reading_datetime,energy_for,filename_maxd,meter_no_hist,meter_no_tou,meter_no_dd
    global filename_hist
    pp = f'PDF_files_for_{active_user}'
    listz = []
    storage = str(pp)
    for f in request.FILES.getlist('prometer'):
        fs = FileSystemStorage(location=storage)
        file_m = fs.save(f.name, f)
        listz.append(f.name)
    #print(listz)

    for file_item in listz:
        #print(file_item)
        #os.path(file_item)
        df1 = tabula.read_pdf(f'{storage}/{file_item}', pages='all',
                              guess=False)
        df = df1[0]
        #print(df)
        col_names = df.columns.tolist()
        df[col_names[0]] = df[col_names[0]].astype(str)
        if col_names[0].find("Historical Energy Registers") != -1:
            global meter_no_hist,hist_val,reading_datetime,energy_for
            meter_no_hist = col_names[0].split()[-1]
            reading_datetime = df[col_names[0]].iloc[-1]
            #final_read = reading_datetime.rsplit(' ', 1)[0]

            for item in df[col_names[0]]:
                global bill_date_hist
                if item.find('End of Period:') != -1:
                    global filename_hist
                    #print(item)
                    date_period = item.split(':')[1]
                    date_read = date_period.split()[0]
                    datex = datetime.strptime(date_read, '%m/%d/%Y').date()
                    a_month = dateutil.relativedelta.relativedelta(months=1)
                    month = datex - a_month
                    month_name = month.strftime('%B')
                    bill_date_hist = f"{month.strftime('%B')} {month.year}"
                    energy_for = f"{month_name} {month.year}"
                    if not os.path.exists(f"{storage}/{month_name}_{month.year}_PDF_files"):
                        os.makedirs(f"{storage}/{month_name}_{month.year}_PDF_files")
                    filename_hist = f"{storage}/{month_name}_{month.year}_PDF_files/{meter_no_hist}_{month_name}_{month.year}_Historical_Registers.pdf"
                    #DELETING EXISTING FILES
                    if os.path.exists(filename_hist):
                        os.remove(filename_hist)
                    os.rename(f'{storage}/{file_item}',filename_hist)
                    break

        elif col_names[0].find('Historical TOU Registers') != -1:
            global meter_no_tou,TOU_file
            TOU_file = file_item
            #print(f"WE SHALL USE {TOU_file}")
            meter_no_tou = col_names[0].split()[-1]
            #print(meter_no_tou)

            for item in df[col_names[0]]:
                global bill_date_tou
                if item.find('End of Period:') != -1:
                    #print(item)
                    global filename_tou
                    date_period = item.split(':')[1]
                    date_read = date_period.split()[0]
                    datex = datetime.strptime(date_read, '%m/%d/%Y').date()
                    a_month = dateutil.relativedelta.relativedelta(months=1)
                    month = datex - a_month
                    month_name = month.strftime('%B')
                    bill_date_tou = f"{month.strftime('%B')} {month.year}"
                    #print(bill_date_tou)
                    if not os.path.exists(f"{storage}/{month_name}_{month.year}_PDF_files"):
                        os.makedirs(f"{storage}/{month_name}_{month.year}_PDF_files")
                    filename_tou = f"{storage}/{month_name}_{month.year}_PDF_files/{meter_no_tou}_{month_name}_{month.year}_TOU_Registers.pdf"
                    if os.path.exists(filename_tou):
                        os.remove(filename_tou)
                    os.rename(f'{storage}/{file_item}',filename_tou)

                    break

        elif col_names[0].find('Historical Maximum Demand Values') != -1:
            global meter_no_dd,max_dd_file
            max_dd_file = file_item
            #print(f"WE SHALL USE{max_dd_file}")
            meter_no_dd = col_names[0].split()[-1]
            #print(meter_no_dd)

            for item in df[col_names[0]]:
                if item.find('End of Period:') != -1:
                    #print(f"THIS IS WHERE WE FOUND IT\n{item}")
                    global bill_date_maxd
                    global filename_maxd
                    date_period = item.split(':')[1]
                    date_read = date_period.split()[0]
                    datex = datetime.strptime(date_read, '%m/%d/%Y').date()
                    a_month = dateutil.relativedelta.relativedelta(months=1)
                    month = datex - a_month
                    month_name = month.strftime('%B')
                    #month = datetime.strptime(date_read, '%m/%d/%Y')
                    month_name =month.strftime('%B')
                    bill_date_maxd = f"{month.strftime('%B')} {month.year}"
                    #print(bill_date_maxd)

                    if not os.path.exists(f"{storage}/{month_name}_{month.year}_PDF_files"):
                        os.makedirs(f"{storage}/{month_name}_{month.year}_PDF_files")
                    filename_maxd = f"{storage}/{month_name}_{month.year}_PDF_files/{meter_no_dd}_{month_name}_{month.year}_Historical Max_DD.pdf"
                    if os.path.exists(filename_maxd):
                        os.remove(filename_maxd)  # file exits, delete it
                    # rename the file
                    #os.rename(originalFilename, outputFilename)
                    os.rename(f'{storage}/{file_item}',filename_maxd)
                    break

        else:
            error = f"{file_item} IS NOT REQUIRED. PLEASE UPLOAD \n" \
                    f"1. HISTORICAL REGISTER \n " \
                    f"2. TOU REGISTER \n " \
                    f"3.MAXIMUM DEMAND REGISTERS"
            return render(request,'prometer_errors.html',{'error':error})

    if meter_no_hist == meter_no_tou == meter_no_dd and bill_date_maxd == bill_date_hist == bill_date_tou:
        df1 = tabula.read_pdf(filename_hist, pages='all',
                              guess=False)
        # df2 = tabula.read_pdf('Gaba Main Log.pdf',pages = '2',guess=True)
        df = df1[0]
        df.columns = ['Description', 'Values']
        df['Description'] = df['Description'].astype(str)
        global new_itemlist, new_val, var_list
        new_itemlist = []
        new_val = []
        var_list = []

        # getting energy values
        desc_list = ['Active Energy Imp', 'Active Energy Exp', 'Apparent Energy Imp', 'Apparent Energy Exp',
                     'Reactive Energy Imp',
                     'Reactive Energy Exp']
        for var in desc_list:
            for item in df['Description']:
                if item.find(var) != -1 and item.find('(M') != -1:
                    x = 1000

                    val = df['Values'][df['Description'] == item].values
                    var_list.append(var)
                    new_itemlist.append(item)
                    new_val.append(float(val) * x)

                elif item.find(var) != -1 and item.find('(k') != -1:
                    # x = 1000
                    val = df['Values'][df['Description'] == item].values
                    var_list.append(var)
                    new_itemlist.append(item)
                    new_val.append(float(val))
        # print(var_list)
        # print(new_itemlist)
        # print(new_val)


        if len(new_itemlist) != len(new_val) != len(var_list):
            error = "Historical energy values did not finish computing"
            return render(request, 'prometer_errors.html',{'error':error})

        #PROCESSING TOU VALUES
        df3 = tabula.read_pdf(filename_tou, pages='all', guess=True)
        df4 = df3[0]
        final_cols = df4.columns.tolist()
        first_col = final_cols[0]
        #print(df4.columns.tolist())
        # print(df4)
        global vals_1
        use_list = ['Active Energy Imp', 'Active Energy Exp']
        vals_1 = []
        vals_2 = []
        for var in use_list:
            for item in df4[first_col]:
                if item.find(var) != -1 and item.find('(M') != -1:
                    x = 1000

                    val1 = df4['Rate 1'][df4[first_col] == item].values
                    val2 = df4['Rate 2'][df4[first_col] == item].values
                    val3 = df4['Rate 3'][df4[first_col] == item].values
                    vals_1.append(float(val1) * x)
                    vals_1.append(float(val2) * x)
                    vals_1.append(float(val3) * x)

                if item.find(var) != -1 and item.find('(k') != -1:
                    val1 = df4['Rate 1'][df4[first_col] == item].values
                    val2 = df4['Rate 2'][df4[first_col] == item].values
                    val3 = df4['Rate 3'][df4[first_col] == item].values
                    vals_1.append(float(val1))
                    vals_1.append(float(val2))
                    vals_1.append(float(val3))
        #print(vals_1)
        if len(vals_1) != 6:
            error = "TOU rates where not all computed"
            return render(request, 'prometer_errors.html', {'error': error})

        #PROCESSING MAXIMUM DEMANDS
        df7 = tabula.read_pdf(filename_maxd, pages='all',
                              guess=True)
        df8 = df7[0]
        # print(df6)
        cols8 = df8.columns.tolist()
        #print(cols8)
        vals_2 = []
        global xunit
        for item in df8[cols8[0]]:
            if item.find('Active Power Imp') != -1 and item.find('(M') != -1:
                xunit = 1000
                val1 = df8[cols8[1]][df8[cols8[0]] == item].values
                val2 = df8[cols8[2]][df8[cols8[0]] == item].values
                val3 = df8[cols8[3]][df8[cols8[0]] == item].values
                # print(val1)
                vals_2.append(str(val1))
                vals_2.append(str(val2))
                vals_2.append(str(val3))

            if item.find('Active Power Imp') != -1 and item.find('(k') != -1:
                xunit = 1
                val1 = df8[cols8[1]][df8[cols8[0]] == item].values
                val2 = df8[cols8[2]][df8[cols8[0]] == item].values
                val3 = df8[cols8[3]][df8[cols8[0]] == item].values
                # print(val1)
                vals_2.append(str(val1))
                vals_2.append(str(val2))
                vals_2.append(str(val3))
        global max_dem_l, max_dem_datetime
        #print(vals_2)
        max_dem_l = []
        max_dem_datetime = []
        if len(vals_2) == 3:
            #print("All values Found")
            characters = ["[", "]", "'", "(", ")"]
            l = 1
            for x in vals_2:
                for char in characters:

                    x = x.replace(char, "")
                # print(x)
                max_dem = x.split()[0]
                final_max_dem = float(max_dem) * xunit
                max_dem_dt = f"{x.split()[1]} {x.split()[2]}"
                # print(f"Max Dem {l} is: {max_dem}")
                # print(f"Max Dem {l} Datetime is: {max_dem_dt}")
                l += 1
                max_dem_l.append(final_max_dem)
                max_dem_datetime.append(max_dem_dt)
            # print(max_dem_l)
            # print(max_dem_datetime)
            # print(xunit)

        else:
            error = "Error in maximum Demand Computation"
            return render(request, 'prometer_errors.html', {'error': error})

        # directory = storage
        # os.chdir(directory)
        # files = glob.glob('*.pdf')
        # for filename in files:
        #     os.unlink(filename)

        #INSERTING INTO DATABASE
        # conn = pymysql.connect(host='localhost', port=3306,
        # user='root', password='', database='meteringdatabase')
        timestamp = datetime.now().replace(microsecond=0)
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        c.execute(f"""CREATE TABLE IF NOT EXISTS {meter_no_hist}(id INT AUTO_INCREMENT PRIMARY KEY,timestamp VARCHAR(20),inserted_by VARCHAR(15),
        meter_readby VARCHAR (15),reading_datetime VARCHAR(25), energy_for VARCHAR(15),Cum_Import VARCHAR(15),                  
        Cum_Export VARCHAR(15),Apparent_Energy_Imp VARCHAR(15),Rate_1 VARCHAR(15),Rate_2 VARCHAR(15),Rate_3 VARCHAR(15),Rate_4 VARCHAR(15),                   
        Rate_5 VARCHAR(15),Rate_6 VARCHAR(15),Max_Dem_1 VARCHAR(10),Max_Dem1_time VARCHAR(20),Max_Dem_2 VARCHAR(10),Max_Dem2_time VARCHAR(20),                   
        Max_Dem_3 VARCHAR(10),Max_Dem3_time VARCHAR(20) )""")

        sql = f"""INSERT INTO {meter_no_hist} (timestamp,inserted_by,meter_readby,reading_datetime,energy_for,Cum_Import,Cum_Export,
              Apparent_Energy_Imp,Rate_1,Rate_2,Rate_3,Rate_4,Rate_5,Rate_6,Max_Dem_1,Max_Dem1_time,Max_Dem_2,Max_Dem2_time,
              Max_Dem_3,Max_Dem3_time) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""
        values = (timestamp,active_user,active_user,reading_datetime,energy_for,new_val[0],new_val[1],new_val[2],vals_1[0],
                  vals_1[1],vals_1[2],vals_1[3],vals_1[4],vals_1[5],max_dem_l[0],max_dem_datetime[0],max_dem_l[1],
                  max_dem_datetime[1],max_dem_l[2],max_dem_datetime[2])

        c.execute(sql,values)
        conn.commit()
        c.close()
        conn.close()

        return render(request, 'form_success.html')


def prometer_100(request):  #INSERTING RECORDS FOR CEWE PROMETER 100
    pymysql.converters.encoders[np.float64] = pymysql.converters.escape_float
    pymysql.converters.conversions = pymysql.converters.encoders.copy()
    pymysql.converters.conversions.update(pymysql.converters.decoders)
    unix = datetime.now().strftime("%Y-%m-%d %H:%M")
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    c = conn.cursor()
    pp = f'files_for_{active_user}'
    listz,list_meter = [],[]
    storage = str(pp)
    global meter_no,meter_no_f,f_billr,f_bill,reading_date,reading_date, meter_no, Energy_For, active_imp, active_exp, apparent_imp
    global rate_1, rate_2,rate_3, rate_4, rate_5, rate_6,month_name, year
    for f in request.FILES.getlist('prometer_100'):
        fs = FileSystemStorage(location=storage)
        file_m = fs.save(f.name, f)
        listz.append(f.name)
    # print(listz)
    if len(listz) != 2:
        warning = "Two CSV files are required. One for Main_Energy and other for TOD_Energy"
        return render(request,'prometer_errors.html', {'warning': warning})

    for file_item in listz:
        meter_no = file_item.split()[0]
        reading_date = file_item.split()[2]
        list_meter.append(meter_no)

    if list_meter[0] != list_meter[1]:
        warning = "The two files are not from the same Meter. Please insert files from the same meter"
        return render(request,'prometer_errors.html', {'warning': warning})

    if meter_no in cewe_P100_list_m:
        meter_no_f = meter_no
        for file_item in listz:
            if file_item.find('MainEnergy') != -1:
                df = pd.read_csv(f'{storage}/{file_item}')
                unix = datetime.now().strftime("%Y-%m-%d %H:%M")
                my_cols = df.columns.tolist()
                df[my_cols[0]] = df[my_cols[0]].astype(str)
                ind = df.loc[df[my_cols[0]] == 'History'].index
                my_ind = ind.tolist()
                global xf
                if my_cols[0].find('kWH'):
                    xf = 1

                else:
                    xf = 1000
                listl = df.loc[df[my_cols[0]] == 'History'].to_numpy()
                col_list = listl[0]

                df1 = df.copy()
                df1 = df1.iloc[my_ind[0] + 1:]  # Removing the default headers
                df1.reset_index(drop=True, inplace=False)
                df1.columns = col_list
                df1['History'] = df1['History'].astype(int)
                df1['Active(I) Total'] = df1['Active(I) Total'].astype(float)
                df1['Active(E) Total'] = df1['Active(E) Total'].astype(float)
                df1['Apparent-Active(I) - type 2'] = df1['Apparent-Active(I) - type 2'].astype(float)
                df1['Apparent-Active(E) - type 6'] = df1['Apparent-Active(E) - type 6'].astype(float)
                df1['Reactive(I)'] = df1['Reactive(I)'].astype(float)
                df1['Reactive(E)'] = df1['Reactive(E)'].astype(float)

                bill_date = df1['Billing Date'][df1['History'] == 1].values
                f_bill = datetime.strptime(bill_date[0], "%d/%m/%Y %H:%M").date()
                active_imp_t = df1['Active(I) Total'][df1['Billing Date'] == bill_date[0]].values  # Active import
                active_imp = active_imp_t[0] * xf
                active_exp_t = df1['Active(E) Total'][df1['Billing Date'] == bill_date[0]].values
                active_exp = active_exp_t[0] * xf
                apparent_imp_f = df1['Apparent-Active(I) - type 2'][df1['Billing Date'] == bill_date[0]].values
                apparent_imp = apparent_imp_f[0] * xf
                apparent_exp_f = df1['Apparent-Active(E) - type 6'][df1['Billing Date'] == bill_date[0]].values
                apparent_exp = apparent_exp_f[0] * xf
                reactive_imp = df1['Reactive(I)'][df1['Billing Date'] == bill_date[0]].values
                reactive_exp = df1['Reactive(E)'][df1['Billing Date'] == bill_date[0]].values

                a_month = dateutil.relativedelta.relativedelta(months=1)
                datex = f_bill - a_month
                month_name = datex.strftime('%B')
                year = datex.strftime('%Y')
                Energy_For = f'{month_name} {year}'
                df.to_excel(f"{month_name}_{year}_files/{meter_no}_{month_name}_{year}_Historical.xlsx")

            if file_item.find("TODEnergy") != -1:
                dfr = pd.read_csv(f'{storage}/{file_item}')
                my_colsr = dfr.columns.tolist()
                dfr[my_colsr[0]] = dfr[my_colsr[0]].astype(str)
                indr = dfr.loc[dfr[my_colsr[0]] == 'History'].index
                my_indr = indr.tolist()
                listr = dfr.loc[dfr[my_colsr[0]] == 'History'].to_numpy()

                col_listr = listr[0]
                dfr1 = dfr.copy()
                dfr1 = dfr1.iloc[my_indr[0] + 1:]
                dfr1.reset_index(drop=True, inplace=False)
                dfr1.columns = col_listr
                dfr1['History'] = dfr1['History'].astype(int)
                dfr1['TOD/SLAB'] = dfr1['TOD/SLAB'].astype(int)
                dfr1['Active(I) Total'] = dfr1['Active(I) Total'].astype(float)
                dfr1['Active(E) Total'] = dfr1['Active(E) Total'].astype(float)

                bill_dater = dfr1['Billing Date'][dfr1['History'] == 1].values
                f_billr = datetime.strptime(bill_dater[0], "%d/%m/%Y %H:%M").date()
                new_df = dfr1.loc[dfr1['Billing Date'] == bill_dater[0]]
                rate_1_f = new_df['Active(I) Total'][new_df['TOD/SLAB'] == 1].values
                rate_1 = rate_1_f[0] * xf
                rate_2_f = new_df['Active(I) Total'][new_df['TOD/SLAB'] == 2].values
                rate_2 = rate_2_f[0] * xf
                rate_3_f = new_df['Active(I) Total'][new_df['TOD/SLAB'] == 3].values
                rate_3 = rate_3_f[0] * xf
                rate_4_f = new_df['Active(E) Total'][new_df['TOD/SLAB'] == 1].values
                rate_4 = rate_4_f[0] * xf
                rate_5_f = new_df['Active(E) Total'][new_df['TOD/SLAB'] == 2].values
                rate_5 = rate_5_f[0] * xf
                rate_6_f = new_df['Active(E) Total'][new_df['TOD/SLAB'] == 3].values
                rate_6 = rate_6_f[0] * xf
                dfr.to_excel(f"{month_name}_{year}_files/{meter_no}_{month_name}_{year}_TOD.xlsx")
        #
        if f_billr != f_bill:
             warning = "CSVs are not from the Billing period"
             return render(request, 'prometer_errors.html', {'warning': warning})

        if f_billr == f_bill:
            c.execute(
                f"""CREATE TABLE IF NOT EXISTS {meter_no_f}(id INT AUTO_INCREMENT PRIMARY KEY,timestamp VARCHAR(20),inserted_by VARCHAR(15),
            meter_readby VARCHAR (15),reading_datetime VARCHAR(20), meter_no VARCHAR(15), energy_for VARCHAR(15),Cum_Import VARCHAR(15),
             Cum_Export VARCHAR(15),Apparent_Power VARCHAR(15),Rate_1 VARCHAR(15),Rate_2 VARCHAR(15),Rate_3 VARCHAR(15),Rate_4 VARCHAR(15),
             Rate_5 VARCHAR(15),Rate_6 VARCHAR(15))""")
            sql = f"""INSERT INTO {meter_no_f} (timestamp,inserted_by,meter_readby,reading_datetime,meter_no,energy_for,Cum_Import,Cum_Export,
            Apparent_Power,Rate_1,Rate_2,Rate_3,Rate_4,Rate_5,Rate_6)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """
            values = (unix, active_user, active_user, reading_date, meter_no_f, Energy_For, active_imp, active_exp,
                      apparent_imp,
                      rate_1, rate_2, rate_3, rate_4, rate_5, rate_6)

            c.execute(sql, values)
            # c.execute(f"""INSERT INTO {meter_no} (timestamp,inserted_by,meter_readby,reading_datetime,meter_no,energy_for,Cum_Import,Cum_Export,
            # Apparent_Power,Rate_1,Rate_2,Rate_3,Rate_4,Rate_5,Rate_6)
            # VALUES (unix,active_user,active_user,reading_date,meter_no,Energy_For,active_imp,active_exp,apparent_imp,rate_1,rate_2,
            #           rate_3,rate_4,rate_5,rate_6)
            # """
            # )
            conn.commit()
            c.close()
            conn.close()
            return render(request, 'form_success.html')

    # if meter_no in cewe_P100_list_c:
    #     meter_no_f = meter_no
    #     print("not there")
    else:
        warning = "Meter No is not on list"
        return render(request,'prometer_errors.html',{'warning':warning})


#INSERTING MONTHLY RECORDS FOR LANDIS
def landis(request):
    conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                           user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
    c = conn.cursor()
    global active_user
    pp = f'files_for_{active_user}'
    storage = str(pp)
    if not os.path.exists(storage):
        os.makedirs(storage)
    uploaded_file = request.FILES['landis']
    fs = FileSystemStorage(location=storage)
    filexx = fs.save(uploaded_file.name, uploaded_file)
    print(filexx)
    file = f"{storage}/{filexx}"
    df = pd.read_excel(file, engine='openpyxl')
    col_list = df.columns.tolist()
    if 'OBIS' in col_list:
        # GETTING METER NO
        meter_no = df['Value'][(df['Designation'] == 'COSEM logical device name') & (df['OBIS'] == '0-0:42.0.0')].any()

        # GROSS APPARENT POWER
        apparent_import = df['Value'][
            (df['Designation'] == 'Gross Apparent energy import +VA (QI+QIV)') & (df['Unit'] == 'kVAh')
            & (df['OBIS'] == '1-1:9.8.0')].any()


        # apparent_export = df['Value'][(df['Designation'] == 'Gross Apparent energy export -VA (QII+QIII)') & (df['Unit'] == 'kVAh')].any()
        # print(apparent_export)

        # GETTING TOTAL EXPORTS & IMPORTS
        total_export = df['Value'][
            (df['Designation'] == 'Gross Active energy export -A (QII+QIII)') & (df['Unit'] == 'kWh')
            & (df['OBIS'] == '1-1:2.8.0')].any()
        total_import = df['Value'][
            (df['Designation'] == 'Gross Active energy import +A (QI+QIV)') & (df['Unit'] == 'kWh')
            & (df['OBIS'] == '1-1:1.8.0')].any()

        # GETTING DIFFERENT RATES
        rate_1 = df['Value'][(df['Designation'] == 'Active energy import +A (QI+QIV) rate 1') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:1.8.1')].any()
        rate_2 = df['Value'][(df['Designation'] == 'Active energy import +A (QI+QIV) rate 2') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:1.8.2')].any()
        rate_3 = df['Value'][(df['Designation'] == 'Active energy import +A (QI+QIV) rate 3') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:1.8.3')].any()
        rate_4 = df['Value'][(df['Designation'] == 'Active energy export -A (QII+QIII) rate 1') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:2.8.1')].any()

        rate_5 = df['Value'][(df['Designation'] == 'Active energy export -A (QII+QIII) rate 2') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:2.8.2')].any()

        rate_6 = df['Value'][(df['Designation'] == 'Active energy export -A (QII+QIII) rate 3') & (df['Unit'] == 'kWh')
                             & (df['OBIS'] == '1-1:2.8.3')].any()

        # GETTING POWER FARLURES FOR VARIOUS PHASES
        L1_power_f = df['Value'][
            (df['Designation'] == 'Number of power failures L1') & (df['OBIS'] == '0-0:96.7.1')].any()
        L2_power_f = df['Value'][(df['Designation'] == 'Number of power failures L2') & (df['OBIS'] == '0-0:96.7.2')].any()


        L3_power_f = df['Value'][(df['Designation'] == 'Number of power failures L3') & (df['OBIS'] == '0-0:96.7.3')].any()

        # CHECKING CONFIG COUNT
        config_count = df['Value'][(df['Designation'] == 'Number of parameterisations (configuration program changes)')
                                   & (df['OBIS'] == '0-0:96.2.0')].any()
        last_config_date_x = df['Value'][(df['Designation'] == 'Date of last parameterisation (configuration program change)')
            & (df['OBIS'] == '0-0:96.2.1')].any()
        last_config_date = last_config_date_x.rsplit(' ', 1)[0]


        # MAXIMUM DEMANDS
        import_max_DD_x = df['Value'][(df['Designation'] == 'Maximum demand +A (QI+QIV)') & (df['Unit'] == 'kW')
                                      & (df['OBIS'] == '1-1:1.6.0')].any()
        import_max_DD = import_max_DD_x.partition(' ')[0]  # GETTING IMPORT MAX DD
        yzz = import_max_DD_x.partition(' ')[2].replace('(', '')
        imp_max_dd_datetime = yzz.rsplit(' ', 1)[0]  # IMPORT MAX DD DATETIME

        export_max_DD_x = df['Value'][(df['Designation'] == 'Maximum demand -A (QII+QIII)') & (df['Unit'] == 'kW')
                                      & (df['OBIS'] == '1-1:2.6.0')].any()
        export_max_DD = export_max_DD_x.partition(' ')[0]  # GETTING EXPORT MAX DD
        yzx = export_max_DD_x.partition(' ')[2].replace('(', '')
        exp_max_dd_datetime = yzx.rsplit(' ', 1)[0]  # EXPORT MAX DD DATETIME

        # GETTING BILLING RESETS
        billing_resets = df['Value'][(df['Designation'] == 'Billing period reset counter') & (df['OBIS'] == '1-0:0.1.0')].any()

        # READING DATETIME
        read_datetime_x = df['Value'][(df['Designation'] == 'Clock') & (df['OBIS'] == '0-0:1.0.0')].any()
        read_datetime = read_datetime_x.rsplit(' ', 1)[0]

        # GETTING TIME STAMP
        time_stamp = now = datetime.now().__format__("%Y-%m-%d %H:%M")

        # BILLING PERIOD
        bill_date_x = df['Value'][
            (df['Designation'] == 'Date of last billing period reset') & (df['OBIS'] == '1-0:0.1.2')].any()
        date_x = bill_date_x.partition(' ')[0]
        new_date = datetime.strptime(date_x, '%Y-%m-%d').date()
        a_month = dateutil.relativedelta.relativedelta(months=1)
        month_year = new_date - a_month
        billing_period = month_year.strftime('%B_%Y')

        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        c.execute(
            f"""CREATE TABLE IF NOT EXISTS {meter_no} (id int(5) NOT NULL AUTO_INCREMENT PRIMARY KEY,time_stamp VARCHAR(20),
        inserted_by varchar(10),meter_read_by varchar(10),read_datetime varchar(20),billing_period varchar(20),
        billing_resets int(5),apparent_import VARCHAR(13),total_export VARCHAR(13),total_import VARCHAR(13),Rate_1 VARCHAR(13),Rate_2 VARCHAR(13),
        Rate_3 VARCHAR(13),Rate_4 VARCHAR(13),Rate_5 VARCHAR(13),Rate_6 VARCHAR(13),import_max_DD VARCHAR(13),imp_max_dd_datetime VARCHAR(20),
        export_max_DD VARCHAR(13),exp_max_dd_datetime VARCHAR(20),config_count VARCHAR(13),last_config_date VARCHAR(20))""")

        sql = f"""INSERT INTO {meter_no} (time_stamp,inserted_by,meter_read_by,read_datetime,billing_period, billing_resets,apparent_import,
        total_export,total_import,Rate_1,Rate_2,Rate_3,Rate_4,Rate_5,Rate_6,import_max_DD,imp_max_dd_datetime,export_max_DD,exp_max_dd_datetime,
        config_count,last_config_date)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
        val = (time_stamp, active_user, active_user, read_datetime, billing_period, billing_resets, apparent_import,
               total_export, total_import, rate_1,
               rate_2, rate_3, rate_4, rate_5, rate_6, import_max_DD, imp_max_dd_datetime, export_max_DD,
               exp_max_dd_datetime, config_count, last_config_date)
        c.execute(sql, val)
        conn.commit()
        c.close()
        conn.close()

        #SAVING FINAL FILE
        final_storage = f'{storage}/landis/{billing_period}_files'
        if not os.path.exists(final_storage):
            os.makedirs(final_storage)
        df.to_excel(f'{final_storage}/{meter_no}_{billing_period}.xlsx', index=None,
                    header=True, sheet_name='Billing Values')
        return render(request, 'form_success.html')
    else:
        error = 'File not correct. insert right file'
        return render(request, 'prometer_errors.html', {'error': error})

#LOAD PROFILES FOR ALL
def new_LP(request):
    return render(request,'insertLP.html')

def monthly_LP(request): # Inserting Load Profile
    engine = create_engine('mysql://root:uxiGJV2uwUWAS533CRDd@database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com:3306/meteringdatabase')
    #engine = create_engine('mysql://root:@localhost/meteringdatabase')
    if request.method == 'POST':
        global active_user, df_au,df1,df4,df2,df3,df_ia,df_ib,df_ic,df_va,df_vb,df_vc,df_pw,df_Pvar
        inserted_by = active_user
        time_stamp = now = datetime.now().__format__("%Y-%m-%d %H:%M")
        pp = f'files_for_{active_user}/Elster'
        storage = str(pp)
        uploaded_file = request.FILES['LP_elster']
        fs = FileSystemStorage(location=storage)
        file_LP = fs.save(uploaded_file.name, uploaded_file)
        print(uploaded_file)
        filex = fs.save(uploaded_file.name, uploaded_file)
        print(filex)
        #CONVERTING CSV TO EXCEL
        temp_store = f"{storage}/temp_LP"
        if not os.path.exists(temp_store):
            os.makedirs(temp_store)
        workbook = xlsxwriter.Workbook(f'{temp_store}/LP_temp.xlsx')
        worksheet = workbook.add_worksheet()
        y,x = 0,0
        with open(f"{storage}/{filex}") as csv_file:
            csv_data = csv.reader(csv_file)
            for row in csv_data:
                y = 0
                for dataxx in row:
                    worksheet.write(x, y, dataxx)
                    y += 1
                x += 1
            workbook.close()
            csv_file.close()

        wb = xl.load_workbook(f'{temp_store}/LP_temp.xlsx')
        sheets = wb.sheetnames
        ws = wb[sheets[0]]
        input_string = str(ws["B2"].value)
        # print(input_string)
        # global meter
        meter_no = input_string.split()[0].replace('-', '')  # meter number
        for i in range(1, ws.max_column):
            cell_value = ws.cell(row=1, column=i).value
            # if cell_value1 != '':
            if not cell_value == None:
                # cell_value = cell_value1.replace('-','')
                manu_1 = 'Elster'
                if cell_value.find(manu_1) == 1:
                    print(f'manufacturer is Elster')
                    break

        meter_no_x = ws["B2"].value
        meter_no = meter_no_x.split()[0].replace('-', '')

        lista = []
        list_1 = []
        list_2 = []
        list_3 = []
        list_4 = []
        for x in range(1, 20):
            for r in range(1, ws.max_column + 1):
                cell = ws.cell(r, x)

                if cell.value == 'Date' and ws.cell(r, x + 1).value == 'Start' and ws.cell(r,
                                                                                           x + 2).value == 'End':  # check this loop for x and r (row and column)
                    for m in range(r + 1, ws.max_row+1):
                        use_date = ws.cell(m, x).value
                        date_time_obj = datetime.strptime(use_date, '%d/%m/%Y').date()
                        month = date_time_obj.strftime("%B")
                        start_time_1 = ws.cell(m, x + 1).value
                        start_time = '%.6s' % start_time_1
                        end_time_1 = ws.cell(m, x + 2).value
                        end_time = '%.6s' % end_time_1
                        list_1.append(date_time_obj)
                        list_2.append(start_time)
                        list_3.append(end_time)
                        list_4.append(month)
                    # print(list_1)
                    df1 = pd.DataFrame(list_1)
                    df1.columns = ['date']
                    # print(df1)
                    df2 = pd.DataFrame(list_2)
                    df2.columns = ['start_time']
                    # print(df2)
                    df3 = pd.DataFrame(list_3)
                    df3.columns = ['end_time']
                    df4 = pd.DataFrame(list_4)
                    df4.columns = ['month']
                    # print(df3)
                    #df = pd.concat([df1, df2, df3], axis=1)

                    # print(df)

                list_ia = []
                # row_ia = 0

                if cell.value == 'A: PhA: Av':
                    for m in range(r, ws.max_row+1):
                        I_A = ws.cell(m, x - 1).value
                        list_ia.append(I_A)
                        lista.append(active_user)
                    df_ia = pd.DataFrame(list_ia)
                    df_ia.columns = ['I_A']
                    df_au = pd.DataFrame(lista)
                    df_au.columns = ['inserted_by']

                    #dfnew = pd.concat([df, df_ia], axis=1)
                    # print(dfnew)

                    break

                list_ib = []

                if cell.value == 'A: PhB: Av':
                    for m in range(r, ws.max_row+1):
                        I_B = ws.cell(m, x - 1).value
                        list_ib.append(I_B)
                    df_ib = pd.DataFrame(list_ib)
                    df_ib.columns = ['I_B']
                    #df_ib = pd.concat([dfnew, df_ib_x], axis=1)
                    # print(df_ib)
                    break

                list_ic = []

                if cell.value == 'A: PhC: Av':
                    for m in range(r, ws.max_row+1):
                        I_C = ws.cell(m, x - 1).value
                        list_ic.append(I_C)
                    df_ic = pd.DataFrame(list_ic)
                    df_ic.columns = ['I_C']
                    #df_ic = pd.concat([df_ib, df_ic_x], axis=1)
                    #print(df_ic)
                    break

                list_va = []

                if cell.value == 'V: PhA: Av':
                    for m in range(r, ws.max_row+1):
                        V_A = ws.cell(m, x - 1).value
                        list_va.append(V_A)
                    df_va = pd.DataFrame(list_va)
                    df_va.columns = ['V_A']
                    break


                list_vb = []
                if cell.value == 'V: PhB: Av':
                    for m in range(r, ws.max_row+1):
                        V_B = ws.cell(m, x - 1).value
                        list_vb.append(V_B)
                    df_vb = pd.DataFrame(list_vb)
                    df_vb.columns = ['V_B']
                    break

                list_vc = []
                if cell.value == 'V: PhC: Av':
                    for m in range(r, ws.max_row+1):
                        V_C = ws.cell(m, x - 1).value
                        list_vc.append(V_C)
                    df_vc = pd.DataFrame(list_vc)
                    df_vc.columns = ['V_C']
                    break

                global df_pw
                list_pw = []
                if cell.value == 'kW: Sys: Av':
                    for m in range(r, ws.max_row+1):
                        PW = ws.cell(m, x - 1).value
                        # print(PW)
                        list_pw.append(PW)
                    df_pw = pd.DataFrame(list_pw)
                    df_pw.columns = ['PW']
                    break


                list_Pvar = []
                if cell.value == 'kvar: Sys: Av':
                    for m in range(r, ws.max_row+1):
                        Pvar = ws.cell(m, x - 1).value
                        # print(PW)
                        list_Pvar.append(Pvar)
                    df_Pvar = pd.DataFrame(list_Pvar)
                    df_Pvar.columns = ['P_var']
                    #now = datetime.datetime.now().date()
                    break


        df = pd.concat([df_au,df1,df2,df3,df_ia,df_ib,df_ic,df_va,df_vb,df_vc,df_pw,df_Pvar],axis=1)

        #df['date']= pd.to_datetime(df['date'])
        #df['date'] = df['date'].dt.date
        month_year_l = df['date'].tolist()
        month_year1 = []
        for item in month_year_l:
            itemx = str(item).strip()
            month_yearl = datetime.strptime(itemx,"%Y-%m-%d")
            month_year = month_yearl.strftime('%B %Y')
            month_year1.append(month_year)
        #month_year1 = df['date'].dt.strftime('%B %Y').tolist()
        #print(month_year1)
        #df[datex] = df['date'].dt.date
        # INSERTING TIMESTAMP

        timestamp_l = [time_stamp] * len(df.index)
        df.insert(1, 'timestamp', timestamp_l, True)
        df.insert(3, 'month_year', month_year1, True)
        #df["datetime"] = df["date"] + df["endtime"]
        df.to_excel(f'{temp_store}/df_temp.xlsx')

        #CHECK IF TABLE EXISTS
        # conn = pymysql.connect(host='localhost', port=3306,
        # user='root', password='', database='meteringdatabase')
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        c.execute("SHOW TABLES ")
        table_list = []
        data = c.fetchall()
        for x in data:
            table_name = x[0]
            table_list.append(table_name)
        current_table = f'{meter_no}_load_profile'

        if current_table in table_list:
            df.to_sql(f'{current_table}', con=engine, if_exists='append', index=False,dtype={'inserted_by': VARCHAR(length=10),'timestamp': VARCHAR(length=20),
            'date': VARCHAR(length=11),'month_year': VARCHAR(length=20),'start_time': String(length=10), 'end_time': String(length=10),
            'I_A': String(length=10), 'I_B': String(length=10),'I_C': String(length=10),'V_A':VARCHAR(length=10),
            'V_B':VARCHAR(length=10),'V_C':VARCHAR(length=10),'PW':VARCHAR(length=10),'P_var':VARCHAR(length=10)})

        else:
            df.to_sql(f'{current_table}', con=engine, if_exists='append', index=False,
                      dtype={'inserted_by': VARCHAR(length=10),'timestamp': VARCHAR(length=20),'date': VARCHAR(length=11), 'month_year': VARCHAR(length=20),
                             'start_time': String(length=10), 'end_time': String(length=10),
                             'I_A': String(length=10), 'I_B': String(length=10),
                             'I_C': String(length=10), 'V_A': VARCHAR(length=10), 'V_B': VARCHAR(length=10),
                             'V_C': VARCHAR(length=10),
                             'PW': VARCHAR(length=10), 'P_var': VARCHAR(length=10)})
            c.execute(f"ALTER TABLE {current_table} ADD id int(5) NOT NULL PRIMARY KEY AUTO_INCREMENT FIRST")
            conn.commit()
            c.close()
            conn.close()

        return render(request,'form_success.html')

def load_profile_cewe_prometer(request):
    if request.method == 'POST':
        global active_user
        inserted_by = active_user
        pp = f'files_for_{active_user}/cewe_prometer'
        storage = str(pp)
        fs = FileSystemStorage(location=storage)
        uploaded_file = request.FILES['LP_cewe_prometer']
        print(uploaded_file)
        filex = fs.save(uploaded_file.name, uploaded_file)
        print(filex)

        file = f"{storage}/{uploaded_file}"
        temp_store = f"{storage}/cewe_LP"
        if not os.path.exists(temp_store):
            os.makedirs(temp_store)

        # GETTING METER NUMBER
        dfmn = tabula.read_pdf(file, pages='1', guess=False)
        colx = dfmn[0].columns.tolist()
        global meter_no
        for item in colx:
            if item.find('Logger') != -1:
                #print(item)
                meter_no = item.split()[1]
                #print(meter_no)
            break
        temp_store = f"{storage}/load_profiles/{meter_no}"
        if not os.path.exists(temp_store):
            os.makedirs(temp_store)
        df = tabula.read_pdf(file, pages='all', guess=True)
        leng = len(df)
        #print(leng)
        df0 = df[0]
        col_len = len(df0.columns.tolist())
        col_names = df0.columns.tolist()
        #print(col_names)
        global df_final
        for i in range(1, leng):
            dfc = df[i]
            dfc_len = len(dfc.columns.tolist())
            #print(len(dfc.columns.tolist()))
            diff = dfc_len - col_len
            #print(diff)
            dfc.drop(dfc.iloc[:, 1:diff + 1], inplace=True, axis=1)
            dfc.columns = col_names
            df_final = pd.concat([df0, dfc], axis=0)
            df0 = df_final
            #dfc.to_csv(f'C:\\Users\\KIMERA\\Desktop\\energy meter downloads and config\\KIM MARCH 2021\\df{i}.csv')
            df0.append(dfc)
        # DF0 IS DF_FINAL
        df0.drop(df0.columns[1], inplace=True, axis=1)
        # RESETTING PANDAS DF INDEX
        df0.reset_index(inplace=True, drop=True)
        final_column_names = ['Datetime', 'V(L1-L2)', 'V(L2-L3)', 'V(L3-L1)', 'I_A', 'I_C', 'Phase Angle', 'PF', 'Freq',
                              'V_THD(All)',
                              'I_THD(All)', 'P_kW', 'P_kVAr', 'P_kVA']
        df0.columns = final_column_names

        #CONVERTING DATE COLUMN TO DATETIME
        df0 = df0.iloc[1:]
        df0['Datetime'] = pd.to_datetime(df0['Datetime'])
        # df0['Mon_Year'] = df0['Datetime'].dt.strftime('%B %Y')
        month_year = df0['Datetime'].dt.strftime('%B %Y').tolist()
        #print(month_year)
        #time_stamp = datetime.now()
        time_stamp = datetime.now().__format__("%Y-%m-%d %H:%M")
        insert_list = [inserted_by] * len(df0.index)
        timestamp_l = [time_stamp] * len(df0.index)
        df0.insert(0, 'inserted_by', insert_list, True)
        df0.insert(1, 'Timestamp', timestamp_l, True)
        df0.insert(3, 'month_year', month_year)
        startx = df0['month_year'].iloc[0]
        endx = df0['month_year'].iloc[-1]

        #print(len(df0.index))
        df0.to_csv(f'{temp_store}/{meter_no}_{startx}_{endx}_load_profile.csv')
        engine = create_engine(
            'mysql://root:uxiGJV2uwUWAS533CRDd@database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com:3306/meteringdatabase')
        table_list = []
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        c.execute("SHOW TABLES ")
        data = c.fetchall()
        for x in data:
            table_name = x[0]
            table_list.append(table_name)
        current_table = f'{meter_no}_load_profile'
        if current_table in table_list:
            df0.to_sql(f'{current_table}', con=engine, if_exists='append', index=False,
            dtype={'inserted_by': VARCHAR(length=10),'Timestamp': VARCHAR(length=20),'Datetime': VARCHAR(length=30),'month_year': VARCHAR(length=30),
                    'V(L1-L2)': Float(),'V(L2-L3)': Float(), 'V(L3-L1)': Float(),
                   'I_A': Float(), 'I_C': Float(), 'Phase Angle': Float(), 'PF': Float(), 'Freq': Float(),
                   'V_THD(All)': VARCHAR(length=5), 'I_THD(All)': VARCHAR(length=5), 'P_kW': Float(),
                   'P_kVAr': Float(), 'P_kVA': VARCHAR(length=10)})
        else:
            df0.to_sql(f'{current_table}', con=engine, if_exists='append', index=False,
                       dtype={'inserted_by': VARCHAR(length=10), 'Timestamp': VARCHAR(length=20),'Datetime': VARCHAR(length=30),
                              'month_year': VARCHAR(length=30),
                              'V(L1-L2)': Float(), 'V(L2-L3)': Float(), 'V(L3-L1)': Float(),
                              'I_A': Float(), 'I_C': Float(), 'Phase Angle': Float(), 'PF': Float(), 'Freq': Float(),
                              'V_THD(All)': VARCHAR(length=5), 'I_THD(All)': VARCHAR(length=5), 'P_kW': Float(),
                              'P_kVAr': Float(), 'P_kVA': VARCHAR(length=10)})

            c.execute(f"ALTER TABLE {current_table} ADD id int(5) NOT NULL PRIMARY KEY AUTO_INCREMENT FIRST")
            conn.commit()
            c.close()
            conn.close()

        return render(request, 'form_success.html')

def EM_details(request):  #ENERGY METER SEARCH DETAILS

    return render(request, 'meter_search.html',{'final_stand_listall':final_stand_listall,'final_sub_listall':final_sub_listall,
    'final_ipp_listall':final_ipp_listall})

#GETTING METER POINT DETAILS FOR ONE METER
def EM_sub_details(request):
    if request.method == 'POST':
        global active_user
        user = active_user
        substation = request.POST['sub_list']
        feeder = request.POST['feed_name']
        # conn = pymysql.connect(host='localhost', port=3306,
        #                        user='root', password='', database='meteringdatabase')

        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                             user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        sql = "select * from substation_meters where substation = %s and component_name = %s"
        val = (substation,feeder)
        c.execute(sql,val)
        data1 = c.fetchone()
        data = list(data1)
        clean_data = data[3:]
        # print(clean_data)
        clean_data = ['N/A' if x == '' else x for x in clean_data]
        print(clean_data)
        print(len(clean_data))

        #MAIN METER NO = clean_data[28]
        if clean_data[28] in meter_work_list:
            #GETTING JOB REPORTS
            c.execute("select CT_analysis,VT_analysis,Energy_Meter_Analysis,field_report from metering_database_work WHERE feeder_name = %s",(feeder))
            datax = c.fetchall()
            #print(list(datax))
            list_ct = datax[0]

        else:
            list_ct = []
        return render(request, 'sub_meter_return.html',{'clean_data':clean_data,'user':user,'list_ct':list_ct,
        'month_names_in_year':month_names_in_year,'use_year_list':use_year_list})

#IPP ENERGY METER DETAILS
def EM_ipp_details(request):
    if request.method == 'POST':
        ipp_name = request.POST['ipp_list']
        ipp_feeder = request.POST['feed_name_ipp']
        # conn = pymysql.connect(host='localhost', port=3306,
        #                        user='root', password='', database='meteringdatabase')

        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
                               user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        c = conn.cursor()
        sql = "select * from ipp_meters where name_ipp = %s and feeder_name = %s"
        val = (ipp_name, ipp_feeder)
        c.execute(sql, val)
        # data = []
        data1 = c.fetchone()
        data = list(data1)
        clean_data = data[3:]
        #print(clean_data)
        clean_data = ['N/A' if x == '' else x for x in clean_data]
        #print(clean_data)
        return render(request, 'ipp_meter_return.html',{'clean_data':clean_data})


def EM_stand_details(request):
    # if request.method == 'POST':
    #     user_act = request.POST['username']
    #     password_given = request.POST['password']

    return render(request, 'form_success.html')


#GENERATING A NODE REPORT
def sub_node_report(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        #sub = json.loads(request.body)
        meter = request.POST.get("select_meter")
        month = request.POST.get("select_month")
        year = request.POST.get("select_year")
        #data: {'select_month': select_month, 'select_month': select_month, 'select_meter': select_meter},
        #print(sub)
        #GETTING RATED CAPACITY
        c.execute("SELECT capacity,component_voltage,core_used FROM substation_meters WHERE meter_no = %s", (meter))
        data = c.fetchone()

        capacity = float(data[0])
        voltage_r = int(data[1])
        voltage = voltage_r/0.11
        ct_rating = int(data[2])
        print(capacity)
        df = pd.read_sql(sql=f"select Datetime,month_year, P_kW,I_T,V_T, PF, Freq FROM {meter}_load_profile", con=conn)
        #print(df)
        month_year = f"{month} {year}"
        df_2 = df[df['month_year'] == month_year]
        if df_2.dropna().empty:
            max, perc_max, min, perc_min, final_mean_p, perc_mean = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            final_max_pf, final_min_pf, final_mean_pf = 'No Data', 'No Data', 'No Data'
            max_opp_f, perc_max_n, min_opp_f, perc_min_n, final_mean_n, perc_mean_n = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            final_max_pf_n, final_min_pf_n, final_mean_pf_n = 'No Data', 'No Data', 'No Data'
            max_I, perc_ct_max, min_I, perc_ct_min, mean_I, perc_ct_mean = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            max_V, perc_v_max, min_V, perc_v_min, mean_V, perc_v_mean = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            max_I_n, perc_ct_max_n, min_I_n, perc_ct_min_n, mean_I_n, perc_ct_mean_n = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            max_V_n, perc_v_max_n, min_V_n, perc_v_min_n, mean_V_n, perc_v_mean_n = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'

            print(df_2)

        #GETTING POWER IN POSITIVE DIRECTION
        df_3 = df_2[df_2['P_kW'] > 0]
        if df_3.dropna().empty:
            max,perc_max,min,perc_min,final_mean_p,perc_mean = 'No Data','No Data','No Data','No Data','No Data','No Data'
            final_max_pf,final_min_pf,final_mean_pf = 'No Data','No Data','No Data'
            max_I, perc_ct_max, min_I, perc_ct_min, mean_I, perc_ct_mean = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'
            max_V, perc_v_max, min_V, perc_v_min, mean_V, perc_v_mean = 'No Data', 'No Data', 'No Data', 'No Data', 'No Data', 'No Data'

        elif df_3.shape[0] != 0:
            max = df_3['P_kW'].max()
            perc_max = round(((max/capacity)*100),2)
            print(max)
            min = df_3['P_kW'].min()
            perc_min = round(((min / capacity) * 100), 2)
            print(min)
            mean_p = df_3['P_kW'].mean()
            final_mean_p = round(mean_p,2)
            print(mean_p)
            perc_mean =  round(((mean_p/(capacity))*100),2)
            print(perc_mean)

            #GETTING POWER FACTOR IN POSITIVE DIRECTION
            max_pf = df_3['PF'].max()
            final_max_pf = round(max_pf, 2)
            print(round(max_pf, 2))
            min_pf = df_3['PF'].min()
            final_min_pf = round(min_pf, 2)
            print(round(min_pf, 2))
            mean_pf = df_3['PF'].mean()
            final_mean_pf = round(mean_pf, 2)
            print(round(mean_pf, 2))

            #GETTING CURRENT
            max_I_f = df_3['I_T'].max()
            max_I = round((max_I_f), 2)
            print(f"THIS IS MAX I: {max_I}")
            print(f"THIS IS CT RATING {ct_rating}")
            perc_ct_max = round((max_I/ct_rating)*100,2) #GETTING PERCETAGE CT RATING
            min_I_f = df_3['I_T'].min()
            min_I = round((min_I_f), 2)
            perc_ct_min = round((min_I / ct_rating)*100, 2)
            print(f"THIS IS MIN I: {min_I}")
            print(f"THIS IS PERC {perc_ct_min}")
            mean_I_f = df_3['I_T'].mean()
            mean_I = round((mean_I_f), 2)
            perc_ct_mean = round((mean_I / ct_rating)*100, 2)
            print(f"THIS IS MEAN I: {mean_I}")
            print(f"THIS IS PERC {perc_ct_mean}")

            # GETTING VOLTAGE
            max_V_f = df_3['V_T'].max()
            max_V = round((max_V_f *1.73), 2)
            print(f"THIS IS MAX V {max_V}")
            perc_v_max = round((max_V / voltage_r)*100, 2)  # GETTING PERCETAGE VOLTAGE
            print(f"THIS IS PERCENTAGE {perc_v_max}")
            min_V_f = df_3['V_T'].min()
            min_V = round((min_V_f *1.73), 2)
            perc_v_min = round((min_V / voltage_r)*100, 2)
            mean_V_f = df_3['V_T'].mean()
            mean_V = round((mean_V_f*1.73), 2)
            perc_v_mean = round((mean_V /voltage_r)*100, 2)


        # GETTING POWER IN NEGATIVE DIRECTION
        df_4 = df_2[df_2['P_kW'] < 0]

        if df_4.dropna().empty:
            max_opp_f,perc_max_n,min_opp_f,perc_min_n,final_mean_n,perc_mean_n = 'No Data','No Data','No Data','No Data','No Data','No Data'
            final_max_pf_n,final_min_pf_n,final_mean_pf_n = 'No Data','No Data','No Data'
            max_I_n, perc_ct_max_n, min_I_n, perc_ct_min_n, mean_I_n, perc_ct_mean_n = 'No Data','No Data','No Data','No Data','No Data','No Data'
            max_V_n, perc_v_max_n, min_V_n, perc_v_min_n, mean_V_n, perc_v_mean_n = 'No Data','No Data','No Data','No Data','No Data','No Data'


        elif df_4.shape[0] != 0:
            max_opp = df_4['P_kW'].max()
            max_opp_f = abs(max_opp)
            perc_max_n = round(((max_opp_f / capacity) * 100), 2)
            print(max_opp_f)
            min_opp = df_4['P_kW'].min()
            min_opp_f = abs(min_opp)
            perc_min_n = round(((min_opp_f / capacity) * 100), 2)
            print(min_opp_f)
            mean_n = df_4['P_kW'].mean()
            final_mean_nx = round(mean_n,2)
            final_mean_n = abs(final_mean_nx)
            perc_mean_n = round(((final_mean_n / capacity) * 100), 2)
            print(mean_n)

            # GETTING POWER FACTOR IN NEGATIVE DIRECTION
            max_pf_n = df_4['PF'].max()
            final_max_pf_n = round(max_pf_n, 2)
            #print(round(max_pf, 2))
            min_pf_n = df_4['PF'].min()
            final_min_pf_n = round(min_pf_n, 2)
            #print(round(min_pf, 2))
            mean_pf_n = df_4['PF'].mean()
            final_mean_pf_n = round(mean_pf_n, 2)
            #print(round(mean_pf, 2))

            # GETTING CURRENT
            max_I_f_n = df_4['I_T'].max()
            max_I_n = round((max_I_f_n), 2)
            perc_ct_max_n = round((max_I_n / ct_rating)*100, 2)  # GETTING PERCETAGE CT RATING
            min_I_f_n = df_4['I_T'].min()
            min_I_n = round((min_I_f_n), 2)
            perc_ct_min_n = round((min_I_n / ct_rating)*100, 2)
            mean_I_f_n = df_4['I_T'].mean()
            mean_I_n = round((mean_I_f_n), 2)
            perc_ct_mean_n = round((mean_I_n / ct_rating)*100, 2)

            # GETTING VOLTAGE
            max_V_f_n = df_4['V_T'].max()
            max_V_n = round((max_V_f_n * 1.73), 2)
            perc_v_max_n = round((max_V_n / voltage_r)*100, 2)  # GETTING PERCETAGE VOLTAGE
            min_V_f_n = df_4['V_T'].min()
            min_V_n = round((min_V_f_n * 1.73), 2)
            perc_v_min_n = round((min_V_n / voltage_r)*100, 2)
            mean_V_f_n = df_4['V_T'].mean()
            mean_V_n = round((mean_V_f_n * 1.73), 2)
            perc_v_mean_n = round((mean_V_n / voltage_r)*100, 2)

        finals = [final_mean_p,max,min,final_mean_n,max_opp_f, min_opp_f, final_mean_pf,final_max_pf,final_min_pf,perc_mean,perc_max,perc_mean,
                  perc_mean_n,perc_max_n,perc_min_n,final_mean_pf_n,final_max_pf_n,final_min_pf_n,perc_ct_mean,perc_ct_max,perc_ct_min,
                  mean_V,perc_v_mean,max_V,perc_v_max,min_V,perc_v_min,perc_ct_mean_n,perc_ct_max_n,
                  perc_ct_min_n,mean_V_n,perc_v_mean_n,max_V_n,perc_v_max_n,min_V_n,perc_v_min_n]


        conn.commit()
        c.close()
        conn.close()
        final_data1 = json.dumps({'finals':finals})

        return HttpResponse(final_data1,content_type ="application/json" )

    # else:
    #     kamp = ['kintu','kamara']
    #     conte = json.dumps(kamp)
    #     return JsonResponse(conte,content_type ="application/json",safe=False)

def ipp_land_details(request): #IPP LANDING PAGE
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        ipp_name = request.POST.get("ipp_name")
        c.execute("select meter_no,meter_manufacturer,meter_type,meter_no_ch,meter_manufacturer_ch,meter_type_ch from ipp_meters where name_ipp = %s",(ipp_name))
        my_data = c.fetchone()
        print(my_data)
        final_data = list(my_data)
        print(final_data)
        conn.commit()
        c.close()
        conn.close()
        final_data1 = json.dumps({'final_data': final_data})
        print(final_data1)


    return HttpResponse(final_data1, content_type="application/json")

#GENERATING A NODE LOSS REPORT
def sub_node_loss(request):
    if request.method == 'POST' and request.is_ajax():
        conn = pymysql.connect(host='database-1.cjvdavipre6m.us-west-2.rds.amazonaws.com', port=3306,
         user='root', password='uxiGJV2uwUWAS533CRDd', database='meteringdatabase')
        #conn = pymysql.connect(host='localhost', port=3306,
                       #user='root', password='', database='meteringdatabase')
        c = conn.cursor()
        #sub = json.loads(request.body)
        meter = request.POST.get("sel_meter")
        month = request.POST.get("month_loss")
        year = request.POST.get("year_loss")
        component = request.POST.get("loss_component")
        print(component)
        monthtt = f"{month}_{year}"
        print(monthtt)

        df = pd.read_sql_query(sql = f"select * FROM {component}_Loss", con=conn)
        df = df.drop(["timestamp"],axis=1)
        df_2 = df[df['month_year'] == monthtt]
        if df_2.dropna().empty:
            # GETTING VALUES
            finals = ["N/A"]*17
            final_data1 = json.dumps({'finals': finals})
            print(final_data1)
            return HttpResponse(final_data1, content_type="application/json")
        #df_3 = df_2[df_2['loss'].values != 0]
        #print(df_2)
        else:
            #GETTING VALUES
            av_loss_f = df_2['loss'].mean()
            av_loss = round(av_loss_f,2)
            min_loss = df_2['loss'].min()
            #min_loss_nonzero = df_3['loss'].min()
            max_loss = df_2['loss'].max()

            min_details = df_2.loc[df_2['loss'].idxmin()].tolist()
            #min_NZ_details = df_3.loc[df_3['loss'].idxmin()].tolist()
            max_details = df_2.loc[df_2['loss'].idxmax()].tolist()
            print(min_loss)
            print(min_details)
            print(max_loss)
            print(max_details)

            list1 = [av_loss,max_loss,min_loss]
            finals = list1 + max_details[1:] + min_details[1:]

            c.close()
            conn.close()
            final_data1 = json.dumps({'finals': finals})
            print(final_data1)
            return HttpResponse(final_data1, content_type="application/json")


#SUBMITTING CONFIGURATION FILES FOR ENERGY METERS ALREADY IN THE DATABASE

